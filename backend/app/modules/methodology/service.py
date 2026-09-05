# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Estimating-methodology service layer.

Stateless service over an :class:`~sqlalchemy.ext.asyncio.AsyncSession`,
following the same shape as the sibling modules:

* Project-scoped CRUD for methodologies, analytical dimensions (+ their values)
  and funding sources.
* Idempotent installation of a built-in template (see
  :mod:`app.modules.methodology.templates`) into a project as a project-scoped
  clone, optionally activating it.
* :meth:`MethodologyService.compute_estimate`, which resolves the cascade
  bases via :func:`app.modules.methodology.bases.resolve_bases`, runs the pure
  :func:`app.modules.methodology.cascade.compute_cascade`, and returns a
  currency-safe structured result.

Access control is the ROUTER's job (every project-touching endpoint calls
:func:`app.dependencies.verify_project_access` before reaching the service);
this layer assumes the caller has already been authorised for the project it
passes in, exactly like the BOQ / risk services.

Money is always :class:`~decimal.Decimal` and is never blended across
currencies: a methodology declares a single currency and the cascade engine
never converts. Where this service aggregates BOQ resource totals it relies on
the BOQ service, which already converts a mixed-currency BOQ into the project
base currency before returning category amounts.
"""

import io
import logging
import re
import uuid
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.methodology import templates as templates_mod
from app.modules.methodology.bases import resolve_bases
from app.modules.methodology.cascade import CascadeError, compute_cascade
from app.modules.methodology.models import (
    AnalyticDimension,
    AnalyticDimensionValue,
    FundingSource,
    Methodology,
)
from app.modules.methodology.schemas import (
    ComputeEstimateRequest,
    DimensionCreate,
    EffectiveVat,
    FundingSourceCreate,
    FundingSourceUpdate,
    MethodologyCreate,
    MethodologyUpdate,
)

logger = logging.getLogger(__name__)

# Active-methodology pointer is stored on the project's metadata blob under
# this key. Projects that never set it fall back to the international default
# (and the existing flat BOQMarkup path remains the platform-wide default for
# projects that never opt into the engine at all).
ACTIVE_METHODOLOGY_META_KEY = "methodology_slug"

# Where the project's per-position resource totals come from when the caller
# does not pass ``resource_totals`` explicitly to compute_estimate.
_SLUG_RE = re.compile(r"[^a-z0-9]+")


def _rate_to_decimal(raw: object) -> Decimal | None:
    """Read a percentage that is stored as a string, without inventing a number.

    Rates travel as decimal strings ("19", "8.1") through the cascade JSON and
    through the project column alike. A value that cannot be read is returned
    as ``None`` rather than as nought: a rate nobody could parse is unknown,
    and nought is a rate somebody chose.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _slugify(value: str) -> str:
    """Lower-kebab-case a free-form name into a slug fragment."""
    return _SLUG_RE.sub("-", value.strip().lower()).strip("-")


class MethodologyService:
    """Business logic for the estimating-methodology engine."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session

    # ── Methodology CRUD ───────────────────────────────────────────────────

    async def list_methodologies(self, project_id: uuid.UUID) -> list[Methodology]:
        """Return every methodology visible to a project.

        Visibility = the platform built-ins / pack templates (``project_id`` is
        NULL) PLUS the project's own clones. Ordered builtins-first, then by
        name, so the project picker shows the catalogue followed by local
        edits.
        """
        stmt = (
            select(Methodology)
            .where((Methodology.project_id.is_(None)) | (Methodology.project_id == str(project_id)))
            .order_by(Methodology.scope.desc(), Methodology.name.asc())
        )
        rows = (await self.session.execute(stmt)).scalars().all()
        return list(rows)

    async def get_methodology(self, methodology_id: uuid.UUID) -> Methodology:
        """Fetch a methodology by id. Raises 404 if absent."""
        obj = await self.session.get(Methodology, methodology_id)
        if obj is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Methodology not found",
            )
        return obj

    async def get_methodology_for_project(self, methodology_id: uuid.UUID, project_id: uuid.UUID) -> Methodology:
        """Fetch a methodology and assert it is visible to ``project_id``.

        A project may read its own clones and any global builtin / pack
        template. A clone owned by a DIFFERENT project is treated as missing
        (404), not forbidden (403), to avoid leaking the existence of other
        tenants' methodology ids (same IDOR policy as verify_project_access).
        """
        obj = await self.get_methodology(methodology_id)
        if obj.project_id is not None and obj.project_id != str(project_id):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Methodology not found",
            )
        return obj

    async def _slug_exists(self, slug: str) -> bool:
        stmt = select(Methodology.id).where(Methodology.slug == slug).limit(1)
        return (await self.session.execute(stmt)).first() is not None

    async def _unique_slug(self, base: str) -> str:
        """Return ``base`` or ``base-2``, ``base-3`` ... until unused.

        Methodology slugs are globally unique (the column is unique), so a
        project clone of a name another project already used gets a numeric
        suffix rather than colliding.
        """
        base = base or "methodology"
        if not await self._slug_exists(base):
            return base
        n = 2
        while True:
            candidate = f"{base}-{n}"
            if not await self._slug_exists(candidate):
                return candidate
            n += 1

    async def create_methodology(self, data: MethodologyCreate) -> Methodology:
        """Create a project-scoped methodology (``scope='project'``).

        The caller-supplied slug (or one derived from the name) is made
        globally unique. The new methodology is always editable and never a
        builtin.
        """
        desired_slug = _slugify(data.slug or data.name)
        slug = await self._unique_slug(desired_slug)

        obj = Methodology(
            slug=slug,
            scope="project",
            project_id=str(data.project_id),
            country_code=data.country_code,
            industry=data.industry,
            name=data.name,
            description=data.description,
            currency=data.currency or "",
            decimals=data.decimals,
            hierarchy_levels=list(data.hierarchy_levels),
            dimension_scheme=list(data.dimension_scheme),
            column_preset=data.column_preset,
            base_mapping=dict(data.base_mapping),
            composites=dict(data.composites),
            cascade_steps=[s.model_dump(mode="json") for s in data.cascade_steps],
            vat_rate=(format(data.vat_rate, "f") if data.vat_rate is not None else None),
            is_builtin=False,
            is_editable=True,
            metadata_=dict(data.metadata),
        )
        self.session.add(obj)
        await self.session.flush()
        logger.info("Methodology created: %s (project=%s)", obj.slug, data.project_id)
        return obj

    async def update_methodology(
        self, methodology_id: uuid.UUID, project_id: uuid.UUID, data: MethodologyUpdate
    ) -> Methodology:
        """Update an editable, project-owned methodology.

        Built-in / pack templates and clones owned by another project cannot be
        edited through here: built-ins are read-only platform data, and
        cross-project edits are blocked by the same visibility rule as reads.
        """
        obj = await self.get_methodology_for_project(methodology_id, project_id)
        if obj.is_builtin or not obj.is_editable or obj.project_id is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="This methodology is read-only. Clone it to make changes.",
            )

        fields = data.model_dump(exclude_unset=True)

        if "cascade_steps" in fields and data.cascade_steps is not None:
            obj.cascade_steps = [s.model_dump(mode="json") for s in data.cascade_steps]
            fields.pop("cascade_steps")
        if "vat_rate" in fields:
            obj.vat_rate = format(data.vat_rate, "f") if data.vat_rate is not None else None
            fields.pop("vat_rate")
        if "metadata" in fields and data.metadata is not None:
            # Merge, do not overwrite, so unrelated extension keys survive a
            # partial PATCH (the json-overwrite-on-PATCH class of bug).
            merged = dict(obj.metadata_ or {})
            merged.update(data.metadata)
            obj.metadata_ = merged
            fields.pop("metadata")

        # Remaining scalar / list fields map 1:1 onto the column names.
        for key, value in fields.items():
            setattr(obj, key, value)

        await self.session.flush()
        logger.info("Methodology updated: %s", obj.slug)
        return obj

    async def delete_methodology(self, methodology_id: uuid.UUID, project_id: uuid.UUID) -> None:
        """Delete an editable, project-owned methodology.

        Built-ins / pack templates and other projects' clones cannot be
        deleted. If the deleted methodology was the project's active one, the
        active pointer is cleared (the project falls back to the international
        default).
        """
        obj = await self.get_methodology_for_project(methodology_id, project_id)
        if obj.is_builtin or obj.project_id is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Built-in methodologies cannot be deleted.",
            )
        slug = obj.slug
        await self.session.delete(obj)
        await self.session.flush()
        # Clear the active pointer if it referenced this methodology.
        await self._clear_active_if_slug(project_id, slug)
        logger.info("Methodology deleted: %s (project=%s)", slug, project_id)

    # ── Built-in templates ─────────────────────────────────────────────────

    @staticmethod
    def list_templates() -> list[dict[str, Any]]:
        """Return the built-in template catalogue (pure data)."""
        return templates_mod.list_templates()

    async def install_template(
        self,
        *,
        project_id: uuid.UUID,
        template_slug: str,
        idempotent: bool = True,
        set_active: bool = False,
    ) -> Methodology:
        """Install a built-in template into a project as a project clone.

        Idempotent by default: if the project already has a clone of this
        template (tracked via ``metadata_.source_template``), that existing
        clone is returned untouched. Pass ``idempotent=False`` to force a
        fresh, numerically-suffixed clone.

        Args:
            project_id: Target project (already access-checked by the router).
            template_slug: Slug of a template in
                :data:`app.modules.methodology.templates.TEMPLATES_BY_SLUG`.
            idempotent: Reuse an existing clone of the same source template.
            set_active: Activate the installed methodology on the project.

        Raises:
            HTTPException 404: Unknown template slug.
        """
        try:
            tpl = templates_mod.get_template(template_slug)
        except templates_mod.TemplateError as exc:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Unknown template: {template_slug}",
            ) from exc

        if idempotent:
            existing = await self._find_installed_clone(project_id, template_slug)
            if existing is not None:
                if set_active:
                    await self.set_active_methodology(project_id, existing.slug)
                return existing

        slug = await self._unique_slug(f"{template_slug}-{_slugify(str(project_id))[:8]}")
        obj = Methodology(
            slug=slug,
            scope="project",
            project_id=str(project_id),
            country_code=tpl.get("country_code"),
            industry=tpl.get("industry"),
            name=tpl["name"],
            description=tpl.get("description"),
            currency=tpl.get("currency", "") or "",
            decimals=int(tpl.get("decimals", 2)),
            hierarchy_levels=list(tpl.get("hierarchy_levels", [])),
            # The methodology's own dimension scheme (definitions); concrete
            # dimension VALUE rows are materialised separately below.
            dimension_scheme=list(tpl.get("dimensions", [])),
            column_preset=tpl.get("column_preset"),
            base_mapping=dict(tpl.get("base_mapping", {})),
            composites=dict(tpl.get("composites", {})),
            cascade_steps=list(tpl.get("cascade_steps", [])),
            vat_rate=tpl.get("vat_rate"),
            is_builtin=False,
            is_editable=True,
            metadata_={"source_template": template_slug},
        )
        self.session.add(obj)
        await self.session.flush()

        # Materialise the template's analytical dimensions as real, editable
        # dimension + value rows scoped to the project + this methodology.
        await self._materialise_dimensions(
            project_id=project_id,
            methodology_slug=slug,
            dimension_specs=tpl.get("dimensions", []),
        )

        if set_active:
            await self.set_active_methodology(project_id, slug)

        logger.info(
            "Installed template %s into project %s as %s",
            template_slug,
            project_id,
            slug,
        )
        return obj

    async def _find_installed_clone(self, project_id: uuid.UUID, template_slug: str) -> Methodology | None:
        """Return the project's existing clone of a template, if any."""
        stmt = select(Methodology).where(
            Methodology.project_id == str(project_id),
            Methodology.scope == "project",
        )
        rows = (await self.session.execute(stmt)).scalars().all()
        for row in rows:
            if (row.metadata_ or {}).get("source_template") == template_slug:
                return row
        return None

    # ── Analytical dimensions ──────────────────────────────────────────────

    async def list_dimensions(
        self, project_id: uuid.UUID, *, methodology_slug: str | None = None
    ) -> list[AnalyticDimension]:
        """List a project's analytical dimensions (optionally one methodology)."""
        stmt = select(AnalyticDimension).where(AnalyticDimension.project_id == str(project_id))
        if methodology_slug is not None:
            stmt = stmt.where(AnalyticDimension.methodology_slug == methodology_slug)
        stmt = stmt.order_by(AnalyticDimension.sort_order.asc(), AnalyticDimension.label.asc())
        rows = (await self.session.execute(stmt)).scalars().all()
        return list(rows)

    async def create_dimension(self, data: DimensionCreate) -> AnalyticDimension:
        """Create an analytical dimension (with optional seed values)."""
        dim = AnalyticDimension(
            project_id=str(data.project_id),
            methodology_slug=data.methodology_slug,
            key=data.key,
            label=data.label,
            kind=data.kind,
            is_required=data.is_required,
            sort_order=data.sort_order,
            metadata_=dict(data.metadata),
        )
        self.session.add(dim)
        await self.session.flush()

        # Seed values. parent_code links resolve within this dimension only.
        code_to_value: dict[str, AnalyticDimensionValue] = {}
        for order, value in enumerate(data.values):
            val = AnalyticDimensionValue(
                dimension_id=dim.id,
                parent_id=None,
                code=value.code,
                label=value.label,
                sort_order=value.sort_order or order,
                metadata_=dict(value.metadata),
            )
            self.session.add(val)
            await self.session.flush()
            code_to_value[value.code] = val

        # Second pass: wire parents now that every value row has an id.
        for value in data.values:
            if value.parent_code and value.parent_code in code_to_value:
                child = code_to_value.get(value.code)
                parent = code_to_value.get(value.parent_code)
                if child is not None and parent is not None and child.id != parent.id:
                    child.parent_id = parent.id
        await self.session.flush()

        logger.info(
            "Dimension created: %s (project=%s, %d values)",
            dim.key,
            data.project_id,
            len(data.values),
        )
        return dim

    async def get_dimension_for_project(self, dimension_id: uuid.UUID, project_id: uuid.UUID) -> AnalyticDimension:
        """Fetch a dimension and assert it belongs to ``project_id`` (404 else)."""
        obj = await self.session.get(AnalyticDimension, dimension_id)
        if obj is None or obj.project_id != str(project_id):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Dimension not found",
            )
        return obj

    async def delete_dimension(self, dimension_id: uuid.UUID, project_id: uuid.UUID) -> None:
        """Delete a project-owned dimension (its values cascade)."""
        obj = await self.get_dimension_for_project(dimension_id, project_id)
        await self.session.delete(obj)
        await self.session.flush()
        logger.info("Dimension deleted: %s (project=%s)", obj.key, project_id)

    async def _materialise_dimensions(
        self,
        *,
        project_id: uuid.UUID,
        methodology_slug: str,
        dimension_specs: Any,
    ) -> None:
        """Create dimension + value rows from a template's dimension specs.

        Skips a dimension whose ``key`` already exists for this project +
        methodology, so re-installing (non-idempotent clone of a fresh slug)
        never duplicates and a hand-edited dimension is preserved.
        """
        for order, spec in enumerate(dimension_specs or []):
            if not isinstance(spec, dict):
                continue
            key = str(spec.get("key", "")).strip()
            if not key:
                continue
            values = [
                DimensionValueCreateLite(
                    code=str(v.get("code", "")),
                    label=str(v.get("label", "")),
                    parent_code=(str(v["parent_code"]) if v.get("parent_code") else None),
                )
                for v in spec.get("values", [])
                if isinstance(v, dict) and v.get("code")
            ]
            dim = AnalyticDimension(
                project_id=str(project_id),
                methodology_slug=methodology_slug,
                key=key,
                label=str(spec.get("label", key)),
                kind=str(spec.get("kind", "flat")),
                is_required=bool(spec.get("is_required", False)),
                sort_order=order,
                metadata_={},
            )
            self.session.add(dim)
            await self.session.flush()
            for v_order, v in enumerate(values):
                self.session.add(
                    AnalyticDimensionValue(
                        dimension_id=dim.id,
                        parent_id=None,
                        code=v.code,
                        label=v.label,
                        sort_order=v_order,
                        metadata_={},
                    )
                )
        await self.session.flush()

    # ── Funding sources ────────────────────────────────────────────────────

    async def list_funding_sources(self, project_id: uuid.UUID) -> list[FundingSource]:
        """List a project's funding-source master entries."""
        stmt = (
            select(FundingSource)
            .where(FundingSource.project_id == str(project_id))
            .order_by(FundingSource.sort_order.asc(), FundingSource.name.asc())
        )
        rows = (await self.session.execute(stmt)).scalars().all()
        return list(rows)

    async def create_funding_source(self, data: FundingSourceCreate) -> FundingSource:
        """Create a funding-source master entry for a project."""
        obj = FundingSource(
            project_id=str(data.project_id),
            code=data.code,
            name=data.name,
            sort_order=data.sort_order,
            metadata_=dict(data.metadata),
        )
        self.session.add(obj)
        await self.session.flush()
        logger.info("Funding source created: %s (project=%s)", obj.code, data.project_id)
        return obj

    async def get_funding_source_for_project(
        self, funding_source_id: uuid.UUID, project_id: uuid.UUID
    ) -> FundingSource:
        """Fetch a funding source and assert it belongs to ``project_id``."""
        obj = await self.session.get(FundingSource, funding_source_id)
        if obj is None or obj.project_id != str(project_id):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Funding source not found",
            )
        return obj

    async def update_funding_source(
        self,
        funding_source_id: uuid.UUID,
        project_id: uuid.UUID,
        data: FundingSourceUpdate,
    ) -> FundingSource:
        """Update a project-owned funding source."""
        obj = await self.get_funding_source_for_project(funding_source_id, project_id)
        fields = data.model_dump(exclude_unset=True)
        if "metadata" in fields and data.metadata is not None:
            merged = dict(obj.metadata_ or {})
            merged.update(data.metadata)
            obj.metadata_ = merged
            fields.pop("metadata")
        for key, value in fields.items():
            setattr(obj, key, value)
        await self.session.flush()
        return obj

    async def delete_funding_source(self, funding_source_id: uuid.UUID, project_id: uuid.UUID) -> None:
        """Delete a project-owned funding source."""
        obj = await self.get_funding_source_for_project(funding_source_id, project_id)
        await self.session.delete(obj)
        await self.session.flush()

    # ── Active methodology pointer (on project metadata) ────────────────────

    async def get_active_slug(self, project_id: uuid.UUID) -> str:
        """Return the project's active methodology slug, or the int'l default."""
        project = await self._get_project(project_id)
        if project is not None:
            meta = getattr(project, "metadata_", None) or {}
            slug = meta.get(ACTIVE_METHODOLOGY_META_KEY)
            if isinstance(slug, str) and slug:
                return slug
        return templates_mod.INTERNATIONAL_SLUG

    async def set_active_methodology(self, project_id: uuid.UUID, slug: str) -> str:
        """Set the project's active methodology slug on its metadata blob.

        The slug must reference a methodology visible to the project (a global
        builtin / pack, or one of the project's own clones) or a known built-in
        template slug. Returns the slug that was set.
        """
        # Validate the slug is resolvable for this project.
        if slug not in templates_mod.TEMPLATES_BY_SLUG:
            stmt = select(Methodology.id).where(
                Methodology.slug == slug,
                (Methodology.project_id.is_(None)) | (Methodology.project_id == str(project_id)),
            )
            if (await self.session.execute(stmt)).first() is None:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Methodology not found for this project",
                )

        project = await self._get_project(project_id)
        if project is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Project not found",
            )
        meta = dict(getattr(project, "metadata_", None) or {})
        meta[ACTIVE_METHODOLOGY_META_KEY] = slug
        project.metadata_ = meta
        await self.session.flush()
        logger.info("Project %s active methodology set to %s", project_id, slug)
        return slug

    async def _clear_active_if_slug(self, project_id: uuid.UUID, slug: str) -> None:
        """Clear the project's active pointer iff it equals ``slug``."""
        project = await self._get_project(project_id)
        if project is None:
            return
        meta = dict(getattr(project, "metadata_", None) or {})
        if meta.get(ACTIVE_METHODOLOGY_META_KEY) == slug:
            meta.pop(ACTIVE_METHODOLOGY_META_KEY, None)
            project.metadata_ = meta
            await self.session.flush()

    async def _get_project(self, project_id: uuid.UUID) -> Any | None:
        """Load the Project row (best-effort import to avoid a hard coupling)."""
        try:
            from app.modules.projects.models import Project

            return await self.session.get(Project, project_id)
        except Exception:
            logger.debug("Project lookup failed for %s", project_id, exc_info=True)
            return None

    # ── Compute estimate ───────────────────────────────────────────────────

    async def _resolve_methodology_config(self, project_id: uuid.UUID, slug: str) -> dict[str, Any]:
        """Resolve a methodology slug to a config dict for the cascade.

        Resolution order:
          1. A methodology row (project clone or global builtin) with this slug.
          2. A built-in template (pure data) with this slug.

        Returns a dict with currency / decimals / base_mapping / composites /
        cascade_steps. Raises 404 if the slug resolves to neither.
        """
        stmt = select(Methodology).where(
            Methodology.slug == slug,
            (Methodology.project_id.is_(None)) | (Methodology.project_id == str(project_id)),
        )
        row = (await self.session.execute(stmt)).scalars().first()
        if row is not None:
            return {
                "slug": row.slug,
                "currency": row.currency or "",
                "decimals": int(row.decimals),
                "base_mapping": dict(row.base_mapping or {}),
                "composites": dict(row.composites or {}),
                "cascade_steps": list(row.cascade_steps or []),
            }

        if slug in templates_mod.TEMPLATES_BY_SLUG:
            tpl = templates_mod.TEMPLATES_BY_SLUG[slug]
            return {
                "slug": tpl["slug"],
                "currency": tpl.get("currency", "") or "",
                "decimals": int(tpl.get("decimals", 2)),
                "base_mapping": dict(tpl.get("base_mapping", {})),
                "composites": dict(tpl.get("composites", {})),
                "cascade_steps": list(tpl.get("cascade_steps", [])),
            }

        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Methodology '{slug}' not found for this project",
        )

    async def _project_vat_rate(self, project_id: uuid.UUID) -> str | None:
        """The project's own consumption-tax rate, or ``None`` if it set none.

        ``default_vat_rate`` is a decimal-string percentage (e.g. ``"21"``),
        the same column :meth:`BOQService.apply_default_markups` reads when it
        seeds a bill. Best-effort by design: a project that cannot be loaded
        must leave the catalogue rate standing rather than fail a computation,
        because the template rate is a correct answer and an error is not.
        """
        try:
            project = await self._get_project(project_id)
        except Exception:  # noqa: BLE001 - best-effort, never break a computation
            logger.debug("default_vat_rate lookup failed for project %s", project_id, exc_info=True)
            return None
        raw = getattr(project, "default_vat_rate", None) if project is not None else None
        if raw is None or str(raw).strip() == "":
            return None
        return str(raw).strip()

    @staticmethod
    def _with_project_vat(cascade_steps: list[Any], vat_rate: str) -> list[Any]:
        """Return the steps with the single tax line moved to ``vat_rate``.

        Keyed on the steps carrying exactly one ``tax`` category, not on the
        methodology's region. A region is the wrong handle here: a clone the
        user has edited has no region to consult, and the property that decides
        whether one rate is a complete swap is a property of the steps alone.
        Where a stack carries no tax line, or more than one, a single project
        rate cannot stand in for it and the steps are returned untouched. That
        is the same rule the bill side applies through
        :data:`~app.modules.boq.markup_templates.NON_SINGLE_TAX_REGIONS`, and
        ``test_a_single_tax_step_is_exactly_a_single_tax_line`` holds the two
        statements of it together.

        The step dicts are shared with the module-level template catalogue, so
        the one being changed is copied. Mutating it in place would re-rate
        every later computation in the process, including other projects'.
        """
        index = MethodologyService._single_tax_index(cascade_steps)
        if index is None:
            return cascade_steps

        patched = list(cascade_steps)
        patched[index] = {**patched[index], "rate": vat_rate}
        return patched

    @staticmethod
    def _single_tax_index(cascade_steps: list[Any]) -> int | None:
        """Position of the one ``tax`` step, or ``None`` when there is not exactly one.

        Both the override and the effective-rate signal have to answer "can one
        project figure stand in for this stack", and they have to answer it the
        same way every time. Two conditions that agree today are one condition
        written twice, so the condition lives here and is read, never restated.
        """
        tax_positions = [
            index
            for index, step in enumerate(cascade_steps)
            if isinstance(step, dict) and str(step.get("category", "")).strip().lower() == "tax"
        ]
        return tax_positions[0] if len(tax_positions) == 1 else None

    async def effective_vat(self, methodology: Methodology, project_id: uuid.UUID) -> EffectiveVat:
        """Which VAT rate this methodology is priced at for this project, and why.

        The read side of the substitution :meth:`compute_estimate` performs. It
        exists because the two used to disagree in silence: the computation and
        both exports have honoured the project's rate since ``e7dbcded9``,
        while a read returned the stored steps, so an editor showing a tax line
        at nought belonged to a bill costed at the project's rate with nothing
        on screen saying so. A hand-made methodology is the sharp case, since
        the create form seeds exactly one tax step at nought and that is
        precisely the shape the override replaces.

        This resolves rather than decides: it applies the same rule as the
        override, through the same helper, and does not itself change a price.
        """
        steps = list(methodology.cascade_steps or [])
        index = self._single_tax_index(steps)
        stored = _rate_to_decimal(steps[index].get("rate")) if index is not None else None

        if index is None:
            # No single tax line, so no one rate describes this stack. Say that
            # plainly rather than reporting a nought somebody would read as
            # zero-rated.
            return EffectiveVat(rate=None, source="none", stored_rate=None, differs_from_stored=False)

        project_rate = _rate_to_decimal(await self._project_vat_rate(project_id))
        if project_rate is None:
            return EffectiveVat(rate=stored, source="methodology", stored_rate=stored, differs_from_stored=False)

        return EffectiveVat(
            rate=project_rate,
            source="project",
            stored_rate=stored,
            # A project on the same rate as the template is not an override
            # worth announcing, which is why this compares figures instead of
            # simply reporting that a project rate exists.
            differs_from_stored=stored is None or project_rate != stored,
        )

    async def _aggregate_boq_resource_totals(self, boq_id: uuid.UUID) -> dict[str, Decimal]:
        """Sum a BOQ's resource costs per resource type via the BOQ service.

        Returns a ``{resource_type: Decimal}`` map (labor / material /
        equipment / subcontractor / other). Best-effort: a failure to load the
        BOQ surfaces as an empty map, which the caller turns into an all-zero
        cascade rather than a 500.
        """
        try:
            from app.modules.boq.service import BOQService

            breakdown = await BOQService(self.session).get_cost_breakdown(boq_id)
        except HTTPException:
            raise
        except Exception:
            logger.debug("BOQ resource aggregation failed for %s", boq_id, exc_info=True)
            return {}

        totals: dict[str, Decimal] = {}
        for cat in getattr(breakdown, "categories", []) or []:
            res_type = getattr(cat, "type", None)
            amount = getattr(cat, "amount", None)
            if not res_type:
                continue
            if not isinstance(amount, Decimal):
                try:
                    amount = Decimal(str(amount))
                except Exception:
                    continue
            totals[str(res_type)] = totals.get(str(res_type), Decimal(0)) + amount
        return totals

    async def compute_estimate(self, data: ComputeEstimateRequest) -> dict[str, Any]:
        """Compute the markup cascade for a project under a methodology.

        Resource totals are taken, in priority order, from
        ``data.resource_totals`` (caller-supplied), else aggregated from
        ``data.boq_id`` via the BOQ cost breakdown. The methodology is
        ``data.methodology_slug`` (a what-if override) or the project's active
        methodology (or the international default).

        Returns a plain dict matching
        :class:`app.modules.methodology.schemas.ComputeEstimateResponse`.

        VAT comes from the project, not from the template, whenever the project
        states one. A template's rate is a catalogue fact about a country and it
        stays the default; but this method is given a ``project_id``, and
        through :meth:`build_export_data` a ``boq_id``, so at that point it is
        pricing one project's bill and has to answer what the bill answers. The
        bill engine has honoured ``default_vat_rate`` since issue #89, and one
        bill asked two ways cannot cost two amounts. A project that set no rate
        of its own is not moved by a cent, and the pure catalogue path,
        :func:`~app.modules.methodology.templates.build_cascade_spec_from_template`,
        does not consult a project at all and is unaffected.

        Raises:
            HTTPException 404: Methodology slug not resolvable for the project.
            HTTPException 422: The methodology's cascade spec is invalid.
        """
        slug = data.methodology_slug or await self.get_active_slug(data.project_id)
        config = await self._resolve_methodology_config(data.project_id, slug)

        project_vat = await self._project_vat_rate(data.project_id)
        if project_vat is not None:
            config["cascade_steps"] = self._with_project_vat(config["cascade_steps"], project_vat)

        # Source the per-resource-type totals.
        if data.resource_totals is not None:
            resource_totals: dict[str, Decimal] = {
                str(k): (v if isinstance(v, Decimal) else Decimal(str(v))) for k, v in data.resource_totals.items()
            }
        elif data.boq_id is not None:
            resource_totals = await self._aggregate_boq_resource_totals(data.boq_id)
        else:
            resource_totals = {}

        # Resolve the cascade leaf bases from the methodology's base mapping.
        # When the methodology declares no mapping (a bare flat template), fall
        # back to a single "direct" base equal to the sum of all resource
        # totals so the cascade still has something to apply to.
        base_mapping = config["base_mapping"]
        if base_mapping:
            bases = resolve_bases(base_mapping, resource_totals)
        else:
            total = sum(resource_totals.values(), Decimal(0))
            bases = {"direct": total}

        spec = templates_mod.build_cascade_spec(
            slug=config["slug"],
            currency=config["currency"],
            decimals=config["decimals"],
            composites=config["composites"],
            cascade_steps=config["cascade_steps"],
        )

        try:
            result = compute_cascade(spec, bases)
        except CascadeError as exc:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Invalid methodology cascade: {exc}",
            ) from exc

        return {
            "project_id": data.project_id,
            "methodology_slug": config["slug"],
            "currency": config["currency"],
            "decimals": config["decimals"],
            "bases": dict(result.bases),
            "composites": dict(result.composites),
            "steps": [
                {
                    "key": s.key,
                    "label": s.label,
                    "category": s.category,
                    "kind": s.kind,
                    "rate": s.rate,
                    "base_amount": s.base_amount,
                    "amount": s.amount,
                    "running_total": s.running_total,
                }
                for s in result.steps
            ],
            "direct_total": result.direct_total,
            "markup_total": result.markup_total,
            "grand_total": result.grand_total,
        }

    # ── Export (Excel / PDF) ────────────────────────────────────────────────

    async def build_export_data(
        self,
        methodology_id: uuid.UUID,
        project_id: uuid.UUID,
        *,
        boq_id: uuid.UUID | None = None,
        resource_totals: dict[str, Decimal] | None = None,
    ) -> dict[str, Any]:
        """Build the data dict that both exporters render.

        Resolves the methodology (scoped to the project for IDOR safety - a
        clone owned by another project 404s exactly like a read), computes its
        cascade against the supplied / aggregated resource totals, and enriches
        the compute result with the human-readable methodology name and the
        project name + preparer for the document cover.

        The resource totals follow the same priority as
        :meth:`compute_estimate`: explicit ``resource_totals`` first, else the
        BOQ aggregation for ``boq_id``, else an all-zero cascade.
        """
        methodology = await self.get_methodology_for_project(methodology_id, project_id)

        compute = await self.compute_estimate(
            ComputeEstimateRequest(
                project_id=project_id,
                methodology_slug=methodology.slug,
                boq_id=boq_id,
                resource_totals=resource_totals,
            )
        )

        project = await self._get_project(project_id)
        project_name = str(getattr(project, "name", "") or "") if project else ""
        prepared_by = await self._resolve_prepared_by(project)

        return {
            "methodology_id": str(methodology_id),
            "methodology_name": methodology.name,
            "methodology_slug": compute["methodology_slug"],
            "project_id": str(project_id),
            "project_name": project_name,
            "prepared_by": prepared_by,
            "currency": compute["currency"],
            "decimals": compute["decimals"],
            "bases": compute["bases"],
            "composites": compute["composites"],
            "steps": compute["steps"],
            "direct_total": compute["direct_total"],
            "markup_total": compute["markup_total"],
            "grand_total": compute["grand_total"],
        }

    async def _resolve_prepared_by(self, project: Any | None) -> str:
        """Best-effort 'Prepared by' name from the project owner (never raises)."""
        if project is None:
            return ""
        owner_id = getattr(project, "owner_id", None)
        if owner_id is None:
            return ""
        try:
            from app.modules.users.models import User

            owner = await self.session.get(User, owner_id)
        except Exception:
            logger.debug("Owner lookup failed for export cover", exc_info=True)
            return ""
        if owner is None:
            return ""
        return str(getattr(owner, "full_name", "") or getattr(owner, "email", "") or "")

    @staticmethod
    def export_filename(data: dict[str, Any], ext: str) -> str:
        """Download filename derived from the methodology name.

        Returns the real (possibly non-ASCII) name; the router wraps it in
        :func:`app.core.content_disposition.attachment_disposition`, which
        derives the RFC 6266 ASCII fallback and UTF-8 ``filename*`` pair.
        """
        raw = str(data.get("methodology_name") or "methodology")
        safe = raw.strip() or "methodology"
        return f"{safe}.{ext}"

    def generate_excel_export(self, data: dict[str, Any]) -> bytes:
        """Render the export data as a formatted .xlsx workbook (bytes).

        One sheet: the methodology identity block, the markup cascade (direct
        cost -> each step -> grand total) with real numeric cells so Excel can
        SUM them, and a resolved-bases appendix. Every user-controlled string
        is routed through :func:`neutralise_formula` to defend against CSV /
        formula injection, mirroring the BOQ exporter.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        from app.core.csv_safety import neutralise_formula

        currency = data.get("currency", "") or ""
        decimals = int(data.get("decimals", 2))
        number_format = f"#,##0.{'0' * decimals}" if decimals > 0 else "#,##0"

        def _num(value: Any) -> Decimal:
            if value is None or value == "":
                return Decimal(0)
            try:
                d = value if isinstance(value, Decimal) else Decimal(str(value).strip())
            except (InvalidOperation, ValueError):
                return Decimal(0)
            return d if d.is_finite() else Decimal(0)

        wb = Workbook()
        ws = wb.active
        ws.title = "Estimate"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color="FFFFFF")
        grand_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        direct_fill = PatternFill(start_color="F0F0F5", end_color="F0F0F5", fill_type="solid")
        grand_fill = PatternFill(start_color="E8E8EE", end_color="E8E8EE", fill_type="solid")
        right = Alignment(horizontal="right")
        top_border = Border(top=Side(style="medium"))

        # ── Identity block ───────────────────────────────────────────────
        ws.cell(row=1, column=1, value="Cost Estimate").font = title_font
        identity = [
            ("Project", data.get("project_name", "")),
            ("Methodology", data.get("methodology_name", "")),
            ("Method ID", data.get("methodology_slug", "")),
            ("Currency", currency),
        ]
        row = 2
        for label, value in identity:
            ws.cell(row=row, column=1, value=label).font = bold
            ws.cell(row=row, column=2, value=neutralise_formula(str(value or "")))
            row += 1

        row += 1  # spacer

        # ── Cascade header ───────────────────────────────────────────────
        headers = ["Step", "Category", "Rate %", "Base", "Amount", "Running total"]
        header_row = row
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        row += 1

        # Direct-cost opening row.
        ws.cell(row=row, column=1, value="Direct cost").font = bold
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = direct_fill
        dc = ws.cell(row=row, column=5, value=_num(data.get("direct_total")))
        dc.number_format = number_format
        dc.font = bold
        dc.alignment = right
        rt = ws.cell(row=row, column=6, value=_num(data.get("direct_total")))
        rt.number_format = number_format
        rt.alignment = right
        row += 1

        for step in data.get("steps", []) or []:
            kind = step.get("kind", "percentage")
            ws.cell(
                row=row,
                column=1,
                value=neutralise_formula(str(step.get("label") or step.get("key", ""))),
            )
            ws.cell(row=row, column=2, value=neutralise_formula(str(step.get("category", ""))))
            if kind == "percentage":
                rate_cell = ws.cell(row=row, column=3, value=_num(step.get("rate")))
                rate_cell.number_format = number_format
                rate_cell.alignment = right
            else:
                ws.cell(row=row, column=3, value="-").alignment = right
            base_cell = ws.cell(row=row, column=4, value=_num(step.get("base_amount")))
            base_cell.number_format = number_format
            base_cell.alignment = right
            amt_cell = ws.cell(row=row, column=5, value=_num(step.get("amount")))
            amt_cell.number_format = number_format
            amt_cell.alignment = right
            run_cell = ws.cell(row=row, column=6, value=_num(step.get("running_total")))
            run_cell.number_format = number_format
            run_cell.alignment = right
            row += 1

        # Grand-total row.
        grand_row = row
        for col in range(1, len(headers) + 1):
            ws.cell(row=grand_row, column=col).border = top_border
            ws.cell(row=grand_row, column=col).fill = grand_fill
        gl = ws.cell(row=grand_row, column=1, value="Grand total")
        gl.font = grand_font
        gt = ws.cell(row=grand_row, column=6, value=_num(data.get("grand_total")))
        gt.font = grand_font
        gt.number_format = number_format
        gt.alignment = right
        # Total markup, stated next to the grand total for an at-a-glance read.
        mk = ws.cell(row=grand_row, column=5, value=_num(data.get("markup_total")))
        mk.number_format = number_format
        mk.alignment = right
        row += 2

        # ── Resolved-bases appendix ──────────────────────────────────────
        bases = data.get("bases", {}) or {}
        composites = data.get("composites", {}) or {}
        if bases or composites:
            ws.cell(row=row, column=1, value="Resolved bases").font = bold
            row += 1
            for col, header in enumerate(("Key", "Type", "Amount"), start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            for key, amount in bases.items():
                ws.cell(row=row, column=1, value=neutralise_formula(str(key)))
                ws.cell(row=row, column=2, value="base")
                ac = ws.cell(row=row, column=3, value=_num(amount))
                ac.number_format = number_format
                ac.alignment = right
                row += 1
            for key, amount in composites.items():
                ws.cell(row=row, column=1, value=neutralise_formula(str(key)))
                ws.cell(row=row, column=2, value="composite")
                ac = ws.cell(row=row, column=3, value=_num(amount))
                ac.number_format = number_format
                ac.alignment = right
                row += 1

        # Auto-width the populated columns.
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for r in ws.iter_rows(min_row=1, max_row=row, min_col=col_idx, max_col=col_idx):
                for cell in r:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

        # Workbook origin metadata (best-effort; touches no data cell).
        try:
            wb.properties.creator = "OpenConstructionERP · DataDrivenConstruction"
            wb.properties.lastModifiedBy = "OpenConstructionERP"
            wb.properties.title = f"Cost Estimate - {data.get('methodology_name', '')}"
            wb.properties.description = "Generated by OpenConstructionERP (https://openconstructionerp.com)"
        except Exception:  # noqa: BLE001 - metadata stamp must never break export
            logger.debug("Workbook metadata stamp failed", exc_info=True)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def generate_pdf_export(self, data: dict[str, Any]) -> bytes:
        """Render the export data as a professional PDF (bytes)."""
        from app.modules.methodology.pdf_export import generate_methodology_pdf

        return generate_methodology_pdf(data)


class DimensionValueCreateLite:
    """Tiny internal struct for template-driven value seeding.

    Avoids constructing a full Pydantic ``DimensionValueCreate`` (which
    requires non-empty code/label) when materialising trusted template data,
    where blank rows are already filtered out by the caller.
    """

    __slots__ = ("code", "label", "parent_code")

    def __init__(self, code: str, label: str, parent_code: str | None) -> None:
        self.code = code
        self.label = label
        self.parent_code = parent_code
