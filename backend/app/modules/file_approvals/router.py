# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""File Approvals (W8) API routes.

Mounted by the module loader at ``/api/v1/file-approvals``.

Endpoints
~~~~~~~~~
* ``GET    /``                            - list workflows
* ``POST   /``                            - submit a file for approval
* ``GET    /{id}``                         - workflow detail
* ``POST   /{id}/steps/{step_id}/decide/`` - record per-step decision
* ``POST   /{id}/withdraw/``               - submitter withdraws
* ``GET    /{id}/stamped/``                - stamped artifact bytes
* ``GET    /stamp-templates/``             - list templates (global + project)
* ``POST   /stamp-templates/``             - create custom template
"""

from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime

from fastapi import APIRouter, Depends, Query, status
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.csv_safety import neutralise_formula
from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.file_approvals.models import FileApprovalWorkflow
from app.modules.file_approvals.schemas import (
    ApprovalStepDecide,
    ApprovalWorkflowCreate,
    ApprovalWorkflowResponse,
    StampTemplateCreate,
    StampTemplateResponse,
)
from app.modules.file_approvals.service import ApprovalService

logger = logging.getLogger(__name__)

router = APIRouter(tags=["File Approvals"])


def _get_service(session: SessionDep) -> ApprovalService:
    return ApprovalService(session)


async def _require_project_access(session: AsyncSession, project_id: uuid.UUID, user_id: str) -> None:
    """Verify the caller may access ``project_id`` (owner, admin, or team member).

    Delegates to the app-wide canonical policy
    ``app.dependencies.verify_project_access`` so legitimate project TEAM
    MEMBERS are no longer wrongly denied. The previous owner-only gate was
    over-strict and 403'd valid non-owner approvers on every workflow
    endpoint, breaking the multi-approver premise.
    """
    # RBAC/broken-access-control fix: relax the owner-only gate to the
    # canonical owner/admin/team-member policy. verify_project_access raises
    # HTTPException(404) on denial, preserving the 404-not-403 IDOR behaviour.
    from app.dependencies import verify_project_access

    await verify_project_access(project_id, user_id, session)


# ── Register export ────────────────────────────────────────────────────────
#
# Excel export of the approvals register, cloned from the proven RFI-log
# exporter (``app.modules.rfi.router.export_rfi_log``). One row per approval
# workflow; a single endpoint yields a compliance-ready audit artifact. The
# workbook build is a pure, module-level function so it can be unit-tested
# without a database.

# Column order for the register sheet - kept as a constant so the test and the
# builder agree on positions.
_EXPORT_HEADERS: tuple[str, ...] = (
    "File Kind",
    "File",
    "Version",
    "Submitted By",
    "Submitted At",
    "Status",
    "Current Step / Approver",
    "Steps / Decisions",
    "Final Decision At",
    "Final Decision By",
    "Notes",
)


def _fmt_dt(value: datetime | None) -> str:
    """ISO-format a datetime for a cell (empty string when null).

    We stringify rather than hand openpyxl the raw value because the
    workflow timestamps are timezone-aware (``DateTime(timezone=True)``) and
    openpyxl refuses tz-aware datetimes. An ISO string starts with a digit,
    so it is never mistaken for a spreadsheet formula.
    """
    return value.isoformat() if value is not None else ""


def _step_summary(steps: list[object]) -> str:
    """Compact one-cell digest of every step: approver/role, decision, note."""
    parts: list[str] = []
    for s in steps:
        approver = getattr(s, "approver_id", None)
        who = getattr(s, "role_label", None) or (str(approver)[:8] if approver else "?")
        decision = getattr(s, "decision", "") or ""
        note = getattr(s, "decision_note", None)
        chunk = f"{who}: {decision}"
        if note:
            chunk += f" ({note})"
        parts.append(chunk)
    return "; ".join(parts)


def build_approvals_workbook(workflows: list[FileApprovalWorkflow]) -> Workbook:
    """Build the approvals-register ``.xlsx`` for a set of workflows.

    Extracted from the export endpoint so it can be exercised DB-free: pass
    in-memory :class:`FileApprovalWorkflow` objects (with ``steps``) and
    inspect the produced cells.

    Every user-controlled string cell is routed through
    :func:`neutralise_formula` (OWASP CSV-injection defence), mirroring the
    RFI-log exporter: ``file_id``, ``file_version_snapshot``, ``notes`` and
    the free-text step digest (which carries each step's ``role_label`` and
    ``decision_note``). Server-derived enums (``file_kind``, ``status``,
    ``decision``) and UUID identifiers pass through unchanged.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Approval Register"

    for col, header in enumerate(_EXPORT_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)

    for row_idx, wf in enumerate(workflows, 2):
        steps = sorted(getattr(wf, "steps", None) or [], key=lambda s: getattr(s, "sort_order", 0))
        # "Current step / approver": the first still-pending step - the same
        # actionable step the drawer highlights. Derived here (like the RFI
        # exporter derives days_open) rather than stored. Terminal workflows
        # have no pending step.
        current = next((s for s in steps if getattr(s, "decision", None) == "pending"), None)
        if current is not None:
            approver = getattr(current, "approver_id", None)
            who = getattr(current, "role_label", None) or (str(approver) if approver else "")
            current_cell = f"#{getattr(current, 'sort_order', 0) + 1}: {who}"
        else:
            # Terminal workflow (approved / rejected / withdrawn): no step is
            # pending. Leave the cell blank - the Status column already carries
            # the outcome, and a literal "-" would be apostrophe-escaped by the
            # CSV-injection guard (a leading "-" is a formula trigger).
            current_cell = ""

        ws.cell(row=row_idx, column=1, value=wf.file_kind)
        ws.cell(row=row_idx, column=2, value=neutralise_formula(wf.file_id))
        ws.cell(row=row_idx, column=3, value=neutralise_formula(wf.file_version_snapshot or ""))
        ws.cell(row=row_idx, column=4, value=str(wf.submitted_by_id) if wf.submitted_by_id else "")
        ws.cell(row=row_idx, column=5, value=_fmt_dt(wf.submitted_at))
        ws.cell(row=row_idx, column=6, value=wf.status)
        ws.cell(row=row_idx, column=7, value=neutralise_formula(current_cell))
        ws.cell(row=row_idx, column=8, value=neutralise_formula(_step_summary(steps)))
        ws.cell(row=row_idx, column=9, value=_fmt_dt(wf.final_decision_at))
        ws.cell(row=row_idx, column=10, value=str(wf.final_decision_by_id) if wf.final_decision_by_id else "")
        ws.cell(row=row_idx, column=11, value=neutralise_formula(wf.notes or ""))

    return wb


# ── Stamp templates ───────────────────────────────────────────────────────


@router.get(
    "/stamp-templates/",
    response_model=list[StampTemplateResponse],
    dependencies=[Depends(RequirePermission("file_approvals.read"))],
)
async def list_stamp_templates(
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
    project_id: uuid.UUID | None = Query(default=None),
) -> list[StampTemplateResponse]:
    """List active stamp templates (globals + ``project_id`` scope)."""
    if project_id is not None:
        await _require_project_access(session, project_id, user_id)
    rows = await service.list_templates(project_id)
    return [StampTemplateResponse.model_validate(r) for r in rows]


@router.post(
    "/stamp-templates/",
    response_model=StampTemplateResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("file_approvals.manage_stamps"))],
)
async def create_stamp_template(
    data: StampTemplateCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> StampTemplateResponse:
    """Create a stamp template (global if ``project_id`` is null)."""
    if data.project_id is not None:
        await _require_project_access(session, data.project_id, user_id)
    row = await service.create_template(data)
    return StampTemplateResponse.model_validate(row)


# ── Workflows ─────────────────────────────────────────────────────────────


@router.get(
    "/",
    response_model=list[ApprovalWorkflowResponse],
    dependencies=[Depends(RequirePermission("file_approvals.read"))],
)
async def list_workflows(
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
    project_id: uuid.UUID = Query(...),
    status_filter: str | None = Query(default=None, alias="status"),
) -> list[ApprovalWorkflowResponse]:
    """List workflows for a project, newest first."""
    await _require_project_access(session, project_id, user_id)
    rows = await service.list_workflows(project_id, status_filter=status_filter)
    return [ApprovalWorkflowResponse.model_validate(r) for r in rows]


@router.post(
    "/",
    response_model=ApprovalWorkflowResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("file_approvals.submit"))],
)
async def submit_for_approval(
    data: ApprovalWorkflowCreate,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> ApprovalWorkflowResponse:
    """Submit a file for approval - creates the workflow + steps."""
    await _require_project_access(session, data.project_id, user_id)
    workflow = await service.submit(data, submitted_by_id=user_id)
    return ApprovalWorkflowResponse.model_validate(workflow)


# NOTE: this static ``/export/`` route MUST stay above the parametric
# ``/{workflow_id}/`` route below. With ``redirect_slashes=False`` and a
# ``uuid.UUID`` path type, a request to ``/export/`` would otherwise be
# matched against ``/{workflow_id}/`` and 422 on the UUID parse instead of
# reaching this handler.
@router.get(
    "/export/",
    dependencies=[Depends(RequirePermission("file_approvals.read"))],
)
async def export_approvals_register(
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
    project_id: uuid.UUID = Query(...),
) -> StreamingResponse:
    """Export the approvals register for a project as an Excel workbook.

    One row per approval workflow (file, submitter, current approver, status,
    dates, per-step decisions, notes). Mirrors the RFI-log exporter's
    permission (``file_approvals.read``), project-scope guard, CSV-injection
    neutralisation and streaming response.
    """
    await _require_project_access(session, project_id, user_id)
    workflows = await service.list_workflows(project_id)
    wb = build_approvals_workbook(workflows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="approval_register.xlsx"'},
    )


@router.get(
    "/{workflow_id}/",
    response_model=ApprovalWorkflowResponse,
    dependencies=[Depends(RequirePermission("file_approvals.read"))],
)
async def get_workflow(
    workflow_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> ApprovalWorkflowResponse:
    """Load a single workflow with steps."""
    workflow = await service.get_workflow(workflow_id)
    await _require_project_access(session, workflow.project_id, user_id)
    return ApprovalWorkflowResponse.model_validate(workflow)


@router.post(
    "/{workflow_id}/steps/{step_id}/decide/",
    response_model=ApprovalWorkflowResponse,
    dependencies=[Depends(RequirePermission("file_approvals.decide"))],
)
async def decide_step(
    workflow_id: uuid.UUID,
    step_id: uuid.UUID,
    data: ApprovalStepDecide,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> ApprovalWorkflowResponse:
    """Record a decision on one approval step.

    On the final approval, the configured stamp template is burned into
    a copy of the file (PDF overlay or JSON sidecar) and the stamped
    path is stored on the workflow.
    """
    workflow = await service.get_workflow(workflow_id)
    await _require_project_access(session, workflow.project_id, user_id)
    workflow = await service.decide(workflow_id, step_id, data, user_id)
    return ApprovalWorkflowResponse.model_validate(workflow)


@router.post(
    "/{workflow_id}/withdraw/",
    response_model=ApprovalWorkflowResponse,
    dependencies=[Depends(RequirePermission("file_approvals.submit"))],
)
async def withdraw_workflow(
    workflow_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> ApprovalWorkflowResponse:
    """Submitter (or admin) withdraws a still-in-review workflow."""
    workflow = await service.get_workflow(workflow_id)
    await _require_project_access(session, workflow.project_id, user_id)
    workflow = await service.withdraw(workflow_id)
    return ApprovalWorkflowResponse.model_validate(workflow)


@router.get(
    "/{workflow_id}/stamped/",
    dependencies=[Depends(RequirePermission("file_approvals.read"))],
)
async def download_stamped(
    workflow_id: uuid.UUID,
    user_id: CurrentUserId,
    session: SessionDep,
    service: ApprovalService = Depends(_get_service),
) -> Response:
    """Return the stamped artifact bytes."""
    workflow = await service.get_workflow(workflow_id)
    await _require_project_access(session, workflow.project_id, user_id)
    data, media_type = await service.read_stamped(workflow_id)
    ext = "pdf" if media_type == "application/pdf" else "json"
    filename = f"approval_{workflow.id}.{ext}"
    return Response(
        content=data,
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
