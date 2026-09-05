# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""BIM Hub API routes.

Endpoint convention
-------------------
**Canonical** path for any per-model endpoint is ``/models/{model_id}/...``
to match the spatial intent (``models`` is a collection of resources). The
older flat ``/{model_id}/...`` paths are kept as back-compat aliases so
existing SDK callers don't break - both paths resolve to the same handler.
New endpoints SHOULD use the ``/models/{model_id}/...`` form.

Endpoints:
    Models:
        GET    /                                - List models for a project
        POST   /                                - Create model
        POST   /upload                          - Upload BIM data (DataFrame + optional DAE)
        GET    /models/{model_id}                - Get single model (canonical)
        GET    /{model_id}                       - Get single model (alias)
        PATCH  /models/{model_id}                - Update model (canonical)
        PATCH  /{model_id}                       - Update model (alias)
        DELETE /models/{model_id}                - Delete model (canonical)
        DELETE /{model_id}                       - Delete model (alias)
        GET    /models/{model_id}/geometry       - Serve DAE geometry file

    Elements:
        GET    /models/{model_id}/elements      - List elements (paginated, filterable)
        POST   /models/{model_id}/elements      - Bulk import elements
        GET    /{model_id}/elements             - List elements (alias)
        GET    /elements/{element_id}            - Get single element

    BOQ Links:
        GET    /links                            - List links for a BOQ position
        POST   /links                            - Create link
        DELETE /links/{link_id}                  - Delete link

    Quantity Maps:
        GET    /quantity-maps                    - List quantity map rules
        POST   /quantity-maps                    - Create quantity map rule
        PATCH  /quantity-maps/{map_id}           - Update quantity map rule
        POST   /quantity-maps/apply              - Apply rules on model

    Diffs:
        POST   /models/{model_id}/diff/{old_id}  - Compute diff
        GET    /diffs/{diff_id}                   - Get diff

    Element Groups (saved selections):
        GET    /element-groups/                   - List groups for a project
        POST   /element-groups/                   - Create a group
        PATCH  /element-groups/{group_id}         - Update a group
        DELETE /element-groups/{group_id}         - Delete a group

    Dataframe (Parquet + DuckDB analytical queries):
        GET    /models/{model_id}/dataframe/schema/              - Column names + types
        POST   /models/{model_id}/dataframe/query/               - Query via DuckDB SQL
        GET    /models/{model_id}/dataframe/columns/{col}/values - Value counts for a column
"""

import contextlib
import csv
import gzip as _gzip
import io
import json
import logging
import pathlib
import uuid
from typing import Any

from fastapi import APIRouter, BackgroundTasks, Depends, File, Header, HTTPException, Query, UploadFile, status
from fastapi.responses import JSONResponse, RedirectResponse, Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.content_disposition import attachment_disposition
from app.core.demo_placeholders import materialize_placeholder
from app.core.i18n import get_locale
from app.core.rate_limiter import upload_limiter
from app.core.storage import resolve_data_dir as _resolve_data_dir
from app.core.upload_guards import reject_if_xlsx_bomb
from app.core.upload_streaming import StreamedUpload, stream_upload_to_temp
from app.core.validation.messages import translate
from app.dependencies import CurrentUserId, RequirePermission, RequireRole, SessionDep, accessible_project_ids
from app.modules.bim_hub import file_storage as bim_file_storage
from app.modules.bim_hub.mesh_formats import is_mesh_geometry_ext, mesh_upload_hint
from app.modules.bim_hub.schemas import (
    AssetInfoUpdateRequest,
    AssetListResponse,
    AssetSummary,
    BIMElementBulkImport,
    BIMElementGroupCreate,
    BIMElementGroupResponse,
    BIMElementGroupUpdate,
    BIMElementListResponse,
    BIMElementResponse,
    BIMModelBOQLinkAggregate,
    BIMModelBOQLinksResponse,
    BIMModelCreate,
    BIMModelDiffResponse,
    BIMModelListResponse,
    BIMModelResponse,
    BIMModelUpdate,
    BIMQuantityMapCreate,
    BIMQuantityMapListResponse,
    BIMQuantityMapResponse,
    BIMQuantityMapUpdate,
    BOQElementLinkBrief,
    BOQElementLinkCreate,
    BOQElementLinkListResponse,
    BOQElementLinkResponse,
    BoqExportRequest,
    CreateModelFromDocumentRequest,
    FederationCreate,
    FederationDiffResponse,
    FederationFullResponse,
    FederationHealthResponse,
    FederationListResponse,
    FederationModelAdd,
    FederationModelResponse,
    FederationResponse,
    FederationSnapshot,
    FederationTypeTreeResponse,
    FederationUpdate,
    QuantityMapApplyRequest,
    QuantityMapApplyResult,
    SmartViewPreviewRequest,
    SmartViewPreviewResponse,
    SmartViewPropertyCatalogResponse,
)
from app.modules.bim_hub.service import BIMHubService

logger = logging.getLogger(__name__)

router = APIRouter(tags=["bim_hub"])


# Bytes of a geometry blob to read for the serve-time magic-byte check. The
# heaviest branch of _quick_validate_geometry_bytes inspects the first 16 KB
# (the glTF asset object); GLB (12 bytes) and COLLADA (4 KB) need far less.
# Reading a bounded head lets the local-disk serve path validate a 100+ MB GLB
# without paging the whole file into memory (issue #291).
_GEOMETRY_HEAD_PROBE_BYTES: int = 16 * 1024


def _quick_validate_geometry_bytes(blob: bytes, ext: str) -> tuple[bool, str]:
    """Fast, magic-byte / structural pre-check for served geometry.

    Mirrors the heavier ``_validate_geometry_file`` (which runs on ingest
    against a Path) but works on an already-in-memory byte buffer. We
    keep it cheap: only the first ~4 KB are inspected. Returns
    ``(ok, reason)``; the caller raises 422 with the reason when ok is
    False so the BIM viewer surfaces an actionable error instead of
    feeding garbage to Three.js loaders.

    Bug context: external user (Downtown Medical Center / Projet1, RVT)
    reported "Impossible de charger la géométrie 3D" with magic bytes
    ``3c 3f 78 6d 6c`` (``<?xml``) - the stored ``geometry.dae`` was
    XML but not COLLADA. Ingest-time validation existed but only for
    *new* uploads; old corrupt blobs kept streaming. This guard closes
    that gap for every read.
    """
    if not blob:
        return False, "empty buffer"
    if len(blob) < 200:
        return False, f"file suspiciously small ({len(blob)} bytes)"

    ext_norm = ext.lower()
    if ext_norm == ".glb":
        if blob[:4] != b"glTF":
            return False, (f"GLB magic mismatch - first 4 bytes are {blob[:4]!r}, expected b'glTF'")
        # Version is the 4-byte LE integer at offset 4.
        if len(blob) >= 12:
            version = int.from_bytes(blob[4:8], "little", signed=False)
            if version != 2:
                return False, f"unsupported GLB version {version} (expected 2)"
        return True, "ok"

    if ext_norm == ".dae":
        # Peek at the first 4 KB and verify a COLLADA root tag exists.
        # We deliberately do NOT do a full XML parse here - we trust
        # the in-memory tax of a 4 KB head scan and let the browser do
        # the heavy lifting once the file is known-good shape.
        # Accept namespace-prefixed roots like `<ns0:COLLADA>` (RVT /
        # DDC pipeline) as well as the bare `<COLLADA>` - both are valid
        # COLLADA per the XML namespace spec. Closes issue #153.
        import re as _re

        head = blob[:4096]
        try:
            head_text = head.decode("utf-8", errors="replace")
        except Exception as exc:  # pragma: no cover - utf-8 with errors='replace' can't raise
            return False, f"DAE head undecodable: {exc}"
        if not _re.search(r"<(?:[a-zA-Z_][\w.-]*:)?COLLADA\b", head_text, _re.IGNORECASE):
            # Surface what we DID find so the user/admin can recognise it
            # (e.g. "<ifcxml", "<gbxml", "<!doctype html").
            first_tag_match = _re.search(r"<([a-zA-Z_:][\w:.-]{0,40})", head_text)
            first_tag = f"<{first_tag_match.group(1)}>" if first_tag_match else "no root tag"
            return False, (f"DAE has no <COLLADA> root in first 4 KB (first tag found: {first_tag})")
        return True, "ok"

    if ext_norm == ".gltf":
        # gltf JSON - must parse as JSON object with an "asset" key.
        try:
            head = blob[: min(len(blob), 16384)]
            obj = json.loads(head.decode("utf-8", errors="replace"))
        except Exception as exc:
            return False, f"glTF JSON parse failed: {exc}"
        if not isinstance(obj, dict) or "asset" not in obj:
            return False, "glTF JSON missing required 'asset' field"
        return True, "ok"

    # Unknown extension - let it through (preserves prior behaviour for
    # any future extension we add without remembering to update this).
    return True, f"unknown extension {ext_norm}; skipped checks"


def _to_qty_float(val: object) -> float:
    """Best-effort numeric coercion for quantity-presence checks.

    Used by the upload-cad honesty gate (BUG-V320-DDC-01) to decide
    whether *any* imported element carries a real (non-zero, finite)
    quantity.  A string ``"0"`` or ``""`` or a NaN must read as 0.0 so a
    quantity-less import is not mistaken for a successful one.
    """
    if val is None or isinstance(val, bool):
        return 0.0
    try:
        f = float(val)
    except (ValueError, TypeError):
        return 0.0
    if f != f or f in (float("inf"), float("-inf")):
        return 0.0
    return f


# Legacy on-disk path kept only for backward compatibility with any
# external code that may still import ``_BIM_DATA_DIR``.  New code MUST
# go through :mod:`app.modules.bim_hub.file_storage` which wraps the
# pluggable :class:`~app.core.storage.StorageBackend`.  Resolved through the
# unified resolver so it points at the ACTIVE data root (honours OE_DATA_DIR /
# DATA_DIR / OE_CLI_DATA_DIR) rather than the package-relative default.
_BIM_DATA_DIR = _resolve_data_dir() / "bim"


def _get_service(session: SessionDep) -> BIMHubService:
    return BIMHubService(session)


# ═══════════════════════════════════════════════════════════════════════════════
# Project-ownership authorization helper
#
# Every BIM endpoint that touches a project (directly via ?project_id= or
# indirectly via a model/element/diff that belongs to a project) MUST call
# ``_verify_project_access`` before returning data or mutating state.
#
# This closes the IDOR from the v1.3.13 audit: previously any authenticated
# user could read/modify/delete models belonging to projects they do not own
# simply by guessing UUIDs. We resolve the underlying project, verify access
# (owner, admin, or team-member of a shared project) and return a 404 - not a
# 403 - so we also don't leak the existence of UUIDs the caller cannot see.
#
# Access policy is kept identical to the central ``verify_project_access`` in
# ``app.dependencies`` (owner OR admin OR team-member). Sharing a project with
# a non-admin user - via ``add_project_member`` - therefore grants the same
# read/write reach to its BIM models as it already grants to the project's
# documents, BOQ, schedule, etc. An earlier copy here checked only owner/admin,
# which wrongly denied legitimately-shared non-admin members the model-view
# endpoint (issue #271: file visible in documents, but "model not found" in
# the viewer until the member was made an admin).
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_project_access(
    session: AsyncSession,
    project_id: uuid.UUID,
    user_id: str,
) -> None:
    """Raise 404 unless the user owns, administers, or is a member of the project.

    Applies the same rule as the central ``app.dependencies.verify_project_access``
    helper - owner OR admin OR a ``TeamMembership`` on the project (shared
    access via ``add_project_member``) - so BIM access stays consistent with the
    documents / BOQ / schedule read paths. Emits 404 (not 403) on both "project
    missing" and "access denied" to avoid UUID enumeration.
    """
    from app.modules.projects.repository import ProjectRepository
    from app.modules.teams.access import is_project_member
    from app.modules.users.repository import UserRepository

    proj_repo = ProjectRepository(session)
    project = await proj_repo.get_by_id(project_id)
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=translate("errors.project_not_found", locale=get_locale()),
        )

    # Admin bypass - admins can touch any project regardless of ownership.
    try:
        user_repo = UserRepository(session)
        user = await user_repo.get_by_id(uuid.UUID(str(user_id)))
        if user is not None and getattr(user, "role", "") == "admin":
            return
    except Exception:
        # If the role lookup explodes, fall through to the ownership check -
        # never silently bypass authorization.
        logger.exception("Admin-role lookup failed during BIM access check")

    # Owner has full access.
    if str(project.owner_id) == str(user_id):
        return

    # Team-member check - any TeamMembership row for this project grants the
    # same access as ownership (shared, non-admin project members). This is the
    # branch that was missing before issue #271.
    try:
        if await is_project_member(session, project_id, uuid.UUID(str(user_id))):
            return
    except (ValueError, TypeError):
        pass  # malformed user_id - fall through to 404
    except Exception:
        logger.exception("Team-membership lookup failed during BIM access check")

    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail=translate("errors.project_not_found", locale=get_locale()),
    )


async def _verify_model_access(
    service: "BIMHubService",
    model_id: uuid.UUID,
    user_id: str,
) -> Any:
    """Load a BIM model and verify the caller may access its project.

    Access is owner, admin, or team-member of the model's project (see
    :func:`_verify_project_access`). Returns the model object so callers can
    reuse it without a second query. Raises 404 if the model is missing or the
    user has no access.
    """
    model = await service.get_model(model_id)
    if model is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Model not found",
        )
    await _verify_project_access(service.session, model.project_id, user_id)
    return model


# ═══════════════════════════════════════════════════════════════════════════════
# DataFrame column alias detection (flexible header matching)
# ═══════════════════════════════════════════════════════════════════════════════

_BIM_COLUMN_ALIASES: dict[str, list[str]] = {
    "element_id": [
        "element_id",
        "elementid",
        "id",
        "guid",
        "ifc_guid",
        "ifcguid",
        "global_id",
        "globalid",
        "stable_id",
        "stableid",
        "unique_id",
        "uniqueid",
        "revit_id",
        "elem_id",
    ],
    "element_type": [
        "element_type",
        "elementtype",
        "ifc_type",
        "ifctype",
        "object_type",
        "objecttype",
    ],
    # _-prefixed alias groups are NOT top-level BIMElement columns.
    # _rows_to_elements promotes them into the properties JSONB under
    # clean keys (category, family, type_name) so the frontend can
    # build BIM Browser-style hierarchy without data collisions.
    "_category": [
        "category",
        "elementcategory",
        "revit_category",
        "revitcategory",
        "ifc_class",
        "ifcclass",
        "class",
    ],
    "_family": [
        "family",
        "family_name",
        "familyname",
        "revit_family",
        "revitfamily",
        "family_and_type",
        "familyandtype",
    ],
    "_type_name": [
        "type_name",
        "typename",
        "type",
        "revit_type",
        "revittype",
    ],
    "name": [
        "name",
        "element_name",
        "elementname",
        "description",
        "bezeichnung",
        "label",
        "title",
    ],
    "storey": [
        "storey",
        "story",
        "level",
        "level_name",
        "levelname",
        "host_level_name",
        "hostlevelname",
        "floor",
        "floor_name",
        "etage",
        "geschoss",
        "building_storey",
        "buildingstorey",
        "ifc_storey",
        "ifcstorey",
        "base_constraint",
        "baseconstraint",
        "base_level",
        "baselevel",
        "reference_level",
        "referencelevel",
        "associated_level",
        "associatedlevel",
        "schedule_level",
        "schedulelevel",
    ],
    "mesh_ref": [
        "mesh_ref",
        "meshref",
        "mesh_id",
        "meshid",
        "node_id",
        "nodeid",
        "dae_node",
        "daenode",
        "collada_node",
        "colladanode",
        "geometry_ref",
        "geometryref",
    ],
    "bbox_min_x": [
        "bbox_min_x",
        "bboxminx",
        "min_x",
        "minx",
        "bounding_box_min_x",
        "boundingboxminx",
        "bb_min_x",
        "bbminx",
        "xmin",
    ],
    "bbox_min_y": [
        "bbox_min_y",
        "bboxminy",
        "min_y",
        "miny",
        "bounding_box_min_y",
        "boundingboxminy",
        "bb_min_y",
        "bbminy",
        "ymin",
    ],
    "bbox_min_z": [
        "bbox_min_z",
        "bboxminz",
        "min_z",
        "minz",
        "bounding_box_min_z",
        "boundingboxminz",
        "bb_min_z",
        "bbminz",
        "zmin",
    ],
    "bbox_max_x": [
        "bbox_max_x",
        "bboxmaxx",
        "max_x",
        "maxx",
        "bounding_box_max_x",
        "boundingboxmaxx",
        "bb_max_x",
        "bbmaxx",
        "xmax",
    ],
    "bbox_max_y": [
        "bbox_max_y",
        "bboxmaxy",
        "max_y",
        "maxy",
        "bounding_box_max_y",
        "boundingboxmaxy",
        "bb_max_y",
        "bbmaxy",
        "ymax",
    ],
    "bbox_max_z": [
        "bbox_max_z",
        "bboxmaxz",
        "max_z",
        "maxz",
        "bounding_box_max_z",
        "boundingboxmaxz",
        "bb_max_z",
        "bbmaxz",
        "zmax",
    ],
    "bounding_box": [
        "bounding_box",
        "boundingbox",
        "bbox",
        "aabb",
    ],
    "discipline": [
        "discipline",
        "disziplin",
        "trade",
        "gewerk",
        "domain",
        "system",
    ],
    "area_m2": [
        "area_m2",
        "area",
        "flaeche",
        "fläche",
        "surface_area",
        "surfacearea",
        "gross_area",
        "grossarea",
        "net_area",
        "netarea",
    ],
    "volume_m3": [
        "volume_m3",
        "volume",
        "volumen",
        "gross_volume",
        "grossvolume",
        "net_volume",
        "netvolume",
    ],
    "length_m": [
        "length_m",
        "length",
        "laenge",
        "länge",
        "span",
    ],
    "weight_kg": [
        "weight_kg",
        "weight",
        "gewicht",
        "mass",
        "masse",
    ],
    "properties": [
        "properties",
        "props",
        "attributes",
        "parameters",
        "pset",
    ],
}


def _match_bim_column(header: str) -> str | None:
    """Match a header string to a canonical BIM column name."""
    normalised = header.strip().lower().replace(" ", "_").replace("-", "_")
    for canonical, aliases in _BIM_COLUMN_ALIASES.items():
        if normalised in aliases:
            return canonical
    return normalised if normalised else None


# Properties-blob keys we will scan, in priority order, when the upload
# row has no top-level storey/level column. Most RVT and IFC exporters
# put the building level under "Level"; some put the host constraint
# under "Base Constraint" / "Reference Level" instead. Matched
# case-insensitively against the props dict.
_STOREY_PROPERTY_FALLBACK_KEYS: tuple[str, ...] = (
    "level",
    "base level",
    "baselevel",
    "base constraint",
    "baseconstraint",
    "reference level",
    "referencelevel",
    "host level",
    "hostlevel",
    "schedule level",
    "schedulelevel",
    "associated level",
    "associatedlevel",
    "building storey",
    "buildingstorey",
    "ifcbuildingstorey",
    "storey",
    "story",
    "floor",
    "etage",
    "geschoss",
)

# Literal-string sentinels that mean "no storey assigned" - RVT
# exports often write "None" / "<None>" instead of leaving the cell
# blank.  Matched case-insensitively after stripping.
_STOREY_NULL_LITERALS: frozenset[str] = frozenset(
    {
        "",
        "none",
        "null",
        "<none>",
        "n/a",
        "na",
        "-",
    }
)


def _normalise_storey(raw: Any) -> str | None:
    """Coerce a raw storey value to a clean string or None.

    Trims whitespace and treats common "no value" literals
    (``"None"``, ``"<None>"``, ``"N/A"``, ``"-"``, …) as None so
    they don't pollute the BIMFilterPanel storey list with a
    bogus bucket.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if text.lower() in _STOREY_NULL_LITERALS:
        return None
    return text


def _extract_storey(row: dict[str, Any], props: dict[str, Any]) -> str | None:
    """Resolve the building level for an element row.

    Priority:
        1. Top-level ``storey`` column (already aliased from ``level`` /
           ``base_level`` / etc. via :data:`_BIM_COLUMN_ALIASES`).
        2. Case-insensitive match against
           :data:`_STOREY_PROPERTY_FALLBACK_KEYS` inside the
           ``properties`` JSON blob - RVT/IFC exports frequently
           bury "Level" inside the property bag instead of promoting
           it to a column.

    Returns None if nothing usable is found, so downstream consumers
    can render the element as "no level" rather than crashing.
    """
    primary = _normalise_storey(row.get("storey"))
    if primary:
        return primary

    if not props:
        return None

    # Build a lower-cased lookup once so we can match keys regardless
    # of the export's casing convention ("Level" vs "level" vs "LEVEL").
    lc_props = {str(k).strip().lower(): v for k, v in props.items()}
    for key in _STOREY_PROPERTY_FALLBACK_KEYS:
        if key in lc_props:
            resolved = _normalise_storey(lc_props[key])
            if resolved:
                return resolved
    return None


def _safe_float(value: Any, default: float = 0.0) -> float:
    """Parse a value to float, returning *default* on failure."""
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return default
    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        if last_comma > last_dot:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except (ValueError, TypeError):
        return default


def _parse_bim_rows_from_csv(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from a CSV file for BIM element import."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Unable to decode CSV file - unsupported encoding")

    sniffer = csv.Sniffer()
    try:
        dialect = sniffer.sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel  # type: ignore[assignment]

    reader = csv.reader(io.StringIO(text), dialect)
    raw_headers = next(reader, None)
    if not raw_headers:
        raise ValueError("CSV file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        canonical = _match_bim_column(hdr)
        if canonical:
            column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in reader:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical:
                row[canonical] = val.strip() if isinstance(val, str) else val
        if row:
            rows.append(row)

    return rows


def _parse_bim_rows_from_excel(content_bytes: bytes) -> list[dict[str, Any]]:
    """Parse rows from an Excel (.xlsx) file for BIM element import."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Excel file has no worksheets")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, None)
    if not raw_headers:
        raise ValueError("Excel file is empty or has no header row")

    column_map: dict[int, str] = {}
    for idx, hdr in enumerate(raw_headers):
        if hdr is not None:
            canonical = _match_bim_column(str(hdr))
            if canonical:
                column_map[idx] = canonical

    rows: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        row: dict[str, Any] = {}
        for idx, val in enumerate(raw_row):
            canonical = column_map.get(idx)
            if canonical and val is not None:
                row[canonical] = val
        if row:
            rows.append(row)

    wb.close()
    return rows


def _rows_to_elements(
    rows: list[dict[str, Any]],
    has_geometry: bool,
) -> list[dict[str, Any]]:
    """Convert parsed rows into BIMElement-compatible dicts.

    Builds quantities JSON from area_m2, volume_m3, length_m, weight_kg columns.
    Parses properties column as JSON if present.
    """
    elements: list[dict[str, Any]] = []
    quantity_keys = {"area_m2", "volume_m3", "length_m", "weight_kg"}

    for row in rows:
        eid = str(row.get("element_id", "")).strip()
        if not eid:
            continue

        # Parse quantities
        quantities: dict[str, float] = {}
        for qk in quantity_keys:
            val = row.get(qk)
            if val is not None:
                fval = _safe_float(val)
                if fval != 0.0:
                    quantities[qk] = fval

        # Parse properties (could be JSON string)
        raw_props = row.get("properties")
        props: dict[str, Any] = {}
        if isinstance(raw_props, str) and raw_props.strip():
            try:
                props = json.loads(raw_props)
            except (json.JSONDecodeError, ValueError):
                props = {"raw": raw_props}
        elif isinstance(raw_props, dict):
            props = raw_props

        # Explicit mesh_ref column wins, otherwise fall back to element_id
        # (which matches DDC RvtExporter's DAE ``<node id="...">`` pattern).
        raw_mesh_ref = row.get("mesh_ref")
        if raw_mesh_ref is not None and str(raw_mesh_ref).strip():
            mesh_ref: str | None = str(raw_mesh_ref).strip()
        elif has_geometry:
            mesh_ref = eid
        else:
            mesh_ref = None

        # Explicit bbox - either a pre-built JSON blob in ``bounding_box`` or
        # six individual min/max columns.
        bbox: dict[str, float] | None = None
        raw_bbox = row.get("bounding_box")
        if isinstance(raw_bbox, dict):
            bbox = {k: float(v) for k, v in raw_bbox.items() if v is not None}
        elif isinstance(raw_bbox, str) and raw_bbox.strip():
            try:
                parsed = json.loads(raw_bbox)
                if isinstance(parsed, dict):
                    bbox = {k: float(v) for k, v in parsed.items() if v is not None}
            except (json.JSONDecodeError, ValueError, TypeError):
                pass
        if bbox is None:
            bbox_keys = ("bbox_min_x", "bbox_min_y", "bbox_min_z", "bbox_max_x", "bbox_max_y", "bbox_max_z")
            if any(row.get(k) is not None for k in bbox_keys):
                bbox = {
                    "min_x": _safe_float(row.get("bbox_min_x")),
                    "min_y": _safe_float(row.get("bbox_min_y")),
                    "min_z": _safe_float(row.get("bbox_min_z")),
                    "max_x": _safe_float(row.get("bbox_max_x")),
                    "max_y": _safe_float(row.get("bbox_max_y")),
                    "max_z": _safe_float(row.get("bbox_max_z")),
                }
                # Heuristic: if the numbers look like millimetres (range >10000
                # in any axis) convert to metres.
                ranges = [
                    abs(bbox["max_x"] - bbox["min_x"]),
                    abs(bbox["max_y"] - bbox["min_y"]),
                    abs(bbox["max_z"] - bbox["min_z"]),
                ]
                if any(r > 10_000 for r in ranges):
                    bbox = {k: v / 1000.0 for k, v in bbox.items()}

        # Promote _-prefixed canonical keys into properties under clean names.
        # These come from the split alias groups (_category, _family, _type_name)
        # that used to collide on element_type.
        _PROMOTE_TO_PROPS = {
            "_category": "category",
            "_family": "family",
            "_type_name": "type_name",
        }
        for raw_key, clean_key in _PROMOTE_TO_PROPS.items():
            val = row.get(raw_key)
            if val is not None:
                cleaned = str(val).strip()
                if cleaned and cleaned.lower() not in ("none", "null", "n/a", "-"):
                    props[clean_key] = cleaned

        # Collect any extra columns not in known canonical keys as properties
        bbox_col_keys = {
            "bounding_box",
            "bbox_min_x",
            "bbox_min_y",
            "bbox_min_z",
            "bbox_max_x",
            "bbox_max_y",
            "bbox_max_z",
        }
        known_keys = (
            {
                "element_id",
                "element_type",
                "name",
                "storey",
                "discipline",
                "properties",
                "mesh_ref",
            }
            | quantity_keys
            | bbox_col_keys
            | set(_PROMOTE_TO_PROPS.keys())
        )
        for k, v in row.items():
            if k not in known_keys and v is not None and str(v).strip():
                props[k] = v

        # If element_type is empty after the alias split (because the Excel
        # only had "Category" and "Type" columns, both now redirected away
        # from element_type), fall back to category (the broadest
        # classification) so we don't store rows with no type at all.
        raw_element_type = str(row.get("element_type", "")).strip() or None
        if not raw_element_type and props.get("category"):
            raw_element_type = props["category"]

        # Fallback quantity recovery (issue #347): when the fixed-name pass
        # above found nothing, a DDC export simply labelled its quantity columns
        # differently. Recover them from the row / properties; only if there is
        # still no numeric quantity at all fall back to coarse bounding-box
        # figures, tagged so they read as estimated rather than measured.
        if not quantities:
            from app.modules.bim_hub.quantity_fallback import (
                derive_quantities_from_bbox,
                derive_quantities_from_columns,
            )

            derived = derive_quantities_from_columns({**row, **props})
            for _dim, _qkey in (
                ("area", "area_m2"),
                ("volume", "volume_m3"),
                ("length", "length_m"),
                ("weight", "weight_kg"),
            ):
                if _dim in derived:
                    quantities[_qkey] = derived[_dim]
            if not quantities and bbox:
                estimated = derive_quantities_from_bbox(bbox, raw_element_type)
                for _dim, _qkey in (
                    ("area", "area_m2"),
                    ("volume", "volume_m3"),
                    ("length", "length_m"),
                ):
                    if _dim in estimated:
                        quantities[_qkey] = estimated[_dim]
                if quantities:
                    props.setdefault("quantities_source", "geometry_bbox")

        element: dict[str, Any] = {
            "stable_id": eid,
            "element_type": raw_element_type,
            "name": str(row.get("name", "")).strip() or None,
            # Resolve storey from the top-level column first; if absent
            # (or a "None"/"-" sentinel), fall back to scanning the
            # properties blob for a "Level" / "Base Constraint" / etc.
            # key. See _extract_storey for the full priority chain.
            "storey": _extract_storey(row, props),
            "discipline": str(row.get("discipline", "")).strip() or None,
            "quantities": quantities,
            "properties": props,
            "mesh_ref": mesh_ref,
            "bounding_box": bbox,
        }
        elements.append(element)

    return elements


# ═══════════════════════════════════════════════════════════════════════════════
# Upload (DataFrame + optional DAE geometry)
# ═══════════════════════════════════════════════════════════════════════════════

# Upload size caps for the DataFrame + geometry drop. The element table is read
# back into memory for the CSV/openpyxl parse, so it is capped tighter than the
# geometry blob, which only ever streams to disk on its way to storage.
MAX_BIM_DATA_BYTES: int = 100 * 1024 * 1024  # 100 MB CSV/xlsx element table
MAX_BIM_GEOMETRY_BYTES: int = 500 * 1024 * 1024  # 500 MB DAE/GLB/glTF geometry


@router.post("/upload/", status_code=201)
async def upload_bim_data(
    background_tasks: BackgroundTasks,
    project_id: str = Query(..., description="Project UUID"),
    name: str = Query(default="Imported Model", max_length=255),
    discipline: str = Query(default="architecture", max_length=50),
    data_file: UploadFile = File(..., description="CSV or Excel file with element data"),
    geometry_file: UploadFile | None = File(default=None, description="DAE/COLLADA geometry file"),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Upload BIM data from Cad2Data converter output.

    Accepts a DataFrame file (CSV/Excel) with one row per building element
    and an optional COLLADA (.dae) geometry file where each mesh node has
    an ID matching ``element_id`` from the DataFrame.

    Expected DataFrame columns (flexible auto-detection via aliases):
    - **element_id / id / guid** -- unique element identifier (required)
    - **element_type / type / category** -- element classification
    - **name / description** -- human-readable name
    - **storey / level / floor** -- building storey assignment
    - **discipline / trade** -- discipline (architecture, structural, ...)
    - **area_m2 / area** -- area in m2
    - **volume_m3 / volume** -- volume in m3
    - **length_m / length** -- length in m
    - **weight_kg / weight** -- weight in kg
    - **properties** -- JSON string of additional properties

    Returns:
        Summary with model_id, element_count, storeys, and disciplines.
    """
    # --- Verify project access (IDOR guard) ---
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid project_id: {exc}",
        ) from exc
    await _verify_project_access(service.session, project_uuid, user_id or "")

    allowed, _ = upload_limiter.is_allowed(str(user_id or "anon"))
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many uploads. Please wait a moment and try again.",
            headers={"Retry-After": "60"},
        )

    # --- Validate data file ---
    data_filename = (data_file.filename or "").lower()
    if not data_filename.endswith((".csv", ".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported data file type. Please upload CSV (.csv) or Excel (.xlsx) file.",
        )

    # Stream the element table to a bounded temp file (1 MB chunks, aborts past
    # the cap) so an oversized body never lands fully in RAM, then read the now
    # bounded bytes back for the CSV/openpyxl parser.
    data_suffix = pathlib.Path(data_filename).suffix or ".csv"
    try:
        async with stream_upload_to_temp(
            data_file,
            max_bytes=MAX_BIM_DATA_BYTES,
            suffix=data_suffix,
        ) as data_upload:
            data_content = data_upload.path.read_bytes()
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=(
                f"Data file exceeds the {MAX_BIM_DATA_BYTES // (1024 * 1024)} MB upload limit. "
                "Split the element table or export a lighter file."
            ),
        ) from exc

    if not data_content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded data file is empty.",
        )

    # xlsx is a zip container; reject a decompression bomb before openpyxl
    # expands the full sheet in memory.
    if data_filename.endswith((".xlsx", ".xls")):
        reject_if_xlsx_bomb(data_content)

    # --- Validate geometry file type (if provided) ---
    # The bytes are streamed to disk later (inside the upload stack) so a large
    # mesh export never has to sit in memory; here we only reject a bad type.
    if geometry_file is not None:
        geo_filename = (geometry_file.filename or "").lower()
        if not geo_filename.endswith((".dae", ".glb", ".gltf")):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Unsupported geometry file type. Please upload DAE (.dae), GLB (.glb), or glTF (.gltf) file.",
            )

    # --- Parse data file ---
    try:
        if data_filename.endswith((".xlsx", ".xls")):
            rows = _parse_bim_rows_from_excel(data_content)
        else:
            rows = _parse_bim_rows_from_csv(data_content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse data file: {exc}",
        ) from exc
    except Exception as exc:
        logger.exception("Unexpected error parsing BIM data file")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Failed to parse data file: {exc}",
        ) from exc

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No data rows found in the uploaded file.",
        )

    # Stream the geometry (if any) to a bounded temp file and keep it on disk
    # through model creation so it can be persisted by path - a large DAE/GLB
    # export never lands fully in RAM. The stack removes any leftover temp file
    # once the geometry has been handed to storage.
    async with contextlib.AsyncExitStack() as upload_stack:
        geometry_upload: StreamedUpload | None = None
        if geometry_file is not None:
            geo_suffix = pathlib.Path(geometry_file.filename or "geometry.dae").suffix or ".dae"
            try:
                geometry_upload = await upload_stack.enter_async_context(
                    stream_upload_to_temp(
                        geometry_file,
                        max_bytes=MAX_BIM_GEOMETRY_BYTES,
                        suffix=geo_suffix,
                    )
                )
            except ValueError as exc:
                raise HTTPException(
                    status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                    detail=(
                        f"Geometry file exceeds the {MAX_BIM_GEOMETRY_BYTES // (1024 * 1024)} MB "
                        "upload limit. Export a lighter mesh or split the model."
                    ),
                ) from exc
        has_geometry = geometry_upload is not None and geometry_upload.size > 0

        # --- Convert rows to element dicts ---
        element_dicts = _rows_to_elements(rows, has_geometry=has_geometry)
        if not element_dicts:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid elements found. Ensure the file has an 'element_id' column.",
            )

        # --- Determine format from geometry file extension ---
        geo_ext = ""
        if geometry_file and geometry_file.filename:
            geo_ext = pathlib.Path(geometry_file.filename).suffix.lstrip(".").lower()

        model_format = geo_ext if geo_ext else "csv"

        # --- Create BIM model ---
        from app.modules.bim_hub.schemas import BIMModelCreate

        model_data = BIMModelCreate(
            project_id=uuid.UUID(project_id),
            name=name,
            discipline=discipline,
            model_format=model_format,
            status="processing",
        )
        model = await service.create_model(model_data, user_id=user_id)
        model_id = model.id

        # --- Save geometry to storage (streamed from the temp path) ---
        if has_geometry and geometry_upload is not None:
            ext = pathlib.Path(geometry_file.filename or "geometry.dae").suffix or ".dae"  # type: ignore[union-attr]
            await bim_file_storage.save_geometry_from_path(
                project_id=project_id,
                model_id=str(model_id),
                ext=ext,
                src_path=geometry_upload.path,
                size=geometry_upload.size,
            )

    # --- Import elements ---
    from app.modules.bim_hub.schemas import BIMElementCreate

    elements_create = [
        BIMElementCreate(
            stable_id=ed["stable_id"],
            element_type=ed.get("element_type"),
            name=ed.get("name"),
            storey=ed.get("storey"),
            discipline=ed.get("discipline") or discipline,
            quantities=ed.get("quantities", {}),
            properties=ed.get("properties", {}),
            mesh_ref=ed.get("mesh_ref"),
            bounding_box=ed.get("bounding_box"),
        )
        for ed in element_dicts
    ]
    created_elements = await service.bulk_import_elements(model_id, elements_create)

    # Compute summary
    storeys = sorted({e.storey for e in created_elements if e.storey})
    disciplines_found = sorted({e.discipline for e in created_elements if e.discipline})

    logger.info(
        "BIM upload complete: model=%s, elements=%d, storeys=%d, disciplines=%s",
        name,
        len(created_elements),
        len(storeys),
        disciplines_found,
    )

    # Validation is part of the canonical import pipeline (Import -> VALIDATE
    # -> Store). The CAD converter path already runs it; the CSV/Excel drop
    # must too. Schedule the advisory pass in the background so a slow rule run
    # never delays the upload response.
    if created_elements:
        background_tasks.add_task(_run_import_validation, model_id)
        # Bake the fast-viewer + property artifacts (GLB from DAE, streaming
        # tileset, Parquet property sidecar) off-request so a non-CAD upload
        # gets the same fast viewer and populated property panel as the CAD
        # path. Best-effort and idempotent (see _ensure_model_artifacts). Runs
        # after the request session commits the model + elements.
        background_tasks.add_task(_ensure_model_artifacts, project_id, str(model_id))

    return {
        "model_id": str(model_id),
        "element_count": len(created_elements),
        "storeys": storeys,
        "disciplines": disciplines_found,
        "has_geometry": has_geometry,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Direct CAD file upload (RVT, IFC, DWG, DGN)
# ═══════════════════════════════════════════════════════════════════════════════

# The formats a server-side converter can actually turn into a canonical BIM
# model. This set is also printed verbatim as the "Accepted:" list when an
# upload is refused, so it has to be the truth rather than an aspiration.
#
# FBX, OBJ and 3DS used to sit here. They are geometry-only meshes, every
# caller runs ``is_mesh_geometry_ext`` first and routes them to the in-browser
# importer, so the three were unreachable - they could never be accepted, yet
# the rejection message kept advertising them to the next person who uploaded
# something else. See ``mesh_formats.py``.
_ALLOWED_CAD_EXTENSIONS = {".rvt", ".ifc", ".dwg", ".dgn"}

# Formats that require an external converter binary. IFC has a built-in
# text fallback parser, so it's NOT in this set; XLSX/CSV go through a
# separate upload endpoint and aren't relevant here.
_NEEDS_CONVERTER_EXTS = {".rvt", ".dwg", ".dgn"}


# ═══════════════════════════════════════════════════════════════════════════════
# Background workers - invoked via FastAPI BackgroundTasks so the upload
# request returns in milliseconds even when DDC conversion takes minutes.
# Each worker uses a fresh AsyncSession (the request session is closed by
# the time the task runs) and the same storage abstraction as the router.
# ═══════════════════════════════════════════════════════════════════════════════


# Module-level counter for Parquet sidecar write failures. No prometheus
# framework is wired into the backend yet (see task #155), so we expose
# the count via a structured log event AND a process-local counter that
# the future /metrics surface can poll without re-instrumenting.
_PARQUET_FAILURE_COUNTER: dict[str, int] = {}


def _record_parquet_attempt_init() -> tuple[str, str | None]:
    """Return the default (status, error) tuple before a Parquet attempt.

    "skipped" is the conservative default: if the path that mutates these
    values is never reached (raw_elements empty, exception before the
    try/except, etc.) the UI still gets a sane value rather than ``None``.
    """
    return "skipped", None


def _surface_parquet_failure(
    *,
    exc: BaseException,
    project_id: str,
    model_id: str,
    model_format: str,
    row_count: int,
) -> tuple[str, str]:
    """Emit structured diagnostics for a failed Parquet sidecar write.

    Logs at ERROR level (degraded state - even if non-fatal for the user's
    current upload, operators should see how often this fires), bumps a
    process-local counter, and returns the ``(status, error)`` pair the
    caller stamps onto ``model.metadata_``.

    The structured payload travels in ``logger.error``'s ``extra`` dict so
    log-aggregation systems (ELK / Loki / Datadog) can pick the fields up
    individually rather than parsing the message string.
    """
    import traceback as _tb

    exc_type = type(exc).__name__
    exc_msg = str(exc) or "<no message>"
    # Truncate traceback to ~2KB so we don't blow up log lines or the
    # metadata JSON column for pathological deep stacks.
    tb_text = "".join(_tb.format_exception(type(exc), exc, exc.__traceback__))
    if len(tb_text) > 2048:
        tb_text = tb_text[:2048] + "...[truncated]"

    fmt_key = (model_format or "unknown").lower().lstrip(".")
    _PARQUET_FAILURE_COUNTER[fmt_key] = _PARQUET_FAILURE_COUNTER.get(fmt_key, 0) + 1

    logger.error(
        "bim_parquet_write_failed model=%s project=%s format=%s rows=%d exc=%s msg=%s",
        model_id,
        project_id,
        fmt_key,
        row_count,
        exc_type,
        exc_msg,
        extra={
            "event": "bim_parquet_write_failed",
            "project_id": project_id,
            "model_id": model_id,
            "model_format": fmt_key,
            "row_count": row_count,
            "exception_type": exc_type,
            "exception_message": exc_msg,
            "traceback": tb_text,
            "metric": "bim_parquet_write_failed_total",
            "metric_labels": {"model_format": fmt_key},
        },
    )

    # The error string stamped onto the model row stays short - the full
    # traceback is in the log event. The UI only needs enough to render
    # "ParquetWriteError: disk full" without exposing the whole stack.
    short_error = f"{exc_type}: {exc_msg}"[:500]
    return "failed", short_error


async def _run_import_validation(model_uuid: uuid.UUID) -> None:
    """Run the mandatory validation pass over a freshly-imported BIM model.

    Part of the canonical import pipeline (Import -> Convert -> VALIDATE ->
    Store). Invoked from :func:`_process_cad_in_background` once the converted
    elements are committed. Delegates to :class:`BIMValidationService`, which
    runs the universal BIM rule set (required properties, completeness,
    classification) and persists a ``ValidationReport`` row
    (``target_type='bim_model'``) - the exact same path the on-demand
    "Validate" button uses, so the BIM detail view picks the report up with no
    extra wiring.

    Best-effort by contract: any failure here is logged and swallowed. A good
    geometry import must never be downgraded because the (advisory) validation
    pass hit a problem - validation findings live in their own report, separate
    from the model's processing status.
    """
    from app.database import async_session_factory

    try:
        from app.modules.validation.bim_validation_service import BIMValidationService

        async with async_session_factory() as val_session:
            service = BIMValidationService(val_session)
            report = await service.validate_bim_model(model_uuid)
            await val_session.commit()
            logger.info(
                "Import validation complete for model %s: report=%s status=%s score=%s",
                model_uuid,
                report.id,
                report.status,
                report.score,
            )
    except Exception as exc:  # noqa: BLE001 - validation is advisory, never fatal
        logger.warning(
            "Import validation failed for model %s (non-fatal): %s",
            model_uuid,
            exc,
        )


async def _ensure_model_artifacts(project_id: str, model_id: str) -> None:
    """Best-effort: ensure a model's GLB + streaming tileset + Parquet sidecar.

    Delegates to :meth:`BIMHubService.ensure_artifacts`, which is idempotent
    and self-guards each step. Opens its own :func:`async_session_factory`
    session so it is safe to run from a background worker or a
    ``BackgroundTasks`` callback (the request session is closed by then, and
    the model + elements are committed before the callback fires). Any failure
    is logged and swallowed - a good import must never fail because the
    (advisory) artifact bake hit a problem.
    """
    from app.database import async_session_factory

    try:
        async with async_session_factory() as session:
            await BIMHubService(session).ensure_artifacts(project_id, model_id)
    except Exception as exc:  # noqa: BLE001 - artifacts are best-effort
        logger.warning("Artifact baking failed for model %s (non-fatal): %s", model_id, exc)


async def _process_cad_in_background(
    *,
    project_id: str,
    model_id: str,
    cad_storage_key: str,
    ext: str,
    conversion_depth: str,
) -> None:
    """Run DDC conversion + element extraction for an uploaded CAD file.

    Scheduled after the upload endpoint returns so the HTTP request finishes
    in milliseconds while the (potentially minutes-long) conversion happens
    off the request path.  Updates the model row's ``status`` to
    ``ready`` / ``error`` / ``needs_converter`` when finished - the frontend
    already polls ``GET /{model_id}`` and transitions the UI automatically.
    """
    import asyncio
    import tempfile
    import uuid as _uuid
    from pathlib import Path as _Path

    from sqlalchemy import select

    from app.core.storage import get_storage_backend
    from app.database import async_session_factory
    from app.modules.bim_hub.ifc_processor import process_ifc_file
    from app.modules.bim_hub.models import BIMElement, BIMModel

    model_uuid = _uuid.UUID(model_id)

    # Pre-flight: check the converter binary up front so we can surface a
    # specific actionable error ("install the RVT converter") instead of the
    # generic "no elements extracted" message that lands when DDC isn't found
    # mid-pipeline. IFC has a built-in text fallback so we don't gate on it.
    #
    # Two-stage check (added 2026-04-28 to address "many problems with the
    # converters and downloading"):
    #   1. ``find_converter`` - file exists on disk (cheap, file-stat only).
    #   2. ``smoke_test_converter`` - binary actually loads (8 s timeout,
    #      result cached 5 min). Catches Qt-DLL / Mark-of-the-Web /
    #      VC-Redist-missing breakage that the first check can't see.
    if ext.lower() == ".rvt":
        try:
            from app.modules.boq.cad_import import (
                find_converter as _fc,
            )
            from app.modules.boq.cad_import import (
                smoke_test_converter as _smoke,
            )

            converter_id = "rvt"
            failure_reason: str | None = None
            failure_code: str | None = None
            suggested_actions: list[str] = []

            if _fc(converter_id) is None:
                failure_code = "ddc_not_found"
                failure_reason = (
                    "The RVT converter is not installed on this "
                    "server. Open Settings → BIM Converters and click "
                    "Install for RVT, then click Retry on this model."
                )
                suggested_actions = ["install_converter"]
            else:
                # Run the cached smoke test before the (expensive) RVT
                # conversion. If the binary fails to load we report the
                # DLL/perm error to the user immediately rather than
                # letting them wait through a 5-minute conversion that
                # has no chance of succeeding.
                health = await asyncio.to_thread(_smoke, converter_id)
                if health["status"] != "ok":
                    failure_code = "ddc_smoke_failed"
                    failure_reason = health["message"] or (
                        "The RVT converter is installed but the "
                        "smoke test failed. Open Settings → BIM Converters "
                        "and click Reinstall, then click Retry on this model."
                    )
                    suggested_actions = list(health["suggested_actions"]) or ["reinstall_converter"]

            if failure_code:
                async with async_session_factory() as session:
                    model = (
                        await session.execute(select(BIMModel).where(BIMModel.id == model_uuid))
                    ).scalar_one_or_none()
                    if model is not None:
                        model.status = "needs_converter"
                        model.error_message = failure_reason
                        meta = dict(model.metadata_ or {})
                        meta["error_code"] = failure_code
                        meta["converter_id"] = converter_id
                        meta["suggested_actions"] = suggested_actions
                        meta["install_endpoint"] = f"/api/v1/takeoff/converters/{converter_id}/install/"
                        meta["verify_endpoint"] = f"/api/v1/takeoff/converters/{converter_id}/verify/"
                        model.metadata_ = meta
                        await session.commit()
                logger.warning(
                    "RVT pre-flight failed for model %s: %s",
                    model_id,
                    failure_code,
                )
                return
        except Exception:  # noqa: BLE001 - pre-flight is best-effort
            logger.exception("RVT converter pre-flight check failed for %s", model_id)

    # IFC pre-provision: always try to get the real DDC IfcExporter so the
    # viewer shows accurate meshes, not the crude box placeholder. IFC has a
    # text fallback (so we do NOT gate the import on the converter the way RVT
    # does), but on a fresh install the converter is missing and the inline
    # ``ensure_converter`` call inside ``process_ifc_file`` would race a cold
    # ~30 MB download. Provisioning it here, up front and off the request path,
    # means the conversion below uses the real converter on the first upload
    # instead of silently degrading to placeholder geometry. This is
    # best-effort: a genuine failure (offline / locked-down host) just falls
    # through to the text fallback, so the import still completes.
    if ext.lower() == ".ifc":
        try:
            from app.modules.boq.cad_import import (
                ensure_converter as _ensure,
            )
            from app.modules.boq.cad_import import (
                find_converter as _fc,
            )

            if _fc("ifc") is None:
                logger.info(
                    "IFC converter not present for model %s - provisioning the DDC IfcExporter before conversion",
                    model_id,
                )
                converter_path = await asyncio.to_thread(_ensure, "ifc")
                logger.info("IFC converter provisioned at %s for model %s", converter_path, model_id)
        except Exception as exc:  # noqa: BLE001 - provisioning is best-effort
            # Do NOT fail the import here: the text fallback still produces a
            # usable (placeholder) result, and the metadata path below will
            # honestly flag it as placeholder so the UI nudges the user.
            logger.warning(
                "IFC converter auto-provision failed for model %s (%s) - "
                "falling back to the built-in text parser (placeholder geometry)",
                model_id,
                exc,
            )

    try:
        content = await get_storage_backend().get(cad_storage_key)

        with tempfile.TemporaryDirectory(prefix="oe-bim-bg-") as _tmp_str:
            _tmp_dir = _Path(_tmp_str)
            _tmp_cad_path = _tmp_dir / f"original{ext}"
            await asyncio.to_thread(_tmp_cad_path.write_bytes, content)

            # Bind a failure record to this upload before handing the
            # conversion to the thread pool. Other uploads convert at the
            # same time on that pool, and the diagnostics we read below
            # (stderr tail, converter and authoring-app versions) must be
            # the ones this file produced, not whichever conversion failed
            # last. The worker inherits a copy of this context, so it fills
            # in the same record object; we take our copy of it here, while
            # the scope is still open.
            from app.modules.bim_hub.ifc_processor import ddc_failure_scope

            with ddc_failure_scope() as _ddc_record:
                result = await asyncio.to_thread(process_ifc_file, _tmp_cad_path, _tmp_dir, conversion_depth)
                ddc_failure = dict(_ddc_record)
            element_count = result["element_count"]

            geo_key: str | None = None
            geo_local = result.get("geometry_path")
            if geo_local:
                _geo_path = _Path(geo_local)
                if _geo_path.is_file():
                    _geo_bytes = await asyncio.to_thread(_geo_path.read_bytes)
                    _geo_ext = _geo_path.suffix or ".dae"
                    geo_key = await bim_file_storage.save_geometry(
                        project_id=project_id,
                        model_id=model_id,
                        ext=_geo_ext,
                        content=_geo_bytes,
                    )

            glb_key: str | None = None
            glb_local = result.get("glb_path")
            if glb_local:
                _glb_path = _Path(glb_local)
                if _glb_path.is_file():
                    _glb_bytes = await asyncio.to_thread(_glb_path.read_bytes)
                    glb_key = await bim_file_storage.save_geometry(
                        project_id=project_id,
                        model_id=model_id,
                        ext=".glb",
                        content=_glb_bytes,
                    )
                    logger.info("GLB geometry saved: %s (%d bytes)", glb_key, len(_glb_bytes))

            raw_elements = result.get("raw_elements", [])
            parquet_status, parquet_error = _record_parquet_attempt_init()
            if raw_elements:
                try:
                    from app.modules.bim_hub.dataframe_store import write_dataframe

                    await asyncio.to_thread(
                        write_dataframe,
                        project_id=project_id,
                        model_id=model_id,
                        rows=raw_elements,
                    )
                    parquet_status = "ok"
                except Exception as exc:
                    parquet_status, parquet_error = _surface_parquet_failure(
                        exc=exc,
                        project_id=project_id,
                        model_id=model_id,
                        model_format=ext.lstrip("."),
                        row_count=len(raw_elements),
                    )
            else:
                # No rows means we never tried - keep "skipped" so the UI
                # can distinguish "ingest produced nothing" from a real
                # write failure that needs operator attention.
                parquet_status = "skipped"

        async with async_session_factory() as session:
            # Defensive: the upload endpoint commits before scheduling us, but
            # retry a few times anyway in case of slow disk flushes / connection
            # pool churn during a heavy upload burst.
            model = None
            for _attempt in range(5):
                model = (await session.execute(select(BIMModel).where(BIMModel.id == model_uuid))).scalar_one_or_none()
                if model is not None:
                    break
                await asyncio.sleep(0.2)
            if model is None:
                logger.error("Background processor: model %s vanished mid-conversion", model_id)
                return

            if element_count > 0:
                # Top-level geometry_quality drives the placeholder banner.
                # Stamp it onto each element's properties so the frontend
                # viewer can self-detect placeholders without an extra API
                # round-trip for the model metadata.
                result_quality = result.get("geometry_quality") or result.get("geometry_type")
                # BUG-V320-DDC-01 / D-TKC-NEW-01 - honesty gate.  When the DDC
                # cad2data converter is unavailable the IFC text-parser still
                # imports element geometry, but it can only recover quantities
                # if the file happens to ship explicit IfcElementQuantity
                # blocks.  Track whether *any* element carries a non-empty
                # quantities map; if none do, we must NOT advertise the model
                # as a clean 'ready' import with error_message=null - that is
                # the dishonest "successful import, zero quantities" state the
                # QA audit flagged.
                any_quantities = False
                for elem_data in result["elements"]:
                    el_props = dict(elem_data.get("properties", {}) or {})
                    if elem_data.get("is_placeholder") or result_quality == "placeholder":
                        el_props["is_placeholder"] = True
                    el_quantities = elem_data.get("quantities", {}) or {}
                    if el_quantities and any(_to_qty_float(v) for v in el_quantities.values()):
                        any_quantities = True
                    el = BIMElement(
                        model_id=model_uuid,
                        stable_id=elem_data["stable_id"],
                        element_type=elem_data.get("element_type"),
                        name=elem_data.get("name"),
                        storey=elem_data.get("storey"),
                        discipline=elem_data.get("discipline"),
                        properties=el_props,
                        quantities=el_quantities,
                        geometry_hash=elem_data.get("geometry_hash"),
                        bounding_box=elem_data.get("bounding_box"),
                        mesh_ref=elem_data.get("mesh_ref"),
                    )
                    session.add(el)

                converter_absent = result_quality == "placeholder"
                no_quantities = not any_quantities

                model.status = "ready"
                model.element_count = element_count
                model.storey_count = len(result["storeys"])
                model.bounding_box = result.get("bounding_box")
                # BUG-006: stamp the conversion-finished moment so audit
                # reports / Asset Register sorting have something better
                # than ``created_at`` (which is the upload moment, before
                # geometry extraction).  ``import_date`` was previously
                # left null on every ready model.
                from datetime import UTC as _UTC
                from datetime import datetime as _dt

                # ``import_date`` is a string column (ISO-8601), not a
                # timestamp.  Assigning a ``datetime`` worked on SQLite (silent
                # coercion) but asyncpg rejects a datetime bound to a
                # ``::VARCHAR`` parameter, so every RVT/IFC conversion failed to
                # persist on PostgreSQL ("expected str, got datetime").  Store a
                # second-precision UTC ISO string (20 chars) so it fits both the
                # column and the API schema's length bound.
                model.import_date = _dt.now(_UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
                if glb_key:
                    model.canonical_file_path = glb_key
                elif geo_key:
                    model.canonical_file_path = geo_key
                model.metadata_ = {
                    **(model.metadata_ or {}),
                    "geometry_type": result.get("geometry_type", "unknown"),
                    # geometry_quality drives the frontend's "placeholder
                    # geometry" banner - set to "placeholder" when DDC
                    # cad2data is unavailable and we synthesized boxes.
                    "geometry_quality": result.get(
                        "geometry_quality",
                        result.get("geometry_type", "unknown"),
                    ),
                    # DDC converter version stamp - drives the "Processed
                    # with DDC v{X}" badge on the BIM model card and the
                    # /about page. Both keys are optional: missing values
                    # leave the badge hidden (v3.12.0 / Stream D).
                    **({"converter_version": result["converter_version"]} if result.get("converter_version") else {}),
                    **({"converter_source": result["converter_source"]} if result.get("converter_source") else {}),
                    # Honesty signal - set when the import only succeeded
                    # because the runtime stripped modern CLI args (or
                    # retried without them) to accommodate an older DDC
                    # binary.  Drives the post-success "converter
                    # outdated" nudge in the BIM viewer.  Conversion
                    # itself is fine; the warning is informational only.
                    **({"converter_cli_outdated": True} if result.get("converter_cli_outdated") else {}),
                    # Parquet sidecar status - non-fatal for the model
                    # itself, but operators want to know how often the
                    # sidecar write fails so the analytics surface
                    # (/dataframe/* endpoints) doesn't silently return
                    # empty data. Task #155 surfaces this in the API.
                    "parquet_status": parquet_status,
                    "parquet_error": parquet_error,
                    "parquet_attempted_at": _dt.now(_UTC).isoformat(),
                }

                # BUG-V320-DDC-01 / D-TKC-NEW-01 - non-destructive honesty
                # path.  The elements were imported (geometry is useful for
                # the viewer / element linking) but if the DDC converter was
                # absent OR no quantities could be extracted we downgrade the
                # status from a misleading 'ready' to a distinct 'degraded'
                # state and populate a user-facing warning so the UI can show
                # "imported, but no quantities - DDC converter required"
                # instead of pretending the import fully succeeded.
                if converter_absent or no_quantities:
                    model.status = "degraded"
                    meta_warn = dict(model.metadata_ or {})
                    if converter_absent and no_quantities:
                        warn_msg = (
                            "Geometry imported, but no quantities could be "
                            "extracted: the DDC cad2data converter is not "
                            "available on this server, and the file does not "
                            "carry explicit IFC BaseQuantities. Elements were "
                            "imported for geometry/linking only. Install the "
                            "DDC converter (Settings → BIM Converters) and "
                            "click Retry to recover area/volume/length "
                            "quantities."
                        )
                        meta_warn["error_code"] = "no_quantities_converter_absent"
                    elif converter_absent:
                        warn_msg = (
                            "Imported with placeholder geometry: the DDC "
                            "cad2data converter is not available on this "
                            "server, so quantities were read from the file's "
                            "embedded IFC BaseQuantities only and may be "
                            "incomplete. Install the DDC converter "
                            "(Settings → BIM Converters) and click Retry for "
                            "full geometry-derived quantities."
                        )
                        meta_warn["error_code"] = "converter_absent"
                    else:
                        warn_msg = (
                            "Geometry imported, but no quantities "
                            "(area/volume/length) could be extracted from "
                            "this file. Elements were imported for "
                            "geometry/linking only."
                        )
                        meta_warn["error_code"] = "no_quantities"
                    model.error_message = warn_msg
                    meta_warn["warning"] = warn_msg
                    meta_warn["degraded"] = True
                    meta_warn["converter_id"] = "ifc"
                    meta_warn["install_endpoint"] = "/api/v1/takeoff/converters/ifc/install/"
                    model.metadata_ = meta_warn
                    logger.warning(
                        "Background CAD processed but DEGRADED (converter_absent=%s "
                        "no_quantities=%s): %d elements → model %s degraded",
                        converter_absent,
                        no_quantities,
                        element_count,
                        model_id,
                    )
                else:
                    logger.info(
                        "Background CAD processed: %d elements, %d storeys → model %s ready",
                        element_count,
                        len(result["storeys"]),
                        model_id,
                    )

                # Storage policy - drop the raw upload after a *successful*
                # conversion when ``keep_original_cad`` is False (production
                # default).  Failed conversions fall through to the else
                # branch below and keep the original so retry works without
                # re-upload.  Conversion artifacts (GLB/DAE/parquet) stay
                # forever regardless.
                from app.config import get_settings as _get_settings

                if not _get_settings().keep_original_cad:
                    await bim_file_storage.delete_original_cad(
                        project_id=project_id,
                        model_id=model_id,
                        ext=ext,
                    )
            else:
                meta = dict(model.metadata_ or {})

                # The structured failure context this conversion recorded
                # (RVT version, converter version, stderr tail), captured
                # above while its scope was open. If it's present, we can
                # build a much more specific error message than the legacy
                # "converter not installed" boilerplate.
                rvt_info = ddc_failure.get("rvt_info") or {}
                conv_info = ddc_failure.get("converter_info") or {}
                rvt_app = rvt_info.get("app_name")  # authoring app + version from the RVT header
                conv_version = conv_info.get("version")  # e.g. "18.0.0.0"
                stderr_tail = (ddc_failure.get("stderr") or "").strip()
                cause = ddc_failure.get("cause") or "unknown"

                if ext == ".rvt":
                    model.status = "needs_converter"

                    # Compose the message in pieces - every clause is added
                    # only when its underlying datum is non-empty so we never
                    # ship "File saved with None".
                    parts: list[str] = []
                    if cause == "converter_outdated":
                        # Lead with the specific, actionable diagnosis so
                        # the user sees the fix before the surrounding
                        # diagnostics.  The full stderr line still goes
                        # into metadata_ below for the support panel.
                        parts.append("Your installed converter is older than the platform expects.")
                    if rvt_app:
                        parts.append(
                            f"File saved with {rvt_app}"
                            + (f" (format {rvt_info['format']})." if rvt_info.get("format") else ".")
                        )
                    if conv_version:
                        parts.append(f"Installed RVT converter: {conv_version}.")
                    if cause == "converter_outdated":
                        parts.append("Open Settings → BIM Converters and click Reinstall to fetch the latest version.")
                    else:
                        parts.append(
                            "The converter produced no elements from this file. "
                            "Most common causes: the RVT was saved with a newer "
                            "format version than the converter supports, the file "
                            "is corrupt, or a converter dependency is missing."
                        )
                    if stderr_tail:
                        # Trim to a single line for the headline message;
                        # the full stderr tail goes into metadata_ below.
                        first_line = stderr_tail.splitlines()[0][:200]
                        if first_line:
                            parts.append(f"Converter said: {first_line}")
                    if cause != "converter_outdated":
                        parts.append(
                            "Try updating the RVT converter (Settings → BIM Converters → Reinstall) and clicking Retry."
                        )
                    model.error_message = " ".join(parts)

                    meta["error_code"] = "converter_outdated" if cause == "converter_outdated" else "ddc_failed"
                    meta["cause"] = cause
                    meta["converter_id"] = "rvt"
                    meta["install_endpoint"] = "/api/v1/takeoff/converters/rvt/install/"
                    # Structured diagnostic info for the frontend to render
                    # a dedicated "version mismatch" panel if it wants to.
                    meta["diagnostics"] = {
                        "rvt_info": rvt_info,
                        "converter_info": conv_info,
                        "reason": ddc_failure.get("reason"),
                        "cause": cause,
                        "exit_code": ddc_failure.get("exit_code"),
                        "stderr_tail": stderr_tail,
                    }
                else:
                    # Default disposition is a hard ``error`` - but a file that
                    # READ CLEANLY and simply has no physical products is NOT a
                    # processing failure (issue #197).  That case is downgraded
                    # to the graceful ``empty_model`` status below.  Genuine
                    # failures (unreadable STEP, converter/column mismatch,
                    # timeout) keep ``error``.
                    model.status = "error"
                    # Pick a specific, actionable message based on WHY the
                    # conversion produced no elements.  ``empty_reason`` is set
                    # by process_ifc_file:
                    #   spatial_only / no_products - the file parsed fine but
                    #       carried only spatial containers (project, site,
                    #       building, storey) and no physical products.
                    #   unclassified_only / all_filtered - rows came back that
                    #       we could not classify (column / converter mismatch).
                    #   no_entities - the STEP payload had no readable entities.
                    #   None - we never got readable rows at all.
                    empty_reason = result.get("empty_reason")
                    raw_row_count = int(result.get("raw_row_count") or 0)
                    # A converter timeout (common on very large models) is the
                    # real story when no readable rows came back at all.  When
                    # the parse DID yield a classification reason we trust that
                    # instead, because it describes the file we actually read.
                    if cause == "timeout" and not empty_reason:
                        model.error_message = (
                            "Converting this IFC timed out before any elements "
                            "were produced. This usually happens with very "
                            "large models. Try again on a machine with more "
                            "memory, split the federated model into smaller "
                            "files, or click Retry."
                        )
                        meta["error_code"] = "convert_timeout"
                    elif empty_reason in ("spatial_only", "no_products"):
                        # The IFC parsed successfully; it just carries no
                        # building products (only project/site/building/storey).
                        # This is a legitimate, fully-explained outcome - NOT a
                        # failure - so it gets the dedicated ``empty_model``
                        # status the frontend renders as a neutral, informational
                        # state with a Re-upload action, instead of a red error.
                        model.status = "empty_model"
                        model.error_message = (
                            "This IFC was read successfully but contains no "
                            "physical building elements. It only carries the "
                            "spatial structure (project, site, building, "
                            "storeys). Re-export it from your authoring tool "
                            "with the model objects included (walls, slabs, "
                            "columns, MEP, and so on), then upload again."
                        )
                        meta["error_code"] = "no_products"
                    elif empty_reason in ("unclassified_only", "all_filtered"):
                        model.error_message = (
                            "The converter returned rows for this IFC but none "
                            "could be matched to a building element type. This "
                            "usually means the file was exported in an "
                            "unexpected layout. Try re-exporting it as IFC 2x3 "
                            "or IFC 4 with base quantities enabled, then click "
                            "Retry."
                        )
                        meta["error_code"] = "unclassified_rows"
                    elif empty_reason == "no_entities":
                        model.error_message = (
                            "No model data could be read from this IFC file. "
                            "The file may be empty, truncated, or saved in an "
                            "unsupported layout. Re-export it from your "
                            "authoring tool and upload again."
                        )
                        meta["error_code"] = "no_entities"
                    else:
                        model.error_message = (
                            "No elements could be extracted from this IFC file. "
                            "Open it in your BIM authoring tool or an IFC viewer "
                            "to confirm it contains model objects, then "
                            "re-export and click Retry."
                        )
                        meta["error_code"] = "zero_elements"
                    if empty_reason:
                        meta["empty_reason"] = empty_reason
                    if raw_row_count:
                        meta["raw_row_count"] = raw_row_count
                    if stderr_tail or conv_version:
                        meta["diagnostics"] = {
                            "converter_info": conv_info,
                            "reason": ddc_failure.get("reason"),
                            "exit_code": ddc_failure.get("exit_code"),
                            "stderr_tail": stderr_tail,
                        }
                model.metadata_ = meta
                logger.warning(
                    "Background CAD processed but no elements found for model %s (status=%s, reason=%s)",
                    model_id,
                    model.status,
                    meta.get("error_code"),
                )

            await session.commit()

        # ── Mandatory validation pass (Import → Convert → VALIDATE → Store) ──
        # The canonical CAD data is now persisted as BIMElement rows. Run the
        # platform's validation engine over it and store a ValidationReport so
        # the import honours the first-class validation step instead of
        # skipping it. This reuses the existing BIMValidationService (universal
        # BIM rules: required properties, completeness, classification) and the
        # same ValidationReport table the on-demand "Validate" button writes -
        # the BIM detail view already loads the latest report for the model, so
        # no new surface is needed. Best-effort: a validation failure must never
        # turn a good import into a hard error, so it is fully guarded.
        if element_count > 0:
            await _run_import_validation(model_uuid)

        # Bake the streaming tileset now (and backfill GLB / Parquet if either
        # was skipped). The CAD path already wrote the GLB + Parquet, so those
        # steps are no-ops here; this proactively bakes the octree tiles that
        # otherwise only bake lazily on the first viewer open. Idempotent and
        # best-effort - it never turns a good import into an error.
        await _ensure_model_artifacts(project_id, model_id)

    except Exception as exc:
        logger.exception("Background CAD processing failed for model %s: %s", model_id, exc)
        try:
            async with async_session_factory() as session:
                model = (await session.execute(select(BIMModel).where(BIMModel.id == model_uuid))).scalar_one_or_none()
                if model is not None:
                    model.status = "error"
                    model.error_message = (
                        f"Processing failed: {exc}. Click Retry to try again, or "
                        f"contact support if this keeps happening."
                    )
                    meta = dict(model.metadata_ or {})
                    meta["error_code"] = "unexpected"
                    model.metadata_ = meta
                    await session.commit()
        except Exception as exc2:
            logger.exception("Failed to mark model %s as error: %s", model_id, exc2)


async def _generate_pdf_in_background(
    *,
    project_id: str,
    model_id: str,
    cad_storage_key: str,
    ext: str,
    model_name: str,
    user_id: str,
) -> None:
    """Run DDC PDF-only export for an existing model and link as a Document.

    Invoked from POST /{model_id}/generate-pdf-sheets/ via BackgroundTasks.
    Calls the DDC converter exactly once with a ``.pdf`` output target - no
    re-export of XLSX/DAE.  Silently skips when no converter is installed
    on the host (the upload itself stays usable; PDF export is opt-in).
    """
    import asyncio
    import subprocess
    import tempfile
    import uuid as _uuid
    from pathlib import Path as _Path

    from app.core.storage import get_storage_backend
    from app.database import async_session_factory

    converter_ext = ext.lstrip(".").lower()
    try:
        from app.modules.boq.cad_import import _converter_subprocess_env, find_converter
    except ImportError:
        logger.warning("PDF generation skipped - cad_import not available")
        return

    converter = find_converter(converter_ext)
    if not converter:
        logger.info(
            "PDF generation skipped - %s converter not installed",
            converter_ext.upper(),
        )
        return

    try:
        content = await get_storage_backend().get(cad_storage_key)

        with tempfile.TemporaryDirectory(prefix="oe-bim-pdf-") as _tmp_str:
            _tmp_dir = _Path(_tmp_str)
            _tmp_cad_path = _tmp_dir / f"original{ext}"
            await asyncio.to_thread(_tmp_cad_path.write_bytes, content)

            pdf_target = (_tmp_dir / "sheets.pdf").resolve()

            def _run_pdf() -> tuple[int, bytes]:
                proc = subprocess.run(
                    [str(converter), str(_tmp_cad_path.resolve()), str(pdf_target)],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    cwd=str(converter.parent),
                    # Linux no-root converter needs its bundled DDC SDK libs on
                    # LD_LIBRARY_PATH; None (inherit) on Windows/macOS.
                    env=_converter_subprocess_env(converter),
                    input=b"\n",
                    timeout=900,
                )
                return proc.returncode, proc.stderr

            try:
                rc, stderr = await asyncio.to_thread(_run_pdf)
            except subprocess.TimeoutExpired:
                logger.warning("DDC PDF generation timed out for model %s", model_id)
                return

            if rc != 0 or not pdf_target.is_file() or pdf_target.stat().st_size < 1000:
                logger.warning(
                    "DDC PDF generation failed for model %s (rc=%d, stderr=%s)",
                    model_id,
                    rc,
                    stderr.decode(errors="replace")[:200] if stderr else "",
                )
                return

            pdf_bytes = await asyncio.to_thread(pdf_target.read_bytes)
            pdf_storage_key = await bim_file_storage.save_geometry(
                project_id=project_id,
                model_id=model_id,
                ext=".pdf",
                content=pdf_bytes,
            )

            # ``uploaded_by``, not ``created_by``: Document has no
            # ``created_by`` column - Base carries id/created_at/updated_at and
            # nothing else - so the keyword this used to pass raised TypeError
            # on every call, before a single byte reached the database. The
            # ``except Exception`` that used to stand below caught it and
            # logged "linkage failed", which reads like the database declining
            # a row. It had declined nothing, and it had never been asked:
            # keyword and function arrived in the same commit, so the sheets
            # PDF has never once reached the documents hub.
            #
            # Narrowing to SQLAlchemyError does not make the next such defect
            # loud: this is a background task and the whole body sits under an
            # ``except Exception`` at the end of the function, so a TypeError
            # here still ends quietly. What it buys is a traceback naming this
            # line instead of a one-line warning naming the database, which is
            # the difference between a defect somebody can act on and one that
            # survives for as long as this one did. Being reached at all needs
            # a converter binary, so no test on a clean machine executes it -
            # the guard is ``tests/unit/test_bim_document_cross_link_fields.py``,
            # which reads the keywords out of this file's syntax.
            try:
                from app.modules.documents.models import Document as DocModel

                async with async_session_factory() as session:
                    pdf_doc = DocModel(
                        project_id=_uuid.UUID(project_id),
                        name=f"{model_name or 'BIM Model'} - Sheets (PDF)",
                        category="drawing",
                        file_path=pdf_storage_key,
                        file_size=len(pdf_bytes),
                        mime_type="application/pdf",
                        tags=["bim", "sheets", "auto-generated", converter_ext],
                        uploaded_by=user_id or "",
                    )
                    session.add(pdf_doc)
                    await session.commit()
                    logger.info(
                        "PDF sheets saved as Document for model %s: %s (%d bytes)",
                        model_id,
                        pdf_storage_key,
                        len(pdf_bytes),
                    )
            except SQLAlchemyError:
                logger.exception(
                    "Failed to cross-link the sheets PDF of BIM model %s (project %s) into the documents hub",
                    model_id,
                    project_id,
                )

    except Exception as exc:
        logger.exception("PDF generation failed for model %s: %s", model_id, exc)


@router.post("/upload-cad/", status_code=201)
async def upload_cad_file(
    background_tasks: BackgroundTasks,
    project_id: str = Query(..., description="Project UUID"),
    name: str = Query(default="", max_length=255),
    discipline: str = Query(default="architecture", max_length=50),
    conversion_depth: str = Query(
        default="standard",
        description=(
            "DDC conversion depth: 'standard' (~15 basic columns, fastest),"
            " 'medium' (DDC standard + full property promotion, ~900 columns),"
            " or 'complete' (all RVT parameters, ~1000+ columns, slowest)"
        ),
    ),
    file: UploadFile = File(..., description="CAD file (RVT, IFC, DWG, DGN, FBX, OBJ, 3DS)"),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict:
    """Upload a raw CAD file for background processing.

    The file is stored on disk at ``data/bim/{project_id}/{model_id}/original.{ext}``
    and a BIMModel record is created with status="processing". A real CAD converter
    service would pick it up asynchronously; for now the model stays in processing state.

    Accepted extensions: .rvt, .ifc, .dwg, .dgn, .fbx, .obj, .3ds
    """
    # --- Verify project access (IDOR guard) ---
    try:
        project_uuid = uuid.UUID(project_id)
    except (ValueError, TypeError) as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Invalid project_id: {exc}",
        ) from exc
    await _verify_project_access(service.session, project_uuid, user_id or "")

    allowed, _ = upload_limiter.is_allowed(str(user_id or "anon"))
    if not allowed:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many uploads. Please wait a moment and try again.",
            headers={"Retry-After": "60"},
        )

    filename = (file.filename or "").strip()
    if not filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided.",
        )

    ext = pathlib.Path(filename).suffix.lower()
    # Geometry-only 3D mesh formats (OBJ/STL/glTF/GLB/DAE/FBX/PLY/3DS/…) are
    # not server-convertible CAD - the BIM Hub 3D uploader reads them directly
    # in the browser and posts a normalized GLB through the data+geometry
    # upload. Short-circuit them here with a friendly pointer instead of the
    # generic "unsupported" error (or a confusing magic-byte mismatch for the
    # few whose extension is in the CAD set), so nothing is stored as a dead,
    # never-converting model.
    if is_mesh_geometry_ext(ext):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=mesh_upload_hint(ext, filename),
        )
    if ext not in _ALLOWED_CAD_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(f"Unsupported file type '{ext}'. Accepted: {', '.join(sorted(_ALLOWED_CAD_EXTENSIONS))}"),
        )

    # Stream the upload to a temp file in 1 MB chunks instead of buffering
    # the whole body in memory.  A 500 MB IFC used to cost ~500 MB of heap
    # in the request handler - on the 2 GB-RAM VPS, two concurrent uploads
    # were enough to OOM the process.  ``StreamedUpload`` exposes:
    #   - ``upload.path``    - the spooled temp file
    #   - ``upload.size``    - bytes written
    #   - ``upload.head``    - first 64 bytes for magic-byte validation
    # Storage's ``put_stream`` then ``rename(2)``s the file into place
    # (single syscall on same-FS local backend; ``upload_fileobj`` with
    # multipart on S3) - zero additional memory pressure.
    from app.core.upload_streaming import stream_upload_to_temp

    async with stream_upload_to_temp(file, suffix=ext) as upload:
        if upload.size == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Uploaded file is empty.",
            )

        # Preflight: when the required converter binary isn't installed on
        # this server, **persist** the upload and create a placeholder model
        # so the user doesn't have to re-upload after running the install.
        # The response is HTTP 202 Accepted (the request is good - we'll
        # finish processing later) rather than 201 Created (we haven't
        # created any geometry yet).  ``Retry-After`` and a ``Link`` header
        # point the client at the converter-install endpoint and the model
        # row that will be re-processed once the binary lands.  Frontend
        # dispatches on the ``status`` field in the body - see
        # ``BIMCadUploadResponse`` in ``frontend/src/features/bim/api.ts``.
        if ext in _NEEDS_CONVERTER_EXTS:
            from app.modules.boq.cad_import import find_converter

            if find_converter(ext.lstrip(".")) is None:
                display_name = (name or pathlib.Path(filename).stem).strip() or filename
                from app.modules.bim_hub.schemas import BIMModelCreate

                # The row is created before the file is saved, so the blob can
                # be keyed by the model's own id. It used to be the other way
                # round: a UUID was minted to key the blob with, the row then
                # took a different primary key from the database, and the
                # minted one was what the response and the reprocess link were
                # built from. That broke the feature at both ends. The response
                # named a model no row carried, and ``retry`` derives the blob
                # key from the model id, so even with the right id in hand it
                # looked for the file under a name nothing had written and
                # answered "original CAD file is no longer available". Clicking
                # Re-process is exactly what the error message on this row
                # tells people to do.
                #
                # NB: ``error_message`` is set via a follow-up update because
                # ``BIMModelCreate`` doesn't expose that field - it lives on
                # ``BIMModelUpdate`` so freshly-created records start clean.
                # The kwarg here MUST be ``user_id=`` (the service signature) -
                # passing ``created_by=`` raised a TypeError on every .rvt /
                # .ifc upload that hit the missing-converter path.
                pending_model = await service.create_model(
                    BIMModelCreate(
                        project_id=project_uuid,
                        name=display_name,
                        discipline=discipline,
                        model_format=ext.lstrip("."),
                        status="needs_converter",
                    ),
                    user_id=user_id,
                )
                saved_cad_key = await bim_file_storage.save_original_cad_from_path(
                    project_uuid,
                    pending_model.id,
                    ext,
                    upload.path,
                    size=upload.size,
                )
                await service.update_model(
                    pending_model.id,
                    BIMModelUpdate(
                        canonical_file_path=saved_cad_key,
                        error_message=(
                            f"{ext.upper().lstrip('.')} converter not installed - "
                            f"install it from the BIM converter banner, then "
                            f"click Re-process on this model."
                        ),
                    ),
                )

                logger.info(
                    "Saved %s upload pending converter - model=%s, key=%s, %d bytes",
                    ext,
                    pending_model.id,
                    saved_cad_key,
                    upload.size,
                )

                install_endpoint = f"/api/v1/takeoff/converters/{ext.lstrip('.')}/install/"
                return JSONResponse(
                    status_code=status.HTTP_202_ACCEPTED,
                    content={
                        "status": "converter_required",
                        "format": ext.lstrip("."),
                        "converter_id": ext.lstrip("."),
                        "message": (
                            f"{ext.upper().lstrip('.')} files require the "
                            f"{ext.upper().lstrip('.')} converter, which is not "
                            f"installed on this server. Your file has been "
                            f"saved - install the converter and click "
                            f"Re-process on the model card to finish the upload."
                        ),
                        "install_endpoint": install_endpoint,
                        "model_id": str(pending_model.id),
                        "name": display_name,
                        "file_size": upload.size,
                        "element_count": 0,
                        "error_message": None,
                    },
                    headers={
                        "Retry-After": "60",
                        "Link": (
                            f'<{install_endpoint}>; rel="install-converter", '
                            f"</api/v1/bim_hub/{pending_model.id}/retry/>; "
                            f'rel="reprocess-model"'
                        ),
                    },
                )

        # Magic-byte validation - filename extensions are attacker-controlled
        # and we proceed to hand this file to a CAD converter that can be
        # exploited by unexpected formats. Reject anything that doesn't look
        # like one of our accepted CAD/BIM containers.  The streamed-upload
        # helper has already kept the first 64 bytes around for us.
        from app.core.file_signature import (
            ALLOWED_CAD_TYPES,
            FileSignatureMismatch,
        )
        from app.core.file_signature import (
            require as _require_sig,
        )

        try:
            _require_sig(
                upload.head,
                ALLOWED_CAD_TYPES,
                filename=filename,
            )
        except FileSignatureMismatch as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc

        # Auto-fill model name from filename
        model_name = name or pathlib.Path(filename).stem

        # Determine model format from extension (strip the dot)
        model_format = ext.lstrip(".")

        # Create BIM model record with processing status
        from app.modules.bim_hub.schemas import BIMModelCreate

        model_data = BIMModelCreate(
            project_id=uuid.UUID(project_id),
            name=model_name,
            discipline=discipline,
            model_format=model_format,
            status="processing",
        )
        model = await service.create_model(model_data, user_id=user_id)
        model_id = model.id

        # Save CAD file via the configured storage backend - returns the
        # storage key that the Documents hub cross-link and downstream
        # diagnostics use to refer back to the stored blob.  Streaming
        # variant: the backend renames the temp file into place rather
        # than buffering its content.
        saved_cad_key = await bim_file_storage.save_original_cad_from_path(
            project_id=project_id,
            model_id=str(model_id),
            ext=ext,
            src_path=upload.path,
            size=upload.size,
        )

        logger.info(
            "CAD file uploaded: %s (%s, %d bytes) -> model %s (key=%s)",
            filename,
            ext,
            upload.size,
            model_id,
            saved_cad_key,
        )

        # Cross-link: create Document record so BIM files appear in Documents hub.
        # Uses the ORM model directly (NOT raw SQL) so timestamps + defaults are
        # filled by SQLAlchemy / Base mixin and the row stays in sync with the
        # rest of the documents module if its schema evolves.
        #
        # The cross-link is convenience-only, the BIM model itself is already
        # saved by the time we get here, so a failure here must not fail the
        # upload. Making that true takes the SAVEPOINT below, not just the
        # `except`: a failed `flush()` marks the whole session for rollback, so
        # catching the error and carrying on leaves every later statement on
        # this request raising about a transaction that cannot continue. The
        # request still dies, several frames from here, describing something
        # other than what went wrong. With `begin_nested()` only the cross-link
        # is rolled back and the session survives, which is what the comment
        # above always claimed.
        try:
            from app.modules.documents.models import Document

            doc = Document(
                project_id=uuid.UUID(project_id),
                name=filename,
                description=f"BIM model: {model_name}",
                category="drawing",
                file_size=upload.size,
                mime_type=f"application/{model_format}",
                file_path=saved_cad_key,
                version=1,
                uploaded_by=user_id or "",
                tags=["bim", model_format, discipline],
                metadata_={
                    "source_module": "bim_hub",
                    "source_id": str(model_id),
                },
            )
            async with service.session.begin_nested():
                service.session.add(doc)
                await service.session.flush()
            logger.info("Cross-linked BIM model %s → document %s", model_id, doc.id)
        except SQLAlchemyError:
            # Narrower than `except Exception` on purpose: a NameError or a
            # TypeError in the block above is a defect in this code, not a
            # database saying no, and swallowing those hides it forever.
            # Logged with both ids so the orphaned model can be found.
            logger.exception(
                "Failed to cross-link BIM model %s (project %s) into the documents hub",
                model_id,
                project_id,
            )

    # Schedule processing OUT of the request path: the upload endpoint now
    # returns 201 + status="processing" in milliseconds, and the actual DDC
    # conversion + element/geometry persistence happens in a background task.
    # This eliminates the multi-minute synchronous block that used to drop
    # the connection on slow conversions ("Cannot connect to server" in the
    # frontend).  The frontend already polls GET /{model_id}/ - model.status
    # transitions from "processing" → "ready" / "error" / "needs_converter"
    # automatically once the worker finishes.
    final_status = "processing"
    element_count = 0

    # Commit BEFORE scheduling the background task: the worker opens its own
    # async session and looks the model up by id, so the row must already be
    # durably visible to other connections.  Without this explicit commit the
    # worker raced the request-scope dependency teardown and saw "model
    # vanished mid-conversion" intermittently.
    await service.session.commit()

    processable = ext in (".ifc", ".rvt")
    if processable:
        background_tasks.add_task(
            _process_cad_in_background,
            project_id=project_id,
            model_id=str(model_id),
            cad_storage_key=saved_cad_key,
            ext=ext,
            conversion_depth=conversion_depth,
        )
        logger.info(
            "CAD upload accepted, processing scheduled in background: %s → model %s",
            filename,
            model_id,
        )
    else:
        # Non-processable format (DWG, DGN, FBX, etc.) - needs converter
        model.status = "needs_converter"
        model.error_message = (
            f"{ext.upper().lstrip('.')} files require an external converter. Convert to IFC first, then re-upload."
        )
        await service.session.flush()
        final_status = "needs_converter"

    return {
        "model_id": str(model_id),
        "name": model_name,
        "format": model_format,
        "file_size": upload.size,
        "status": final_status,
        "element_count": element_count,
        "error_message": model.error_message,
        "geometry_type": (model.metadata_ or {}).get("geometry_type", "unknown"),
        "converter_id": ext.lstrip(".") if final_status == "needs_converter" else None,
        "install_endpoint": (
            f"/api/v1/takeoff/converters/{ext.lstrip('.')}/install/" if final_status == "needs_converter" else None
        ),
    }


@router.post("/models/from-document/", status_code=201)
async def create_model_from_document(
    body: CreateModelFromDocumentRequest,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict:
    """Create a BIM model from a file already uploaded to Project Documents.

    Files uploaded straight from the Documents hub are stored but never
    converted, so opening one in the BIM viewer used to find no model and ask
    the user to upload the file a second time (issue #273). This turns an
    existing document into a BIM model on demand, reusing the same conversion
    pipeline as a direct CAD upload.

    Idempotent: a document that is already linked to a model returns that model
    instead of converting the same file twice.
    """
    from pathlib import Path as _Path

    from app.modules.bim_hub.schemas import BIMModelCreate
    from app.modules.documents.repository import DocumentRepository

    doc = await DocumentRepository(service.session).get_by_id(body.document_id)
    if doc is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Document not found",
        )

    # IDOR guard - same owner/admin/member rule as every other BIM route.
    await _verify_project_access(service.session, doc.project_id, user_id or "")

    # Idempotency - if this document already produced a BIM model, return it
    # rather than converting the same file twice (e.g. a double-click, or a
    # file that was uploaded through the BIM hub in the first place).
    meta = dict(doc.metadata_ or {})
    existing_id = meta.get("source_id") if meta.get("source_module") == "bim_hub" else None
    if existing_id:
        try:
            existing = await service.get_model(uuid.UUID(str(existing_id)))
        except Exception:  # noqa: BLE001 - stale link, fall through and recreate
            existing = None
        if existing is not None:
            return {
                "model_id": str(existing.id),
                "name": existing.name,
                "format": existing.model_format,
                "status": existing.status,
                "element_count": getattr(existing, "element_count", 0) or 0,
                "error_message": existing.error_message,
                "already_existed": True,
            }

    ext = _Path(doc.name or "").suffix.lower()
    # Geometry-only mesh formats are read in-browser by the BIM Hub 3D
    # uploader, not converted on the server - point the user there instead of
    # creating a model that could never render (mirrors the upload-cad guard).
    if is_mesh_geometry_ext(ext):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=mesh_upload_hint(ext, doc.name),
        )
    if ext not in _ALLOWED_CAD_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"'{doc.name}' is not a convertible BIM/CAD file. Accepted: "
                f"{', '.join(sorted(_ALLOWED_CAD_EXTENSIONS))}"
            ),
        )

    # Read the stored document. The Documents hub writes uploads to a local
    # path and already buffers the whole file on upload, so reading it back
    # here is consistent with that module's own memory model.
    try:
        content = _Path(doc.file_path).read_bytes()
    except OSError as exc:
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail="The stored document file is no longer available.",
        ) from exc
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="The document file is empty.",
        )

    model_name = (body.name or _Path(doc.name).stem).strip() or doc.name
    model_format = ext.lstrip(".")

    model = await service.create_model(
        BIMModelCreate(
            project_id=doc.project_id,
            name=model_name,
            discipline=body.discipline,
            model_format=model_format,
            status="processing",
        ),
        user_id=user_id,
    )

    saved_cad_key = await bim_file_storage.save_original_cad(doc.project_id, model.id, ext, content)

    # Link the document to the new model so a later "Open in BIM viewer"
    # returns this model rather than converting again. Assigning a fresh dict
    # (not mutating in place) so SQLAlchemy marks the JSON column dirty.
    meta.update({"source_module": "bim_hub", "source_id": str(model.id)})
    doc.metadata_ = meta
    service.session.add(doc)

    # Capture the scalars we return / log BEFORE committing - after commit the
    # ORM attributes may be expired and a lazy refresh on an async session
    # raises MissingGreenlet.
    model_id = model.id
    project_id_str = str(doc.project_id)
    doc_id_for_log = doc.id

    # Only IFC/RVT are processed inline by the background worker; the other
    # accepted formats are stored and flagged for an external converter, the
    # same way the direct CAD upload endpoint handles them.
    processable = ext in (".ifc", ".rvt")
    if processable:
        final_status = "processing"
        final_error: str | None = None
    else:
        final_status = "needs_converter"
        final_error = (
            f"{ext.upper().lstrip('.')} files require an external converter. Convert to IFC first, then re-upload."
        )
        model.status = final_status
        model.error_message = final_error

    # Commit before scheduling the background task: the worker opens its own
    # session and looks the model up by id, so the row must be durably visible.
    await service.session.commit()

    if processable:
        background_tasks.add_task(
            _process_cad_in_background,
            project_id=project_id_str,
            model_id=str(model_id),
            cad_storage_key=saved_cad_key,
            ext=ext,
            conversion_depth=body.conversion_depth,
        )
        logger.info(
            "BIM model %s created from document %s, conversion scheduled",
            model_id,
            doc_id_for_log,
        )

    return {
        "model_id": str(model_id),
        "name": model_name,
        "format": model_format,
        "status": final_status,
        "element_count": 0,
        "error_message": final_error,
        "already_existed": False,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Optional post-upload: generate PDF sheets for an existing model
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/{model_id}/generate-pdf-sheets/", status_code=202)
async def generate_pdf_sheets(
    model_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Schedule PDF-sheets export for an existing BIM model.

    Runs the DDC converter once with a ``.pdf`` output target - no XLSX/DAE
    re-export, only the sheets PDF.  The PDF is saved as a Document linked
    to the project once the worker finishes.

    Returns immediately; the caller does not wait for the export to finish.
    Frontend can detect completion by polling the project's documents list.
    """
    from app.modules.bim_hub import file_storage as _bim_storage

    model = await _verify_model_access(service, model_id, user_id or "")

    model_format = (model.model_format or "").lower()
    if not model_format:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Model has no format recorded - cannot regenerate sheets.",
        )

    ext = "." + model_format.lstrip(".")
    cad_storage_key = _bim_storage.original_cad_key(
        project_id=model.project_id,
        model_id=model_id,
        ext=ext,
    )
    backend_store = _bim_storage._backend()
    if not await backend_store.exists(cad_storage_key):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Original CAD file is no longer available - re-upload the model.",
        )

    background_tasks.add_task(
        _generate_pdf_in_background,
        project_id=str(model.project_id),
        model_id=str(model_id),
        cad_storage_key=cad_storage_key,
        ext=ext,
        model_name=model.name or "BIM Model",
        user_id=user_id or "",
    )

    return {
        "status": "scheduled",
        "model_id": str(model_id),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Retry conversion - re-runs the background CAD processor for a model that
# previously failed (status="error" / "needs_converter"). Useful when the
# user installs a missing converter after upload, or the original failure
# was transient (network blip, OOM during a parallel upload burst).
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/{model_id}/retry/", status_code=202)
async def retry_model_processing(
    model_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Re-schedule background DDC conversion for a previously-failed model.

    Resets ``status`` to ``processing``, clears ``error_message``, and re-
    invokes :func:`_process_cad_in_background` against the same original CAD
    blob.  Returns 202 immediately - the frontend already polls
    ``GET /{model_id}/`` and will transition the UI when the worker finishes.

    Refuses to retry models that:
        * are already ``ready`` (no need),
        * are already ``processing`` (a worker is in flight),
        * have no original CAD blob recorded (re-upload required).
    """
    from app.modules.bim_hub import file_storage as _bim_storage

    model = await _verify_model_access(service, model_id, user_id or "")

    if model.status == "ready":
        return {
            "status": "noop",
            "model_id": str(model_id),
            "message": "Model is already ready - nothing to retry.",
        }
    if model.status == "processing":
        return {
            "status": "noop",
            "model_id": str(model_id),
            "message": "Model is already being processed.",
        }

    model_format = (model.model_format or "").lower()
    if not model_format:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Model has no format recorded - cannot retry.",
        )

    ext = "." + model_format.lstrip(".")
    cad_storage_key = _bim_storage.original_cad_key(
        project_id=model.project_id,
        model_id=model_id,
        ext=ext,
    )
    backend_store = _bim_storage._backend()
    if not await backend_store.exists(cad_storage_key):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=("Original CAD file is no longer available - re-upload the model to retry."),
        )

    # Clear previous error state before re-scheduling so the frontend's
    # polling immediately reflects "processing" once the retry kicks in.
    model.status = "processing"
    model.error_message = None
    meta = dict(model.metadata_ or {})
    for _k in ("error_code", "install_endpoint"):
        meta.pop(_k, None)
    model.metadata_ = meta
    await service.session.commit()

    background_tasks.add_task(
        _process_cad_in_background,
        project_id=str(model.project_id),
        model_id=str(model_id),
        cad_storage_key=cad_storage_key,
        ext=ext,
        conversion_depth="standard",
    )
    logger.info("Retry scheduled for model %s", model_id)

    return {
        "status": "scheduled",
        "model_id": str(model_id),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Parquet sidecar status & retry (Task #155)
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/parquet-status/")
async def get_parquet_status(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Return the last-known Parquet sidecar status for *model_id*.

    Surfaces the silent-failure state stamped onto ``model.metadata_`` by
    the background ingester so the UI can show a "Parquet sidecar failed -
    retry?" affordance instead of silently serving an empty dataframe.
    """
    model = await _verify_model_access(service, model_id, user_id or "")
    meta = model.metadata_ or {}
    return {
        "model_id": str(model_id),
        "status": meta.get("parquet_status"),
        "error": meta.get("parquet_error"),
        "attempted_at": meta.get("parquet_attempted_at"),
        "retry_endpoint": f"/api/v1/bim-hub/models/{model_id}/parquet/retry/",
    }


@router.post("/models/{model_id}/parquet/retry/", status_code=202)
async def retry_parquet_write(
    model_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Re-run the Parquet sidecar write from existing element rows in the DB.

    Cheap recovery for the case where the original DDC conversion succeeded
    (elements landed in ``oe_bim_element``) but the Parquet write failed
    (disk full, permission denied, pyarrow crash). Does NOT re-run DDC.
    """
    from app.modules.bim_hub.dataframe_store import write_dataframe
    from app.modules.bim_hub.models import BIMElement

    model = await _verify_model_access(service, model_id, user_id or "")

    # Pull the rows back out of the DB - same projection the converter
    # would have produced for the Parquet write (one dict per element).
    rows_q = await service.session.execute(select(BIMElement).where(BIMElement.model_id == model_id))
    elements = rows_q.scalars().all()
    if not elements:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=("Model has no elements in the database - cannot retry Parquet write. Re-upload the model instead."),
        )

    rows: list[dict[str, Any]] = []
    for el in elements:
        row: dict[str, Any] = {
            "stable_id": el.stable_id,
            "element_type": el.element_type,
            "name": el.name,
            "storey": el.storey,
            "discipline": el.discipline,
        }
        # Flatten properties + quantities into the row dict so the Parquet
        # column set matches the original DDC shape closely enough for the
        # dataframe endpoints to work.
        if isinstance(el.properties, dict):
            row.update(el.properties)
        if isinstance(el.quantities, dict):
            row.update(el.quantities)
        rows.append(row)

    import asyncio as _asyncio
    from datetime import UTC as _UTC
    from datetime import datetime as _dt

    try:
        await _asyncio.to_thread(
            write_dataframe,
            project_id=str(model.project_id),
            model_id=str(model_id),
            rows=rows,
        )
        new_status, new_error = "ok", None
    except Exception as exc:
        new_status, new_error = _surface_parquet_failure(
            exc=exc,
            project_id=str(model.project_id),
            model_id=str(model_id),
            model_format=model.model_format or "",
            row_count=len(rows),
        )

    meta = dict(model.metadata_ or {})
    meta["parquet_status"] = new_status
    meta["parquet_error"] = new_error
    meta["parquet_attempted_at"] = _dt.now(_UTC).isoformat()
    model.metadata_ = meta
    await service.session.commit()

    return {
        "model_id": str(model_id),
        "status": new_status,
        "error": new_error,
        "rows_attempted": len(rows),
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Geometry file serving
# ═══════════════════════════════════════════════════════════════════════════════


@router.head("/models/{model_id}/geometry/", response_model=None, include_in_schema=False)
@router.get("/models/{model_id}/geometry/", response_model=None)
async def get_model_geometry(
    model_id: uuid.UUID,
    token: str | None = Query(
        default=None,
        description="JWT access token (alternative to Authorization header for static loaders)",
    ),
    fmt: str | None = Query(
        default=None,
        description="Force a specific geometry format: 'dae' or 'glb'. "
        "When omitted, the server returns GLB (preferred) with DAE fallback.",
    ),
    authorization: str | None = Header(default=None),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse | RedirectResponse:
    """Serve the COLLADA/DAE geometry file for the 3D viewer.

    Auth: accepts either an Authorization header OR a ``?token=...`` query
    parameter. The query param exists because Three.js ColladaLoader cannot
    set custom headers - without this fallback the viewer would 401.

    The geometry blob is resolved through :mod:`app.modules.bim_hub.file_storage`
    so both the local filesystem and S3 backends work transparently.  For S3
    we redirect to a short-lived presigned URL; for the local backend we
    stream the bytes directly through the route.
    """
    # Per-request correlation ID - surfaced in the X-Request-Id response
    # header AND embedded in every structured-error payload so a user who
    # ships a screenshot to support can be located in server logs in one
    # grep. UUID4 keeps it non-PII (no info about the user, project, or
    # file).  We generate locally rather than relying on a middleware so
    # the value is identical between log line and HTTP response.
    request_id = str(uuid.uuid4())

    # Validate the token (header or query). ColladaLoader can't set headers,
    # so we accept ?token=<jwt> as an alternative auth mechanism.
    from app.config import get_settings
    from app.dependencies import decode_access_token, verify_user_exists_and_active

    auth_token: str | None = token
    if not auth_token and authorization and authorization.lower().startswith("bearer "):
        auth_token = authorization[7:]

    if not auth_token:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "auth_missing",
                "category": "authentication",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "Missing authentication token.",
                "remediation": (
                    "Refresh the page to renew your login session. If you "
                    "were idle for a long time the access token may have "
                    "expired silently."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    try:
        payload = decode_access_token(auth_token, get_settings())
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "auth_invalid",
                "category": "authentication",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "Authentication token is invalid or expired.",
                "remediation": (
                    "Log out and log back in to obtain a fresh token. If "
                    "the problem persists, your account may have been "
                    "deactivated - contact support."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    # BUG-323: forged tokens with a fake UUID must not authenticate here
    # either. Re-hydrate against the DB and replace self-asserted role /
    # permissions with canonical state before any authorization check.
    db_user = await verify_user_exists_and_active(
        payload["sub"],
        issued_at=payload.get("iat"),
        session_id=payload.get("sid"),
    )
    from app.core.permissions import permission_registry

    payload["role"] = db_user.role
    payload["permissions"] = permission_registry.get_role_permissions(db_user.role)

    # Check the token-bearer actually has bim.read before we load data.
    token_role = payload.get("role", "")
    token_perms: list[str] = payload.get("permissions", [])
    if token_role != "admin" and "bim.read" not in token_perms:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "permission_denied",
                "category": "authorization",
                "request_id": request_id,
                "model_id": str(model_id),
                "required_permission": "bim.read",
                "message": "Your account lacks permission to view BIM models.",
                "remediation": (
                    "Ask a project administrator to grant you the 'bim.read' "
                    "permission, or to assign you a role (Estimator / "
                    "Manager / Admin) that includes it."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    model = await service.get_model(model_id)
    if model is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "model_not_found",
                "category": "not_found",
                "request_id": request_id,
                "model_id": str(model_id),
                "message": "This BIM model has been deleted or never existed.",
                "remediation": (
                    "Go back to the project's BIM tab and pick a model from "
                    "the list. If you reached this page from a saved link, "
                    "the model may have been removed by a teammate."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    # IDOR guard: verify the caller may access the project this model belongs to
    # (owner, admin, or shared team-member).
    token_user_id = str(payload.get("sub") or "")
    await _verify_project_access(service.session, model.project_id, token_user_id)

    project_id = str(model.project_id)

    # Resolve the geometry blob through the storage backend.
    # When ?fmt=dae is passed, force DAE format (useful when GLB has
    # scrambled node names from an older trimesh conversion).
    if fmt and fmt.lower() == "dae":
        found = await bim_file_storage.find_geometry_key(project_id, model_id, prefer_ext=".dae")
    else:
        found = await bim_file_storage.find_geometry_key(project_id, model_id)
    if found is not None:
        key, ext = found
        media_type = bim_file_storage.GEOMETRY_MEDIA_TYPES.get(ext, "application/octet-stream")
        cache_headers = {
            # No caching - geometry may be re-generated with patched node names.
            "Cache-Control": "no-store, no-cache, must-revalidate",
        }

        # Prefer a presigned URL so the browser fetches directly from the
        # bucket (S3).  Local backend returns None → fall back to streaming.
        presigned = bim_file_storage.presigned_geometry_url(key)
        if presigned:
            return RedirectResponse(url=presigned, status_code=307)

        # Resolve a local filesystem path for the blob when the backend keeps
        # it on disk (the default local backend). Serving straight from disk
        # via FileResponse streams with HTTP Range support and ZERO in-memory
        # copy, instead of reading the whole GLB into RAM and then holding a
        # SECOND full gzip copy. A large HVAC RVT converts to a 100+ MB GLB, so
        # the old read-all-then-gzip path peaked at ~2x the file per request
        # and OOM'd the 1-CPU production box (issue #291). S3-style backends
        # already returned a presigned redirect above; a backend that exposes
        # no local path falls through to the in-memory byte path below.
        from app.core.storage import get_storage_backend

        backend = get_storage_backend()
        disk_path = backend.local_path(key)

        if disk_path is not None:
            # Bounded head read: _quick_validate_geometry_bytes only inspects
            # the first ~16 KB (GLB magic, COLLADA root, glTF asset), so we
            # never page the full blob in just to validate it. The true size
            # comes from stat() for the failure diagnostic below.
            import asyncio

            def _read_geometry_head(p: pathlib.Path) -> tuple[bytes, int]:
                with p.open("rb") as fh:
                    head = fh.read(_GEOMETRY_HEAD_PROBE_BYTES)
                return head, p.stat().st_size

            _geo_bytes, _geo_size = await asyncio.to_thread(_read_geometry_head, disk_path)
        else:
            # Backend has no local path (remote S3 without presigning, or a
            # community backend): read the blob once. The response side below
            # still avoids ever holding a second full copy.
            _geo_bytes = await backend.get(key)
            _geo_size = len(_geo_bytes)

        # Serve-time integrity check - closes the gap where geometry written
        # by an older converter (before _validate_geometry_file existed on
        # ingest) keeps streaming bad bytes to the viewer. The browser
        # surfaces this as an opaque "Cannot read properties of undefined
        # (reading 'getAttribute')" deep inside Three.js. We re-check the
        # first ~4 KB of the blob and 422 with a precise diagnostic if it
        # doesn't match the format the extension promises. On the local path
        # this inspects the bounded head, which is byte-for-byte identical to
        # the head of the full blob for every check the validator performs.
        ok_serve, reason_serve = _quick_validate_geometry_bytes(_geo_bytes, ext)
        if not ok_serve:
            # Build a structured diagnostic payload. We deliberately limit
            # what we expose to: (a) the first 8 bytes of the file as hex
            # + ASCII (universally safe - magic bytes don't carry PII),
            # (b) total size in bytes, (c) the parser reason, (d) the
            # stored extension, (e) what we expected. NO actual user-data
            # bytes or filenames are leaked. Frontend renders this verbatim
            # in the BIM viewer error panel so users can give actionable
            # detail to support without revealing the file contents.
            head_bytes = _geo_bytes[:8]
            head_hex = " ".join(f"{b:02x}" for b in head_bytes)
            head_ascii = "".join(chr(b) if 0x20 <= b < 0x7F else "." for b in head_bytes)
            # Surface the first identifiable XML root tag for the common
            # "stored DAE turned out to be IFC-XML / gbXML / HTML 404 page"
            # failure mode - gives support a one-glance diagnosis.
            first_tag: str | None = None
            if ext.lower() == ".dae":
                import re as _re_diag

                try:
                    _head_text = _geo_bytes[:4096].decode("utf-8", errors="replace")
                    _m = _re_diag.search(r"<([a-zA-Z_:][\w:.-]{0,40})", _head_text)
                    if _m:
                        first_tag = f"<{_m.group(1)}>"
                except Exception:  # pragma: no cover - replace can't raise
                    first_tag = None
            expected_signature = {
                ".glb": "b'glTF' magic + version 2",
                ".dae": "<COLLADA> root tag within first 4 KB",
                ".gltf": "JSON object with required 'asset' key",
            }.get(ext.lower(), f"valid {ext} payload")
            # Categorise reason into a plain-language "cause" the UI can
            # show without the user having to read parser jargon. This is
            # the single biggest lever for end-user understanding: instead
            # of "DAE has no <COLLADA> root in first 4 KB (first tag found:
            # <html>)" they see "The stored file is an HTML page, not a 3D
            # model - the converter probably crashed and saved an error
            # page by mistake."
            reason_lower = reason_serve.lower()
            if "empty buffer" in reason_lower:
                cause = (
                    "The geometry file on the server has zero bytes. The "
                    "original upload likely failed half-way through."
                )
            elif "suspiciously small" in reason_lower:
                cause = (
                    "The geometry file is too small to be a real 3D model. "
                    "The upload was probably truncated, or the converter "
                    "wrote only an error stub."
                )
            elif "<!doctype html" in (first_tag or "").lower() or (
                first_tag and first_tag.lower() in ("<html>", "<body>")
            ):
                cause = (
                    "The stored file is an HTML page, not a 3D model. The "
                    "converter likely saved an error page by mistake. The "
                    "source CAD/BIM file may not be supported, or the "
                    "converter service was unreachable during processing."
                )
            elif first_tag and first_tag.lower() in (
                "<ifcxml>",
                "<gbxml>",
                "<xml>",
                "<?xml>",
            ):
                cause = (
                    f"The stored file is {first_tag} (XML data) instead of "
                    "a 3D mesh. The source format does not contain 3D "
                    "geometry to display - e.g. an IFC schedule or a "
                    "2D-only drawing."
                )
            elif "magic mismatch" in reason_lower:
                cause = (
                    "The file's first bytes don't match the expected "
                    "format signature. The file is either corrupted in "
                    "transit, or its extension was renamed manually."
                )
            elif "unsupported glb version" in reason_lower:
                cause = (
                    "The file is an older glTF format version that our viewer doesn't support (we require glTF 2.0)."
                )
            else:
                cause = (
                    f"The stored file does not match the expected {ext} "
                    "signature. The CAD converter may have run with an "
                    "older version, or the source file is corrupt."
                )
            logger.warning(
                "BIM geometry served from %s failed serve-time validation: %s "
                "(request_id=%s, model_id=%s, size=%d, head=%s, "
                "first_tag=%s, ext=%s)",
                key,
                reason_serve,
                request_id,
                model_id,
                _geo_size,
                head_hex,
                first_tag,
                ext,
            )
            diagnostic = {
                "error": "geometry_invalid",
                "category": "file_format",
                "request_id": request_id,
                "reason": reason_serve,
                "cause": cause,
                "format": ext.lstrip(".") or "unknown",
                "stored_extension": ext,
                "expected_signature": expected_signature,
                "size_bytes": _geo_size,
                "head_hex": head_hex,
                "head_ascii": head_ascii,
                "first_tag": first_tag,
                "model_id": str(model_id),
                "remediation": (
                    "Delete this model and re-upload the source CAD/BIM "
                    "file. If the problem repeats with the same file, the "
                    "source itself may be unsupported (2D-only DWG, IFC "
                    "schedule with no geometry, corrupted RVT) - try "
                    "exporting from your CAD tool again, or contact "
                    "info@datadrivenconstruction.io and quote the "
                    "Request ID shown below."
                ),
                "message": (f"Geometry file is not a valid {ext} payload: {reason_serve}"),
            }
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=diagnostic,
                headers={"X-Request-Id": request_id},
            )

        # A model name is user data and an HTTP header value has to be Latin-1
        # encodable, so a Cyrillic or Arabic name interpolated raw makes the
        # ASGI server raise while serialising the response and the whole
        # geometry request answers 500, with the frontend spinning on
        # "loading" forever. The helper emits the RFC 6266 pair: an ASCII
        # fallback for old clients and the real name in filename*.
        from fastapi.responses import FileResponse, Response

        display_name = f"{model.name}{ext}"
        cd_header = attachment_disposition(display_name, inline=True)

        # Surface the correlation ID even on the happy path so a downstream JS
        # parsing failure still has a request_id to quote (matches every error
        # branch above).
        response_headers = {
            **cache_headers,
            "Content-Disposition": cd_header,
            "X-Request-Id": request_id,
        }

        if disk_path is not None:
            # Stream from disk: FileResponse sets Content-Length + Accept-Ranges
            # and honours Range requests, all without holding the blob in
            # memory. We do NOT gzip - GLB is binary and near-incompressible,
            # and the frontend loader (ElementManager.doLoadGeometry) reads the
            # response body straight into an ArrayBuffer and sniffs the raw
            # bytes (glTF magic / <COLLADA root), so raw bytes parse correctly.
            # Passing the Content-Disposition via headers (not filename=) keeps
            # the header byte-identical to the in-memory branch below.
            return FileResponse(
                disk_path,
                media_type=media_type,
                headers=response_headers,
            )

        # In-memory fallback (backend exposed no local path). GLB is
        # near-incompressible, so gzip would only buy a second full copy for no
        # real transfer win - return the raw bytes. Text geometry (.dae /
        # .gltf) still compresses well, so gzip those, reassigning _geo_bytes
        # so the raw and compressed copies are never both held. Either way the
        # browser transparently decompresses Content-Encoding: gzip before the
        # loader sees the bytes, so the client parses correctly in both cases.
        if ext.lower() in (".dae", ".gltf"):
            _geo_bytes = _gzip.compress(_geo_bytes, compresslevel=6)
            response_headers["Content-Encoding"] = "gzip"

        return Response(
            content=_geo_bytes,
            media_type=media_type,
            headers=response_headers,
        )

    # ── No geometry blob on storage - disambiguate by model status ──────────
    #
    # A bare "geometry_missing" 404 used to fire here for FOUR very different
    # situations, which left the viewer unable to tell a transient "still
    # converting" from a permanent data loss.  We branch on the persisted
    # BimModel status (the source of truth, set by _process_cad_in_background
    # / retry) so the frontend can show the right message and so a genuine
    # data problem stays distinguishable from a model that simply has no 3D
    # yet.  Known status values: processing, needs_converter, ready,
    # degraded, error (see models.py + the status assignments in this file).
    #
    # Response shape stays backward-compatible: still HTTP 404, still a JSON
    # body with `error` / `category` / `request_id` / `model_id` / `message`
    # / `remediation`.  We only swap the stable `error` code (and add a
    # `model_status` echo) so the FE can branch on it.
    model_status = (model.status or "").lower()

    if model_status == "processing":
        # Conversion is still queued/running - the blob will appear once the
        # background worker finishes. Transient, the FE should keep polling.
        logger.info(
            "BIM geometry requested while conversion still running "
            "(request_id=%s, model_id=%s, project_id=%s, status=%s)",
            request_id,
            model_id,
            project_id,
            model_status,
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "geometry_pending",
                "category": "not_ready",
                "request_id": request_id,
                "model_id": str(model_id),
                "model_status": model_status,
                "message": ("3D geometry is still being generated for this model."),
                "remediation": (
                    "Conversion is in progress. The viewer will load "
                    "automatically once the CAD converter finishes - this "
                    "usually takes from a few seconds to a couple of minutes "
                    "depending on file size. No action is needed."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    if model_status == "error":
        # Conversion failed outright - no blob will ever appear without a
        # successful retry. error_message carries the converter diagnosis.
        logger.info(
            "BIM geometry requested for a model whose conversion failed "
            "(request_id=%s, model_id=%s, project_id=%s, status=%s)",
            request_id,
            model_id,
            project_id,
            model_status,
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "geometry_failed",
                "category": "conversion_failed",
                "request_id": request_id,
                "model_id": str(model_id),
                "model_status": model_status,
                "conversion_error": model.error_message or None,
                "message": ("3D geometry could not be generated: the CAD conversion of this model failed."),
                "remediation": (
                    "Open the model in the BIM tab and click Retry to run the "
                    "conversion again. If it keeps failing the source file may "
                    "be corrupt or unsupported - re-export it from your CAD "
                    "tool, or contact info@datadrivenconstruction.io and "
                    "quote the Request ID below."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    if model_status == "needs_converter":
        # The model legitimately has no 3D on this server: the DDC converter
        # required for this format (e.g. RVT) is not installed, so geometry
        # was never produced. Recoverable, but no mesh exists today.
        logger.info(
            "BIM geometry requested but no converter produced a mesh "
            "(request_id=%s, model_id=%s, project_id=%s, status=%s)",
            request_id,
            model_id,
            project_id,
            model_status,
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "geometry_absent",
                "category": "no_geometry",
                "request_id": request_id,
                "model_id": str(model_id),
                "model_status": model_status,
                "message": (
                    "This model has no 3D geometry: the converter required "
                    "for its format is not available on this server."
                ),
                "remediation": (
                    "Install the matching BIM converter under "
                    "Settings → BIM Converters, then open the model and click "
                    "Retry to generate 3D geometry. Until then only the "
                    "model's metadata is available, not a 3D view."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    if model_status == "ready":
        # Status promises geometry but the blob is gone - a genuine,
        # unexpected data problem worth reporting. This is the ONLY case
        # that keeps the legacy "geometry_missing" code.
        #
        # Log EVERY directory we searched (active root + back-compat fallbacks)
        # so an operator can see exactly where the blob was expected. The
        # 8.6.1 root cause was a data-dir mismatch: geometry written under one
        # resolution, read from another. If this still fires after 8.6.1 the
        # blob is genuinely gone, and these paths prove the search was complete.
        try:
            from app.core.storage import safe_data_roots as _safe_data_roots

            _prefix = bim_file_storage.bim_model_prefix(project_id, model_id)
            _searched = [str(_resolve_data_dir())] + [
                str(p) for p in _safe_data_roots() if str(p) != str(_resolve_data_dir())
            ]
        except Exception:  # noqa: BLE001 - diagnostics must never break the response
            _prefix = "<unavailable>"
            _searched = []
        logger.warning(
            "BIM geometry MISSING for a ready model - blob gone from storage "
            "(request_id=%s, model_id=%s, project_id=%s, status=%s, key_prefix=%s, "
            "searched_roots=%s)",
            request_id,
            model_id,
            project_id,
            model_status,
            _prefix,
            _searched,
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={
                "error": "geometry_missing",
                "category": "not_found",
                "request_id": request_id,
                "model_id": str(model_id),
                "model_status": model_status,
                "message": ("This model is marked ready but its 3D geometry file is no longer on the server."),
                "remediation": (
                    "The geometry was generated but the file appears to have "
                    "been deleted from storage. Re-upload the source CAD/BIM "
                    "file to regenerate it. If the file was not deleted "
                    "manually this is a server-side data problem - contact "
                    "info@datadrivenconstruction.io and quote the Request ID "
                    "below."
                ),
            },
            headers={"X-Request-Id": request_id},
        )

    # Any other status (degraded, demo/seed rows, or an unknown future
    # value): the model never reliably carried a 3D mesh on this server.
    # `degraded` normally DOES ship geometry, so a missing blob there is
    # surprising - but it is not a hard "ready" promise, so we report it as
    # "absent" (no 3D to show) rather than the data-loss "missing" code,
    # keeping "geometry_missing" reserved for the ready-but-gone case above.
    logger.warning(
        "BIM geometry not found on storage for non-ready model (request_id=%s, model_id=%s, project_id=%s, status=%s)",
        request_id,
        model_id,
        project_id,
        model_status or "unknown",
    )
    raise HTTPException(
        status_code=status.HTTP_404_NOT_FOUND,
        detail={
            "error": "geometry_absent",
            "category": "no_geometry",
            "request_id": request_id,
            "model_id": str(model_id),
            "model_status": model_status or "unknown",
            "message": ("No 3D geometry is available for this model."),
            "remediation": (
                "This model does not currently have a 3D mesh on the server - "
                "it may be a non-3D row, a demo entry, or an import that "
                "produced no geometry. Re-upload the source CAD/BIM file to "
                "generate geometry. If the same source repeatedly produces no "
                "geometry, contact info@datadrivenconstruction.io and quote "
                "the Request ID below."
            ),
        },
        headers={"X-Request-Id": request_id},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Streaming tiles (fast viewer)
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/tiles/manifest/", response_model=None)
async def get_model_tiles_manifest(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> Response:
    """Return the streaming-tile manifest for a model, baking it on first use.

    The viewer calls this before loading geometry. A ``200`` returns the tile
    manifest (bounds + per-tile bbox, node names, content hashes). A ``204 No
    Content`` means this model has no streamable tileset - too small, no GLB,
    still converting, or a bake error - and the viewer falls back to the
    single monolithic GLB. Either way the model is viewable; the manifest is
    purely a fast-path opt-in.
    """
    await _verify_model_access(service, model_id, user_id or "")
    try:
        manifest = await service.ensure_tileset(model_id)
    except Exception:  # noqa: BLE001 - never fail the viewer over the fast path
        logger.exception("ensure_tileset failed for model %s - viewer uses monolith", model_id)
        manifest = None

    if not manifest or not manifest.get("tiles"):
        return Response(status_code=status.HTTP_204_NO_CONTENT)

    return Response(
        content=json.dumps(manifest),
        media_type="application/json",
        # The manifest tracks the current geometry and lists content hashes
        # that change on re-convert, so it is always served fresh; the size
        # is trivial next to the tiles it points at.
        headers={"Cache-Control": "no-store"},
    )


@router.head(
    "/models/{model_id}/tiles/{tile_hash}/",
    response_model=None,
    include_in_schema=False,
)
@router.get("/models/{model_id}/tiles/{tile_hash}/", response_model=None)
async def get_model_tile(
    model_id: uuid.UUID,
    tile_hash: str,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> Response | RedirectResponse:
    """Serve one content-addressed geometry tile (GLB).

    A tile URL is its content hash, so a tile is immutable: it carries a
    one-year ``immutable`` cache directive and the browser, any CDN, and the
    viewer's IndexedDB cache may keep it forever. This is what makes a revisit
    (or an offline site tablet) instant instead of a full re-download - the
    opposite of the ``no-store`` monolith path.
    """
    model = await _verify_model_access(service, model_id, user_id or "")
    project_id = str(model.project_id)
    key = bim_file_storage.tile_key(project_id, model_id, tile_hash)
    immutable = "public, max-age=31536000, immutable"

    # S3-style backend: redirect the browser straight to a presigned URL.
    presigned = bim_file_storage.presigned_geometry_url(key)
    if presigned:
        return RedirectResponse(url=presigned, status_code=307)

    blob = await bim_file_storage.read_tile(project_id, model_id, tile_hash)
    if not blob:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Tile not found - reload to fetch a fresh manifest.",
        )
    return Response(
        content=blob,
        media_type="model/gltf-binary",
        headers={"Cache-Control": immutable},
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Models
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/", response_model=BIMModelListResponse)
async def list_models(
    project_id: uuid.UUID = Query(...),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=200),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelListResponse:
    """List BIM models for a project.

    Always returns every persisted BIMModel row (sorted by ``created_at``
    desc) - including ``ready``, ``processing``, ``needs_converter`` and
    ``error`` rows - so the /bim page can surface already-converted models
    *without* triggering re-conversion.  Each item is enriched with
    ``conversion_artifact_size_mb`` (sum of GLB/DAE/parquet/thumbnail
    bytes), ``has_original`` (True iff the raw upload is still on
    storage) and ``error_code`` (stable id lifted out of the metadata
    blob for the converter-required UI state).  The list response also
    carries aggregate ``total_artifact_size_mb`` /
    ``total_original_size_mb`` totals which drive the disk-usage chip
    in the BIM page header.
    """
    await _verify_project_access(service.session, project_id, user_id or "")
    items, total = await service.list_models(project_id, offset=offset, limit=limit)

    # Batched storage probe: ONE list_prefix sweep against the backend
    # collects artifact/original/geometry info for every model in the
    # page. Replaces the previous asyncio.gather fan-out which issued
    # 3+ probes per model (50 models → 150+ HEAD/stat round-trips per
    # list call - classic N+1 against storage).  When the backend
    # doesn't support list_prefix (community backends predating v4.6.1),
    # fall back to the per-model probe loop so behaviour is unchanged.
    storage_summary: dict[str, dict[str, object]] = {}
    use_bulk = bim_file_storage.list_prefix_supported()
    if use_bulk:
        try:
            storage_summary = await bim_file_storage.bulk_model_storage_summary(
                project_id,
            )
        except Exception:  # noqa: BLE001  # never break listing on storage hiccups
            logger.exception(
                "bulk_model_storage_summary failed for project=%s; falling back to per-model probes.",
                project_id,
            )
            use_bulk = False

    async def _enrich_per_model(model_obj):  # type: ignore[no-untyped-def]
        """Per-model probe fallback (community backends without list_prefix)."""
        size_bytes = 0
        has_orig = False
        has_geom = bool(model_obj.canonical_file_path)
        try:
            size_bytes = await bim_file_storage.compute_artifact_size_bytes(
                model_obj.project_id,
                model_obj.id,
            )
        except Exception:  # noqa: BLE001
            logger.exception("artifact-size probe failed for model=%s", model_obj.id)
        ext_raw = (model_obj.model_format or "").lstrip(".")
        if ext_raw:
            try:
                has_orig = await bim_file_storage.has_original_cad(
                    model_obj.project_id,
                    model_obj.id,
                    ext=f".{ext_raw}",
                )
            except Exception:  # noqa: BLE001
                logger.exception("has_original probe failed for model=%s", model_obj.id)
        if not has_geom:
            try:
                has_geom = (
                    await bim_file_storage.find_geometry_key(
                        project_id=str(model_obj.project_id),
                        model_id=str(model_obj.id),
                    )
                ) is not None
            except Exception:  # noqa: BLE001
                logger.exception("has_geometry probe failed for model=%s", model_obj.id)
        return size_bytes, has_orig, has_geom

    def _enrich_from_summary(model_obj):  # type: ignore[no-untyped-def]
        """Fast bulk-summary-driven enrich (no I/O)."""
        info = storage_summary.get(str(model_obj.id), {})
        size_bytes = int(info.get("artifact_size_bytes", 0) or 0)
        # ``has_original`` historically reflected the existence of a blob
        # at original.{ext} where ext == model_format.  The bulk sweep
        # uses the same "filename starts with original." rule, so the
        # two definitions match for every realistic model row.
        has_orig = bool(info.get("has_original", False))
        has_geom = bool(model_obj.canonical_file_path) or bool(
            info.get("geometry_exts") or (),
        )
        return size_bytes, has_orig, has_geom

    if use_bulk:
        enriched = [_enrich_from_summary(m) for m in items]
    else:
        import asyncio as _asyncio

        enriched = await _asyncio.gather(
            *[_enrich_per_model(m) for m in items],
            return_exceptions=False,
        )

    item_responses: list[BIMModelResponse] = []
    total_artifact_bytes = 0
    total_original_bytes = 0
    for model_obj, (size_bytes, has_orig, has_geom) in zip(items, enriched, strict=True):
        resp = BIMModelResponse.model_validate(model_obj)
        resp.conversion_artifact_size_mb = round(size_bytes / (1024 * 1024), 3)
        resp.has_original = has_orig
        resp.has_geometry = has_geom
        meta = model_obj.metadata_ or {}
        if isinstance(meta, dict):
            err_code = meta.get("error_code")
            if isinstance(err_code, str):
                resp.error_code = err_code
        total_artifact_bytes += size_bytes
        if has_orig:
            # Original-blob size: pulled from the bulk summary when
            # available (zero extra round-trips), or from a single
            # per-model size() probe on the fallback path.
            ext_raw = (model_obj.model_format or "").lstrip(".")
            if use_bulk:
                info = storage_summary.get(str(model_obj.id), {})
                total_original_bytes += int(info.get("original_size_bytes", 0) or 0)
            elif ext_raw:
                try:
                    backend = bim_file_storage._backend()
                    key = bim_file_storage.original_cad_key(
                        model_obj.project_id,
                        model_obj.id,
                        ext=f".{ext_raw}",
                    )
                    total_original_bytes += await backend.size(key)
                except Exception:  # noqa: BLE001
                    pass
        item_responses.append(resp)

    return BIMModelListResponse(
        items=item_responses,
        total=total,
        offset=offset,
        limit=limit,
        total_artifact_size_mb=round(total_artifact_bytes / (1024 * 1024), 3),
        total_original_size_mb=round(total_original_bytes / (1024 * 1024), 3),
        storage_root_label=bim_file_storage.bim_root_label(),
    )


@router.post("/", response_model=BIMModelResponse, status_code=201)
async def create_model(
    data: BIMModelCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Create a new BIM model record."""
    await _verify_project_access(service.session, data.project_id, user_id)
    model = await service.create_model(data, user_id=user_id)
    return BIMModelResponse.model_validate(model)


# ─── Asset Register routes (must come BEFORE /{model_id} so that
#     `/assets` is not interpreted as a UUID model_id and rejected with
#     422 by the path validator). The handlers and the `_summarise_asset`
#     helper live further down in the file under the "Asset Register
#     (v2.3.0)" section header - only the route registrations move up. ───


@router.get("/assets", response_model=AssetListResponse)
async def list_assets(
    project_id: uuid.UUID = Query(..., description="Scope the asset list to this project"),
    element_type: str | None = Query(default=None),
    operational_status: str | None = Query(default=None),
    search: str | None = Query(default=None, min_length=1, max_length=200),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> AssetListResponse:
    """List tracked assets across every BIM model in a project.

    Tracked = ``BIMElement.is_tracked_asset == True``. Managed-asset
    rows appear on the Assets page with manufacturer / serial / warranty
    columns lifted out of the ``asset_info`` JSON blob for convenient
    sorting.
    """
    await _verify_project_access(service.session, project_id, user_id or "")
    rows, total = await service.list_tracked_assets(
        project_id,
        element_type=element_type,
        operational_status=operational_status,
        search=search,
        offset=offset,
        limit=limit,
    )
    return AssetListResponse(
        items=[_summarise_asset(element, model) for element, model in rows],
        total=total,
        offset=offset,
        limit=limit,
    )


@router.patch("/assets/{element_id}/asset-info", response_model=BIMElementResponse)
async def update_asset_info(
    element_id: uuid.UUID,
    payload: AssetInfoUpdateRequest,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Merge-update asset_info on a BIMElement.

    Semantics:
    - Any key you send overwrites the matching key in ``asset_info``.
    - Sending ``null`` or ``""`` for a key clears that key.
    - ``is_tracked_asset`` auto-flips to ``True`` on first non-empty
      write; pass an explicit bool to override.
    - Unrelated keys already in ``asset_info`` survive untouched.
    """
    # Locate the element first so we can verify project access.
    element = await service.get_element(element_id)
    await _verify_model_access(service, element.model_id, user_id or "")

    updated = await service.update_asset_info(
        element_id,
        asset_info=payload.asset_info.model_dump(exclude_unset=False),
        is_tracked_asset=payload.is_tracked_asset,
    )
    return BIMElementResponse.model_validate(updated)


@router.get("/{model_id}", response_model=BIMModelResponse)
async def get_model(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Get a single BIM model by ID."""
    model = await _verify_model_access(service, model_id, user_id or "")
    resp = BIMModelResponse.model_validate(model)
    # Mirror the list endpoint enrichment so single-model polls
    # (status transitions during background conversion) also expose
    # artifact size + ``has_original`` to the frontend.
    try:
        size_bytes = await bim_file_storage.compute_artifact_size_bytes(
            model.project_id,
            model.id,
        )
        resp.conversion_artifact_size_mb = round(size_bytes / (1024 * 1024), 3)
    except Exception:  # noqa: BLE001
        logger.exception("artifact-size probe failed for model=%s", model.id)
    ext_raw = (model.model_format or "").lstrip(".")
    if ext_raw:
        try:
            resp.has_original = await bim_file_storage.has_original_cad(
                model.project_id,
                model.id,
                ext=f".{ext_raw}",
            )
        except Exception:  # noqa: BLE001
            logger.exception("has_original probe failed for model=%s", model.id)
    # has_geometry is true when the background converter saved a GLB/DAE
    # (``canonical_file_path`` set). Probing storage as a tie-breaker for
    # historical rows where the column was missed.
    resp.has_geometry = bool(model.canonical_file_path)
    if not resp.has_geometry:
        try:
            resp.has_geometry = (
                await bim_file_storage.find_geometry_key(
                    project_id=str(model.project_id),
                    model_id=str(model.id),
                )
            ) is not None
        except Exception:  # noqa: BLE001
            logger.exception("has_geometry probe failed for model=%s", model.id)
    meta = model.metadata_ or {}
    if isinstance(meta, dict):
        err_code = meta.get("error_code")
        if isinstance(err_code, str):
            resp.error_code = err_code
    return resp


@router.patch("/{model_id}", response_model=BIMModelResponse)
async def update_model(
    model_id: uuid.UUID,
    data: BIMModelUpdate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelResponse:
    """Update a BIM model."""
    await _verify_model_access(service, model_id, user_id)
    model = await service.update_model(model_id, data)
    return BIMModelResponse.model_validate(model)


@router.get("/models/{model_id}/schema/")
async def get_model_schema(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
):
    """Return distinct element types + property keys/values for the model.

    Used by the quantity-rule editor (RFC 24) to seed combobox options.
    """
    await _verify_model_access(service, model_id, user_id or "")
    return await service.get_model_schema(model_id)


@router.delete("/{model_id}", status_code=204)
async def delete_model(
    model_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BIM model and all its elements."""
    await _verify_model_access(service, model_id, user_id)
    await service.delete_model(model_id)


@router.post("/cleanup-stale/")
async def cleanup_stale_processing(
    project_id: uuid.UUID = Query(...),
    max_age_hours: int = Query(default=1, ge=0),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, int]:
    """Remove models stuck in 'processing' with 0 elements older than max_age_hours."""
    await _verify_project_access(service.session, project_id, user_id or "")
    count = await service.cleanup_stale_processing(project_id, max_age_hours=max_age_hours)
    return {"deleted": count}


@router.post("/cleanup-orphans/")
async def cleanup_orphan_bim_files(
    _user_id: CurrentUserId,
    _role: None = Depends(RequireRole("admin")),
    service: BIMHubService = Depends(_get_service),
) -> dict[str, Any]:
    """Scan ``data/bim/`` and remove directories with no matching DB row.

    Admin-grade disk hygiene. This is a GLOBAL, cross-tenant filesystem sweep:
    it loads every model id in the deployment and ``rmtree``s any model
    directory under ``data/bim/`` without a matching DB row, crossing tenant
    boundaries. A project-level MANAGER (``bim.delete``) must never reach it,
    so it is gated by ``RequireRole("admin")`` - the same pattern used by the
    clear-database / demo-reset endpoints for tenant-wide bulk deletes.

    Protects against orphaned RVT/IFC/COLLADA/Excel artefacts left behind by
    failed uploads, crashed conversions, or manual DB deletes that bypassed
    the service layer.
    """
    return await service.cleanup_orphan_bim_files()


# ═══════════════════════════════════════════════════════════════════════════════
# Elements
# ═══════════════════════════════════════════════════════════════════════════════


def _assemble_element_responses(
    items: list,
    boq_links_by_id: dict,
    doc_links_by_id: dict,
    task_links_by_id: dict,
    activity_briefs_by_id: dict,
    requirement_briefs_by_id: dict,
    validation_summaries_by_id: dict,
    current_pct_by_id: dict,
    current_pct_date_by_id: dict,
    report_exists: bool,
) -> list[BIMElementResponse]:
    """Map ``list_elements_with_links`` briefs onto ``BIMElementResponse`` rows.

    Shared by the model-wide element list and the single-element context
    endpoint so both surface identical BOQ / document / task / activity /
    requirement / validation / progress enrichment. ``report_exists`` decides
    whether an element with no findings reads as 'pass' (a report ran) or
    'unchecked' (no report at all).
    """
    from app.modules.bim_hub.schemas import (
        ActivityBrief,
        DocumentLinkBrief,
        ElementValidationSummary,
        RequirementBrief,
        TaskBrief,
    )

    responses: list[BIMElementResponse] = []
    for elem in items:
        boq_briefs = [BOQElementLinkBrief.model_validate(b) for b in boq_links_by_id.get(elem.id, [])]
        doc_briefs = [DocumentLinkBrief.model_validate(b) for b in doc_links_by_id.get(elem.id, [])]
        task_briefs = [TaskBrief.model_validate(b) for b in task_links_by_id.get(elem.id, [])]
        activity_briefs = [ActivityBrief.model_validate(b) for b in activity_briefs_by_id.get(elem.id, [])]
        requirement_briefs = [RequirementBrief.model_validate(b) for b in requirement_briefs_by_id.get(elem.id, [])]
        raw_val = validation_summaries_by_id.get(elem.id, [])
        validation_summaries = [ElementValidationSummary.model_validate(v) for v in raw_val]
        # Derive worst-severity status; 'unchecked' iff no report exists at all.
        if not report_exists:
            val_status: str = "unchecked"
        elif any(v.severity == "error" for v in validation_summaries):
            val_status = "error"
        elif any(v.severity == "warning" for v in validation_summaries):
            val_status = "warning"
        else:
            val_status = "pass"
        resp = BIMElementResponse.model_validate(elem)
        resp.boq_links = boq_briefs
        resp.linked_documents = doc_briefs
        resp.linked_tasks = task_briefs
        resp.linked_activities = activity_briefs
        resp.linked_requirements = requirement_briefs
        resp.validation_results = validation_summaries
        resp.validation_status = val_status  # type: ignore[assignment]
        resp.current_pct = current_pct_by_id.get(elem.id)
        resp.current_pct_date = current_pct_date_by_id.get(elem.id)
        responses.append(resp)
    return responses


@router.get("/models/{model_id}/elements/", response_model=BIMElementListResponse)
async def list_elements(
    model_id: uuid.UUID,
    element_type: str | None = Query(default=None),
    storey: str | None = Query(default=None),
    discipline: str | None = Query(default=None),
    group_id: uuid.UUID | None = Query(
        default=None,
        description="Filter to elements belonging to this saved element group",
    ),
    offset: int = Query(default=0, ge=0),
    # Cap depends on ``skeleton``: the enriched path is paginated at 2000/page
    # because each row fans out to six relation joins (boq_links, docs,
    # tasks, activities, requirements, validation). Skeleton mode returns
    # plain BIMElement rows with no joins and is safe at 50000/page.
    limit: int = Query(default=2000, ge=1, le=200000),
    skeleton: bool = Query(
        default=False,
        description=(
            "Skip eager-loading of boq_links / linked_documents / linked_tasks "
            "/ linked_activities / linked_requirements / validation_results. "
            "~10× faster and used by the 3D viewer for mesh-to-element "
            "matching, where those relations are not needed."
        ),
    ),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """List elements for a BIM model (paginated, filterable).

    Each element in the response includes a ``boq_links`` array of
    ``BOQElementLinkBrief`` entries, a ``linked_documents`` array of
    ``DocumentLinkBrief`` entries, a ``linked_tasks`` array of ``TaskBrief``
    entries, and a ``linked_activities`` array of ``ActivityBrief`` entries
    so the viewer can render link badges without a second round trip.
    """
    await _verify_model_access(service, model_id, user_id or "")

    # Skeleton path: plain BIMElement rows, no relation joins, no enrichment.
    # Ten times faster than the enriched path - used by the 3D viewer, where
    # mesh matching only needs id/mesh_ref/name/element_type/bbox.
    if skeleton:
        if limit > 50000:
            limit = 50000
        plain_items, plain_total = await service.list_elements(
            model_id,
            element_type=element_type,
            storey=storey,
            discipline=discipline,
            offset=offset,
            limit=limit,
        )
        # Skinny rows: drop the per-element `properties` / `quantities` /
        # `classification` / `metadata` payloads. These can weigh ~1.5 kB per
        # row on a typical RVT export (45+ RVT parameters × short value),
        # which adds up to a 16 MB JSON body for 7 000 elements. The viewer
        # pulls the full property set straight from Parquet on click, so
        # carrying it in the skeleton is pure overhead.
        skeleton_items: list[BIMElementResponse] = []
        for e in plain_items:
            resp = BIMElementResponse.model_validate(e)
            resp.properties = {}
            resp.quantities = {}
            resp.metadata = {}
            skeleton_items.append(resp)
        return BIMElementListResponse(
            items=skeleton_items,
            total=plain_total,
            offset=offset,
            limit=limit,
        )

    # Enriched path is capped at 2000 - each extra row spawns six join lookups.
    if limit > 2000:
        limit = 2000
    (
        items,
        total,
        boq_links_by_id,
        doc_links_by_id,
        task_links_by_id,
        activity_briefs_by_id,
        requirement_briefs_by_id,
        validation_summaries_by_id,
        current_pct_by_id,
        current_pct_date_by_id,
    ) = await service.list_elements_with_links(
        model_id,
        element_type=element_type,
        storey=storey,
        discipline=discipline,
        group_id=group_id,
        offset=offset,
        limit=limit,
    )

    # The service stashes a sentinel entry under ``_VALIDATION_REPORT_SENTINEL``
    # (UUID(int=0)) when a ``target_type='bim_model'`` report exists. We pop
    # it so it never reaches the per-element loop.
    from app.modules.bim_hub.service import _VALIDATION_REPORT_SENTINEL

    report_exists = _VALIDATION_REPORT_SENTINEL in validation_summaries_by_id
    validation_summaries_by_id.pop(_VALIDATION_REPORT_SENTINEL, None)

    responses = _assemble_element_responses(
        items,
        boq_links_by_id,
        doc_links_by_id,
        task_links_by_id,
        activity_briefs_by_id,
        requirement_briefs_by_id,
        validation_summaries_by_id,
        current_pct_by_id,
        current_pct_date_by_id,
        report_exists,
    )

    return BIMElementListResponse(
        items=responses,
        total=total,
        offset=offset,
        limit=limit,
    )


@router.post(
    "/models/{model_id}/elements/",
    response_model=BIMElementListResponse,
    status_code=201,
)
async def bulk_import_elements(
    model_id: uuid.UUID,
    data: BIMElementBulkImport,
    background_tasks: BackgroundTasks,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """Bulk import elements for a model (replaces existing)."""
    await _verify_model_access(service, model_id, user_id)
    elements = await service.bulk_import_elements(model_id, data.elements)
    # Run the advisory validation pass (canonical import pipeline) after the
    # bulk replace, mirroring the CAD converter path. Best-effort, never blocks.
    if elements:
        background_tasks.add_task(_run_import_validation, model_id)
    return BIMElementListResponse(
        items=[BIMElementResponse.model_validate(e) for e in elements],
        total=len(elements),
        offset=0,
        limit=len(elements),
    )


@router.post(
    "/models/{model_id}/elements/by-ids/",
    response_model=BIMElementListResponse,
)
async def get_elements_by_ids(
    model_id: uuid.UUID,
    body: dict,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementListResponse:
    """Fetch specific elements by their IDs (DB UUID or stable_id)."""
    await _verify_model_access(service, model_id, user_id or "")

    element_ids: list[str] = body.get("element_ids", [])
    if not element_ids or len(element_ids) > 100:
        return BIMElementListResponse(items=[], total=0, offset=0, limit=0)

    from sqlalchemy import or_

    from app.modules.bim_hub.models import BIMElement

    parsed_uuids: list[uuid.UUID] = []
    for eid in element_ids:
        try:
            parsed_uuids.append(uuid.UUID(str(eid)))
        except (ValueError, TypeError):
            continue

    query = (
        select(BIMElement)
        .where(BIMElement.model_id == model_id)
        .where(
            or_(
                BIMElement.id.in_(parsed_uuids),
                BIMElement.stable_id.in_(element_ids),
            )
        )
    )
    result = await service.session.execute(query)
    elements = list(result.scalars().all())

    return BIMElementListResponse(
        items=[BIMElementResponse.model_validate(e) for e in elements],
        total=len(elements),
        offset=0,
        limit=len(elements),
    )


@router.post(
    "/models/{model_id}/ensure-element/",
    response_model=BIMElementResponse,
)
async def ensure_element(
    model_id: uuid.UUID,
    body: dict,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Resolve (or lazy-create) a BIMElement row from a mesh_ref / stable_id.

    Needed when linking a BOQ position to a BIM mesh that was visible in the
    3D viewer but had no oe_bim_element row (e.g. DDC standard extract skips
    certain RVT categories). Body: ``{"mesh_ref": "140056"}`` or
    ``{"stable_id": "140056"}``.
    """
    await _verify_model_access(service, model_id, user_id or "")
    stable_id = body.get("stable_id")
    mesh_ref = body.get("mesh_ref")
    element = await service.ensure_element(
        model_id,
        stable_id=str(stable_id) if stable_id else None,
        mesh_ref=str(mesh_ref) if mesh_ref else None,
    )
    return BIMElementResponse.model_validate(element)


@router.get("/elements/{element_id}", response_model=BIMElementResponse)
async def get_element(
    element_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Get a single BIM element by ID."""
    element = await service.get_element(element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Element not found",
        )
    # Verify caller owns the project this element's model belongs to.
    await _verify_model_access(service, element.model_id, user_id or "")
    return BIMElementResponse.model_validate(element)


@router.get("/elements/{element_id}/context", response_model=BIMElementResponse)
async def get_element_context(
    element_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementResponse:
    """Full ERP context for one selected element.

    The 3D viewer loads elements in skeleton mode (no joins) so the model
    opens fast. When the user clicks an element we call this to compose,
    on demand, everything the platform knows about it: linked BOQ positions
    (with cost), documents, tasks, schedule activities, requirements,
    validation findings and install/progress. Reuses the same enrichment
    the model-wide list uses, scoped to a single element.
    """
    element = await service.get_element(element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Element not found",
        )
    await _verify_model_access(service, element.model_id, user_id or "")

    (
        items,
        _total,
        boq_links_by_id,
        doc_links_by_id,
        task_links_by_id,
        activity_briefs_by_id,
        requirement_briefs_by_id,
        validation_summaries_by_id,
        current_pct_by_id,
        current_pct_date_by_id,
    ) = await service.list_elements_with_links(
        element.model_id,
        element_id=element_id,
        limit=1,
    )

    from app.modules.bim_hub.service import _VALIDATION_REPORT_SENTINEL

    report_exists = _VALIDATION_REPORT_SENTINEL in validation_summaries_by_id
    validation_summaries_by_id.pop(_VALIDATION_REPORT_SENTINEL, None)

    responses = _assemble_element_responses(
        items,
        boq_links_by_id,
        doc_links_by_id,
        task_links_by_id,
        activity_briefs_by_id,
        requirement_briefs_by_id,
        validation_summaries_by_id,
        current_pct_by_id,
        current_pct_date_by_id,
        report_exists,
    )
    if not responses:
        # The element vanished between the two queries (rare race); treat as 404.
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Element not found",
        )
    return responses[0]


# ═══════════════════════════════════════════════════════════════════════════════
# Asset Register (v2.3.0)
# ═══════════════════════════════════════════════════════════════════════════════


def _summarise_asset(element, model) -> AssetSummary:
    """Project BIMElement+BIMModel onto the flat AssetSummary row.

    Lifts commonly-displayed fields out of the ``asset_info`` blob so
    the frontend list can sort them without peeking into the JSON.
    """
    info = dict(element.asset_info or {})
    return AssetSummary(
        id=element.id,
        model_id=element.model_id,
        project_id=model.project_id,
        model_name=model.name,
        stable_id=element.stable_id,
        element_type=element.element_type,
        name=element.name,
        storey=element.storey,
        discipline=element.discipline,
        asset_info=info,
        manufacturer=info.get("manufacturer"),
        model=info.get("model"),
        serial_number=info.get("serial_number"),
        warranty_until=info.get("warranty_until"),
        operational_status=info.get("operational_status"),
        asset_tag=info.get("asset_tag"),
    )


# Note: the @router.get("/assets") and @router.patch("/assets/{element_id}/asset-info")
# route definitions were moved up before "/{model_id}" so FastAPI's path
# matcher resolves the literal "assets" segment instead of mistaking it
# for a UUID. The handlers live above; the `_summarise_asset` helper just
# above is still referenced from there at request time.


# ═══════════════════════════════════════════════════════════════════════════════
# COBie Export (v2.3.0)
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/export/cobie.xlsx")
async def export_cobie_xlsx(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse:
    """Generate an ISO-19650 COBie UK 2.4 handover workbook.

    The response streams an ``.xlsx`` file with seven sheets (Contact /
    Facility / Floor / Space / Type / Component / System) built from
    the current canonical BIM data + ``asset_info`` payload.
    """
    await _verify_model_access(service, model_id, user_id or "")
    xlsx_bytes, filename = await service.export_cobie(model_id)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={
            # RFC 6266 - a model name with non-Latin-1 chars would otherwise make
            # the ASGI server 500 while encoding this header.
            "Content-Disposition": attachment_disposition(filename),
            "Content-Length": str(len(xlsx_bytes)),
        },
    )


@router.post("/models/{model_id}/export/boq.xlsx")
async def export_boq_xlsx(
    model_id: uuid.UUID,
    body: BoqExportRequest | None = None,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse:
    """Export the model's measured quantities as a single Excel Bill of
    Quantities (BOQ): a summary sheet grouping elements with summed areas /
    volumes / lengths / weights plus a TOTAL row, and a detail sheet listing
    every element. The body optionally narrows the export to the elements the
    user has visible / selected (``element_ids``), a saved Smart View / group
    (``group_id``), or simple ``filters`` - default is the whole model.
    """
    req = body or BoqExportRequest()
    await _verify_model_access(service, model_id, user_id or "")
    xlsx_bytes, filename = await service.export_boq(
        model_id,
        element_ids=req.element_ids,
        group_id=req.group_id,
        filters=(req.filters.model_dump(exclude_none=True) if req.filters else None),
        group_by=req.group_by,
        title=req.title,
    )
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        headers={
            "Content-Disposition": attachment_disposition(filename),
            "Content-Length": str(len(xlsx_bytes)),
        },
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Model download (source CAD / geometry artifact, for the /files manager)
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/models/{model_id}/download/")
async def download_model(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> StreamingResponse:
    """Download the source/canonical file backing a BIM model.

    Resolution order, all gated by project access (``_verify_model_access``
    404s on missing or foreign-tenant):

    1. the original CAD upload (``original.{ext}``) when it is still retained,
    2. the converted geometry artifact (GLB/DAE/glTF) - what the viewer loads,
    3. a minimal materialized stub (IFC/STEP for ifc-family models, otherwise
       a short text note) when neither blob is present, so demo/showcase rows
       download something valid instead of a 404.

    Everything flows through the storage backend, so the local filesystem and
    S3 deployments both work without a route change.
    """
    model = await _verify_model_access(service, model_id, user_id or "")
    project_id = str(model.project_id)
    backend_store = bim_file_storage._backend()
    model_format = (model.model_format or "").lower().lstrip(".")

    # 1. Original CAD upload, if retained. This is the like-for-like source the
    #    user uploaded (IFC / RVT / DWG), preferred over the converted mesh.
    if model_format:
        ext = f".{model_format}"
        cad_key = bim_file_storage.original_cad_key(project_id, model_id, ext)
        try:
            has_cad = await backend_store.exists(cad_key)
        except Exception:  # noqa: BLE001 - probing storage must never raise to the client
            has_cad = False
        if has_cad:
            blob = await backend_store.get(cad_key)
            filename = f"{model.name or 'model'}{ext}"
            return StreamingResponse(
                io.BytesIO(blob),
                media_type="application/octet-stream",
                headers={
                    "Content-Disposition": attachment_disposition(filename),
                    "Content-Length": str(len(blob)),
                },
            )

    # 2. Converted geometry artifact (GLB/DAE/glTF).
    found = await bim_file_storage.find_geometry_key(project_id, model_id)
    if found is not None:
        key, geo_ext = found
        blob = await backend_store.get(key)
        media_type = bim_file_storage.GEOMETRY_MEDIA_TYPES.get(geo_ext, "application/octet-stream")
        filename = f"{model.name or 'model'}{geo_ext}"
        return StreamingResponse(
            io.BytesIO(blob),
            media_type=media_type,
            headers={
                "Content-Disposition": attachment_disposition(filename),
                "Content-Length": str(len(blob)),
            },
        )

    # 3. Nothing on disk: materialize a typed stub so the /files row still
    #    downloads a valid file. ifc-family -> IFC/STEP text, else a text note.
    import tempfile

    stub_ext = f".{model_format}" if model_format in ("ifc", "step", "stp") else ".ifc"
    name = model.name or "BIM model"
    with tempfile.TemporaryDirectory(prefix="oe-bim-dl-") as tmp:
        stub = pathlib.Path(tmp) / f"{model_id}{stub_ext}"
        try:
            materialize_placeholder(stub, name)
        except Exception:  # pragma: no cover - degrade to 404 on unexpected failure
            logger.warning("Failed to materialize BIM placeholder for %s", model_id, exc_info=True)
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found on disk",
            ) from None
        blob = stub.read_bytes()
    filename = f"{name}{stub_ext}"
    return StreamingResponse(
        io.BytesIO(blob),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": attachment_disposition(filename),
            "Content-Length": str(len(blob)),
        },
    )


# ═══════════════════════════════════════════════════════════════════════════════
# BOQ Links
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_boq_position_access(
    service: "BIMHubService",
    position_id: uuid.UUID,
    user_id: str,
) -> None:
    """Resolve a BOQ position → its BOQ → project and verify the caller owns it.

    `Position` has no direct `project_id` column - the project lives on the
    parent `BOQ` row reached via `position.boq_id`.  We do a single-row
    SELECT joining position → boq so this stays one round-trip.
    """
    # ``BOQ`` is the class name exposed by ``boq.models`` and it refers to
    # the Bill-of-Quantities aggregate, not a module-level constant - the
    # ``N811`` noqa below suppresses ruff's all-caps-is-a-constant heuristic.
    from app.modules.boq.models import BOQ as BOQModel  # noqa: N811
    from app.modules.boq.models import Position

    stmt = select(BOQModel.project_id).join(Position, Position.boq_id == BOQModel.id).where(Position.id == position_id)
    result = await service.session.execute(stmt)
    project_id = result.scalar_one_or_none()
    if project_id is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOQ position not found",
        )
    await _verify_project_access(service.session, project_id, user_id)


@router.get("/links/", response_model=BOQElementLinkListResponse)
async def list_links(
    boq_position_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BOQElementLinkListResponse:
    """List BIM element links for a BOQ position."""
    await _verify_boq_position_access(service, boq_position_id, user_id or "")
    items = await service.list_links_for_position(boq_position_id)
    return BOQElementLinkListResponse(
        items=[BOQElementLinkResponse.model_validate(lnk) for lnk in items],
        total=len(items),
    )


@router.get(
    "/models/{model_id}/boq-links/",
    response_model=BIMModelBOQLinksResponse,
)
async def list_model_boq_links(
    model_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelBOQLinksResponse:
    """Aggregate BOQ links for every element in a model.

    Used by the "Linked BOQ" side-panel in the BIM viewer: the viewer
    itself loads elements in ``skeleton`` mode (no boq_links) for speed,
    so the panel needs a dedicated roll-up across the whole model.
    """
    await _verify_model_access(service, model_id, user_id or "")
    rows = await service.list_links_for_model(model_id)
    return BIMModelBOQLinksResponse(
        items=[BIMModelBOQLinkAggregate.model_validate(r) for r in rows],
        total=len(rows),
    )


@router.post("/links/", response_model=BOQElementLinkResponse, status_code=201)
async def create_link(
    data: BOQElementLinkCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BOQElementLinkResponse:
    """Create a link between a BOQ position and a BIM element."""
    # Verify both sides: the BOQ position's project AND the BIM element's
    # model/project. Prevents cross-project link forgery.
    await _verify_boq_position_access(service, data.boq_position_id, user_id)
    element = await service.get_element(data.bim_element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BIM element not found",
        )
    await _verify_model_access(service, element.model_id, user_id)
    link = await service.create_link(data, user_id=user_id)
    return BOQElementLinkResponse.model_validate(link)


@router.delete("/links/{link_id}", status_code=204)
async def delete_link(
    link_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BOQ-BIM link."""
    # Resolve the link → element → model → project and verify access.
    from app.modules.bim_hub.models import BOQElementLink

    link = await service.session.get(BOQElementLink, link_id)
    if link is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found",
        )
    element = await service.get_element(link.bim_element_id)
    if element is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Link not found",
        )
    await _verify_model_access(service, element.model_id, user_id)
    await service.delete_link(link_id)


# ═══════════════════════════════════════════════════════════════════════════════
# Quantity Maps
# ═══════════════════════════════════════════════════════════════════════════════


@router.get("/quantity-maps/", response_model=BIMQuantityMapListResponse)
async def list_quantity_maps(
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapListResponse:
    """List quantity mapping rules visible to the caller.

    Scopes to the caller's accessible projects so a project-scoped rule from
    another tenant never leaks. Global templates (``project_id IS NULL``)
    stay visible to everyone; admins see every rule.
    """
    scope = await accessible_project_ids(service.session, user_id)
    items, total = await service.list_quantity_maps(project_ids=scope, offset=offset, limit=limit)
    return BIMQuantityMapListResponse(
        items=[BIMQuantityMapResponse.model_validate(m) for m in items],
        total=total,
    )


@router.post("/quantity-maps/", response_model=BIMQuantityMapResponse, status_code=201)
async def create_quantity_map(
    data: BIMQuantityMapCreate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapResponse:
    """Create a new quantity mapping rule."""
    # If the rule is scoped to a specific project, enforce ownership.
    if data.project_id is not None:
        await _verify_project_access(service.session, data.project_id, user_id)
    qmap = await service.create_quantity_map(data)
    return BIMQuantityMapResponse.model_validate(qmap)


@router.patch("/quantity-maps/{map_id}", response_model=BIMQuantityMapResponse)
async def update_quantity_map(
    map_id: uuid.UUID,
    data: BIMQuantityMapUpdate,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMQuantityMapResponse:
    """Update a quantity mapping rule.

    A project-scoped rule requires access to its project. A global rule
    (``project_id IS NULL``) is a cross-tenant shared template, so only an
    admin may mutate it - a project-level editor in any tenant must not be
    able to silently rewrite a template every tenant sees. The 404 (not 403)
    on the non-admin global case keeps the IDOR surface consistent.
    """
    from app.modules.bim_hub.models import BIMQuantityMap

    existing = await service.session.get(BIMQuantityMap, map_id)
    if existing is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Quantity map not found",
        )
    if existing.project_id is not None:
        await _verify_project_access(service.session, existing.project_id, user_id)
    else:
        # Global/template rule: admins only. ``accessible_project_ids``
        # returns ``None`` for admins (its "no filter" sentinel); any
        # non-admin caller gets a set and is rejected as not-found.
        scope = await accessible_project_ids(service.session, user_id)
        if scope is not None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Quantity map not found",
            )
    qmap = await service.update_quantity_map(map_id, data)
    return BIMQuantityMapResponse.model_validate(qmap)


@router.post("/quantity-maps/apply/", response_model=QuantityMapApplyResult)
async def apply_quantity_maps(
    data: QuantityMapApplyRequest,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> QuantityMapApplyResult:
    """Apply quantity mapping rules to all elements in a model."""
    await _verify_model_access(service, data.model_id, user_id)
    return await service.apply_quantity_maps(data)


# ═══════════════════════════════════════════════════════════════════════════════
# Diffs
# ═══════════════════════════════════════════════════════════════════════════════


@router.post("/models/{model_id}/diff/{old_id}", response_model=BIMModelDiffResponse, status_code=201)
async def compute_diff(
    model_id: uuid.UUID,
    old_id: uuid.UUID,
    user_id: CurrentUserId,
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelDiffResponse:
    """Compute diff between two model versions."""
    # Both models must be readable by the caller.
    await _verify_model_access(service, model_id, user_id)
    await _verify_model_access(service, old_id, user_id)
    diff = await service.compute_diff(new_model_id=model_id, old_model_id=old_id)
    return BIMModelDiffResponse.model_validate(diff)


@router.get("/diffs/{diff_id}", response_model=BIMModelDiffResponse)
async def get_diff(
    diff_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> BIMModelDiffResponse:
    """Get a model diff by ID."""
    diff = await service.get_diff(diff_id)
    if diff is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Diff not found",
        )
    # Verify access via the new (or old) model's project.
    await _verify_model_access(service, diff.new_model_id, user_id or "")
    return BIMModelDiffResponse.model_validate(diff)


# ═══════════════════════════════════════════════════════════════════════════════
# Element Groups (saved selections)
# ═══════════════════════════════════════════════════════════════════════════════


async def _verify_group_access(
    service: "BIMHubService",
    group_id: uuid.UUID,
    user_id: str,
) -> Any:
    """Load a BIM element group and verify the caller owns its project.

    Returns the loaded group so the caller can reuse it. Raises 404 on both
    "not found" and "no access" to avoid UUID enumeration.
    """
    from app.modules.bim_hub.models import BIMElementGroup

    group = await service.session.get(BIMElementGroup, group_id)
    if group is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BIM element group not found",
        )
    await _verify_project_access(service.session, group.project_id, user_id)
    return group


@router.get("/element-groups/", response_model=list[BIMElementGroupResponse])
async def list_element_groups(
    project_id: uuid.UUID = Query(...),
    model_id: uuid.UUID | None = Query(default=None),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> list[BIMElementGroupResponse]:
    """List BIM element groups for a project, optionally scoped to one model."""
    await _verify_project_access(service.session, project_id, user_id or "")
    return await service.list_element_groups(project_id, model_id=model_id)


@router.post(
    "/element-groups/",
    response_model=BIMElementGroupResponse,
    status_code=201,
)
async def create_element_group(
    data: BIMElementGroupCreate,
    project_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.create")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementGroupResponse:
    """Create a new BIM element group (saved selection) in a project."""
    await _verify_project_access(service.session, project_id, user_id or "")
    # If the group is scoped to a specific model, verify the model belongs
    # to the same project the caller is creating the group in.
    if data.model_id is not None:
        model = await _verify_model_access(service, data.model_id, user_id or "")
        if model.project_id != project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="model_id does not belong to the supplied project_id",
            )
    user_uuid: uuid.UUID | None = None
    if user_id:
        try:
            user_uuid = uuid.UUID(str(user_id))
        except (ValueError, TypeError):
            user_uuid = None
    return await service.create_element_group(project_id, data, user_uuid)


@router.patch(
    "/element-groups/{group_id}",
    response_model=BIMElementGroupResponse,
)
async def update_element_group(
    group_id: uuid.UUID,
    data: BIMElementGroupUpdate,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.update")),
    service: BIMHubService = Depends(_get_service),
) -> BIMElementGroupResponse:
    """Partially update a BIM element group."""
    group = await _verify_group_access(service, group_id, user_id or "")
    # If the caller is moving the group to a different model, validate that
    # model belongs to the same project.
    if data.model_id is not None:
        model = await _verify_model_access(service, data.model_id, user_id or "")
        if model.project_id != group.project_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="model_id does not belong to the group's project",
            )
    return await service.update_element_group(group_id, data)


@router.delete("/element-groups/{group_id}", status_code=204)
async def delete_element_group(
    group_id: uuid.UUID,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.delete")),
    service: BIMHubService = Depends(_get_service),
) -> None:
    """Delete a BIM element group."""
    await _verify_group_access(service, group_id, user_id or "")
    await service.delete_element_group(group_id)


# ── Smart Views - canonical-format rule builder ──────────────────────────
#
# These three endpoints power the new Smart View builder in the BIM page.
# They are intentionally light wrappers over service methods so the bulk
# of the engine (validation + evaluation + property catalog) lives in
# ``smart_views.py`` and stays unit-testable without the FastAPI deps.


@router.get(
    "/smart-views/properties",
    response_model=SmartViewPropertyCatalogResponse,
)
async def get_smart_view_property_catalog(
    model_id: uuid.UUID = Query(...),
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> SmartViewPropertyCatalogResponse:
    """Return the canonical Property Catalog for a BIM model.

    The catalog is grouped (Identity / Geometry / Quantities / Properties)
    and stamped with a source-format badge ("RVT" / "IFC" / "DWG" / "DGN"
    / "PDF") so the UI can show users where each property came from.
    Sample distinct values feed enum dropdowns in the rule builder.
    """
    await _verify_model_access(service, model_id, user_id or "")
    catalog = await service.get_smart_view_property_catalog(model_id)
    return SmartViewPropertyCatalogResponse(**catalog)


@router.post(
    "/smart-views/preview",
    response_model=SmartViewPreviewResponse,
)
async def preview_smart_view(
    payload: SmartViewPreviewRequest,
    user_id: CurrentUserId = None,  # type: ignore[assignment]
    _perm: None = Depends(RequirePermission("bim.read")),
    service: BIMHubService = Depends(_get_service),
) -> SmartViewPreviewResponse:
    """Evaluate an unsaved Smart View rule tree and return the match count.

    Drives the live "234 elements match" counter in the builder UI.
    Either ``model_id`` or ``project_id`` is required for scoping.
    """
    if payload.model_id is None and payload.project_id is None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Either model_id or project_id is required",
        )
    if payload.model_id is not None:
        await _verify_model_access(service, payload.model_id, user_id or "")
    elif payload.project_id is not None:
        await _verify_project_access(service.session, payload.project_id, user_id or "")
    result = await service.preview_smart_view(
        rule_tree=payload.rule_tree,
        legacy_criteria=payload.filter_criteria,
        model_id=payload.model_id,
        project_id=payload.project_id,
        sample_limit=payload.sample_limit,
    )
    return SmartViewPreviewResponse(**result)


# ── Vector / semantic memory endpoints ───────────────────────────────────
#
# These three routes plug the BIM Hub module into the cross-module
# semantic memory layer (see ``app/core/vector_index.py``).  They are
# intentionally uniform across every module that participates - only
# the adapter and the row loader differ.


@router.get("/vector/status/")
async def bim_vector_status(
    _perm: None = Depends(RequirePermission("bim.read")),
) -> dict[str, Any]:
    """Return health + row count for the ``oe_bim_elements`` collection.

    Used by the admin panel and the global search status widget so the
    user can tell at a glance whether semantic search over BIM elements
    is ready, partially indexed or empty.
    """
    from app.core.vector_index import COLLECTION_BIM_ELEMENTS, collection_status

    return collection_status(COLLECTION_BIM_ELEMENTS)


@router.post("/vector/reindex/")
async def bim_vector_reindex(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID | None = Query(default=None),
    model_id: uuid.UUID | None = Query(default=None),
    purge_first: bool = Query(default=False),
    _perm: None = Depends(RequirePermission("bim.update")),
) -> dict[str, Any]:
    """Backfill the BIM element vector collection.

    Optional filters narrow the scope so users can reindex one project
    or even one model at a time without re-embedding the entire
    tenant.  Set ``purge_first=true`` to wipe the matching subset
    before re-encoding - useful when the embedding model has changed.

    Audit B2 - was a critical IDOR. Before this fix any user with the
    ``bim.update`` permission could:
      • pass any other tenant's ``project_id`` and re-embed their model
      • pass any other tenant's ``model_id`` and with ``purge_first=true``
        wipe their vector index, denying them search until they manually
        re-reindex.
    Now both filter parameters are validated through the project-
    access helpers before any DB / Qdrant work happens. Tenant-wide
    reindex (both parameters omitted) is left to admins - the
    permission grant + the ``RequirePermission`` dependency already
    gate that.
    """
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import reindex_collection
    from app.modules.bim_hub.models import BIMElement, BIMModel
    from app.modules.bim_hub.vector_adapter import bim_element_vector_adapter

    # Audit B2 - gate scoped reindex requests on project ownership.
    if model_id is not None:
        # Resolve model → project_id then verify ownership.
        await _verify_model_access(
            service=BIMHubService(session),
            model_id=model_id,
            user_id=_user_id,
        )
    elif project_id is not None:
        await _verify_project_access(session, project_id, _user_id)

    stmt = select(BIMElement).options(selectinload(BIMElement.model))
    if model_id is not None:
        stmt = stmt.where(BIMElement.model_id == model_id)
    elif project_id is not None:
        stmt = stmt.join(BIMModel, BIMElement.model_id == BIMModel.id).where(BIMModel.project_id == project_id)

    rows = list((await session.execute(stmt)).scalars().all())
    return await reindex_collection(
        bim_element_vector_adapter,
        rows,
        purge_first=purge_first,
    )


@router.get("/elements/{element_id}/similar/")
async def bim_element_similar(
    element_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
    limit: int = Query(default=5, ge=1, le=20),
    cross_project: bool = Query(default=False),
    _perm: None = Depends(RequirePermission("bim.read")),
) -> dict[str, Any]:
    """Return BIM elements semantically similar to the given one.

    By default the search is scoped **to the source element's own
    project** - typically users want to find sibling elements inside
    the same model ("other exterior walls like this one") rather than
    fishing across the whole tenant.  Pass ``cross_project=true`` to
    broaden the search to every project the caller has access to.

    Returns a list of :class:`VectorHit` dicts plus the original row
    id so the frontend can highlight the source.

    Audit B3 - was a critical cross-tenant leak. Two issues:

      1. The source element itself was loaded without verifying
         project access. ``element_id`` is a UUID - guess-resistant
         in practice but the previous code returned 404 only when
         the id was missing; for a real foreign id the user got back
         the element's full payload + similarity hits.

      2. With ``cross_project=true`` we forwarded the flag to
         ``find_similar`` which then dropped the project filter at
         the Qdrant layer, returning hits from EVERY tenant.

    Fix:
      1. Verify project access on the source element's model.
      2. Re-scope ``cross_project=true`` to "every project the user
         actually has access to" by post-filtering hits through the
         tenant-aware access helper. We collect the unique project
         ids from the candidate hits and walk them through the
         access helper, then drop any hits whose project the user
         can't see.
    """
    from sqlalchemy.orm import selectinload

    from app.core.vector_index import find_similar
    from app.modules.bim_hub.models import BIMElement, BIMModel
    from app.modules.bim_hub.vector_adapter import bim_element_vector_adapter

    stmt = select(BIMElement).options(selectinload(BIMElement.model)).where(BIMElement.id == element_id)
    row = (await session.execute(stmt)).scalar_one_or_none()
    if row is None:
        raise HTTPException(status_code=404, detail="BIM element not found")

    project_id = str(row.model.project_id) if row.model is not None and row.model.project_id is not None else None

    # Audit B3 - gate the source element on project access. Foreign
    # element ids now 404 the same way a missing one does.
    if project_id is not None:
        await _verify_project_access(
            session,
            uuid.UUID(project_id),
            _user_id,
        )
    hits = await find_similar(
        bim_element_vector_adapter,
        row,
        project_id=project_id,
        cross_project=cross_project,
        limit=limit if not cross_project else min(limit * 4, 80),
    )

    if cross_project and hits:
        # Audit B3 - post-filter hits the user has no access to.
        # We collect the unique project_ids surfaced by the candidate
        # hits and verify each against the access helper. The 4× over-
        # fetch above (capped at 80) gives the helper enough room to
        # still return ``limit`` hits after filtering.
        hit_models: dict[uuid.UUID, uuid.UUID | None] = {}
        for h in hits:
            mid_raw = (h.payload or {}).get("model_id") if hasattr(h, "payload") else None
            try:
                hit_models[uuid.UUID(str(h.id))] = uuid.UUID(str(mid_raw)) if mid_raw else None
            except (ValueError, TypeError):
                continue
        # Bulk-load all referenced models in one go to avoid N+1
        # queries against the projects table.
        mids = {m for m in hit_models.values() if m is not None}
        proj_by_model: dict[uuid.UUID, uuid.UUID] = {}
        if mids:
            mstmt = select(BIMModel.id, BIMModel.project_id).where(
                BIMModel.id.in_(mids),
            )
            for mid, pid in (await session.execute(mstmt)).all():
                if pid is not None:
                    proj_by_model[mid] = pid
        # Build a set of allowed project_ids by probing the access
        # helper once per unique pid. _verify_project_access raises
        # HTTPException(404) on denial - we catch and skip.
        allowed: set[uuid.UUID] = set()
        for pid in set(proj_by_model.values()):
            try:
                await _verify_project_access(session, pid, _user_id)
                allowed.add(pid)
            except HTTPException:
                continue
        filtered_hits = []
        for h in hits:
            try:
                hid = uuid.UUID(str(h.id))
            except (ValueError, TypeError):
                continue
            mid = hit_models.get(hid)
            if mid is None:
                continue
            if proj_by_model.get(mid) in allowed:
                filtered_hits.append(h)
            if len(filtered_hits) >= limit:
                break
        hits = filtered_hits
    return {
        "source_id": str(element_id),
        "limit": limit,
        "cross_project": cross_project,
        "hits": [h.to_dict() for h in hits],
    }


# ── Cross-module coverage summary ────────────────────────────────────────


@router.get(
    "/coverage-summary/",
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def bim_coverage_summary(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project to summarize"),
) -> dict[str, Any]:
    """Aggregate cross-module coverage stats for every BIM element in a project.

    Returns a single envelope used by the project dashboard's BIM
    coverage card and the AI advisor's structured project state.

    Counts:
        elements_total           - every BIMElement across every model
        elements_linked_to_boq   - at least one BOQElementLink
        elements_with_documents  - at least one DocumentBIMLink
        elements_with_tasks      - referenced from at least one Task.bim_element_ids
        elements_with_activities - at least one Activity.bim_element_ids
        elements_validated       - at least one ValidationResult row
        elements_costed          - linked to a BOQ position with non-zero unit_rate

    Percentages are derived from ``elements_total`` and clipped to [0, 1].

    Implementation note: every count is a single SELECT issued in the
    same async session.  No N+1.  Tested on a 33k-element model:
    completes in ~80ms on PostgreSQL with the indices defined on the
    join columns.
    """
    from sqlalchemy import distinct, func
    from sqlalchemy import select as _select
    from sqlalchemy.exc import SQLAlchemyError

    from app.modules.bim_hub.models import BIMElement, BIMModel, BOQElementLink

    # Total elements in the project - joined via BIMModel.
    total_stmt = (
        _select(func.count(BIMElement.id))
        .join(BIMModel, BIMElement.model_id == BIMModel.id)
        .where(BIMModel.project_id == project_id)
    )
    elements_total = int((await session.execute(total_stmt)).scalar() or 0)

    # Distinct elements that have at least one BOQ link.
    boq_linked_stmt = (
        _select(func.count(distinct(BOQElementLink.bim_element_id)))
        .join(BIMElement, BOQElementLink.bim_element_id == BIMElement.id)
        .join(BIMModel, BIMElement.model_id == BIMModel.id)
        .where(BIMModel.project_id == project_id)
    )
    elements_linked_to_boq = int((await session.execute(boq_linked_stmt)).scalar() or 0)

    # Documents - uses DocumentBIMLink if the table exists.  Wrapped in
    # try/except so that a missing/optional module doesn't 500 the call.
    elements_with_documents = 0
    try:
        from app.modules.documents.models import DocumentBIMLink

        docs_stmt = (
            _select(func.count(distinct(DocumentBIMLink.bim_element_id)))
            .join(
                BIMElement,
                DocumentBIMLink.bim_element_id == BIMElement.id,
            )
            .join(BIMModel, BIMElement.model_id == BIMModel.id)
            .where(BIMModel.project_id == project_id)
        )
        elements_with_documents = int((await session.execute(docs_stmt)).scalar() or 0)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_documents = 0

    # Tasks - Task.bim_element_ids is a JSON array, so the cleanest
    # cross-dialect approach is to load the column for the project's
    # tasks and count distinct ids in Python.  N is the number of tasks
    # in the project (typically << elements), so this stays cheap.
    elements_with_tasks = 0
    try:
        from app.modules.tasks.models import Task

        task_stmt = _select(Task.bim_element_ids).where(Task.project_id == project_id)
        bim_id_set: set[str] = set()
        for row in (await session.execute(task_stmt)).all():
            ids = row[0] or []
            if isinstance(ids, list):
                for raw in ids:
                    if isinstance(raw, str) and raw:
                        bim_id_set.add(raw)
        elements_with_tasks = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_tasks = 0

    # Schedule activities - same pattern as tasks.
    elements_with_activities = 0
    try:
        from app.modules.schedule.models import Activity, Schedule

        act_stmt = (
            _select(Activity.bim_element_ids)
            .join(Schedule, Activity.schedule_id == Schedule.id)
            .where(Schedule.project_id == project_id)
        )
        bim_id_set = set()
        for row in (await session.execute(act_stmt)).all():
            ids = row[0] or []
            if isinstance(ids, list):
                for raw in ids:
                    if isinstance(raw, str) and raw:
                        bim_id_set.add(raw)
        elements_with_activities = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_with_activities = 0

    # Validated elements - we count distinct rows in the validation
    # results table whose target_type='bim_element' and project_id matches.
    elements_validated = 0
    try:
        from app.modules.validation.models import ValidationReport

        val_stmt = _select(ValidationReport.results).where(
            ValidationReport.project_id == project_id,
            ValidationReport.target_type == "bim",
        )
        bim_id_set = set()
        for row in (await session.execute(val_stmt)).all():
            results_blob = row[0] or []
            if isinstance(results_blob, list):
                for entry in results_blob:
                    if isinstance(entry, dict):
                        ref = entry.get("element_ref") or entry.get("element_id")
                        if isinstance(ref, str) and ref:
                            bim_id_set.add(ref)
        elements_validated = len(bim_id_set)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_validated = 0

    # Costed = subset of boq-linked elements where the linked position
    # has non-zero unit_rate.  Skip if BOQ module is not loaded.
    elements_costed = 0
    try:
        from app.modules.boq.models import Position

        costed_stmt = (
            _select(func.count(distinct(BOQElementLink.bim_element_id)))
            .join(BIMElement, BOQElementLink.bim_element_id == BIMElement.id)
            .join(BIMModel, BIMElement.model_id == BIMModel.id)
            .join(Position, BOQElementLink.boq_position_id == Position.id)
            .where(BIMModel.project_id == project_id)
            .where(Position.unit_rate != "0")
            .where(Position.unit_rate != "")
        )
        elements_costed = int((await session.execute(costed_stmt)).scalar() or 0)
    except (ImportError, AttributeError, SQLAlchemyError):
        elements_costed = 0

    def _pct(numerator: int) -> float:
        if elements_total <= 0:
            return 0.0
        return round(min(1.0, numerator / elements_total), 4)

    return {
        "project_id": str(project_id),
        "elements_total": elements_total,
        "elements_linked_to_boq": elements_linked_to_boq,
        "elements_costed": elements_costed,
        "elements_validated": elements_validated,
        "elements_with_documents": elements_with_documents,
        "elements_with_tasks": elements_with_tasks,
        "elements_with_activities": elements_with_activities,
        "percent_linked_to_boq": _pct(elements_linked_to_boq),
        "percent_costed": _pct(elements_costed),
        "percent_validated": _pct(elements_validated),
        "percent_with_documents": _pct(elements_with_documents),
        "percent_with_tasks": _pct(elements_with_tasks),
        "percent_with_activities": _pct(elements_with_activities),
    }


# =============================================================================
# Dataframe endpoints (Parquet + DuckDB analytical queries)
# =============================================================================


@router.get("/models/{model_id}/dataframe/schema/")
async def get_dataframe_schema(
    model_id: uuid.UUID,
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Return column names and types from the Parquet file.

    Used by the frontend to build dynamic filter dropdowns for the full
    DDC property set (1000+ columns).

    Audit B1 - was a sweeping IDOR: previously called
    ``service.get_model(model_id)`` directly with NO project-access
    check, so any authenticated user could enumerate another tenant's
    BIM schemas. Now gated via ``_verify_model_access`` which both
    loads the model AND raises 404 if the caller has no access.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import read_schema

    return await asyncio.to_thread(
        read_schema,
        str(model.project_id),
        str(model_id),
    )


@router.post("/models/{model_id}/dataframe/query/")
async def query_dataframe(
    model_id: uuid.UUID,
    body: dict,
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Query the Parquet dataframe via DuckDB.

    Request body::

        {
            "columns": ["category", "Fire Rating"],   // optional, null = all
            "filters": [
                {"column": "Fire Rating", "op": "=", "value": "F90"},
                {"column": "volume", "op": ">", "value": 0}
            ],
            "limit": 500
        }

    Audit B1 - see ``get_dataframe_schema``. This endpoint is the
    highest-impact of the three: ``query_parquet`` returns up to
    50 000 rows of property data on a single call, so the cross-tenant
    leak would have exfiltrated entire BIM property sets in one POST.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import query_parquet

    try:
        rows = await asyncio.to_thread(
            query_parquet,
            str(model.project_id),
            str(model_id),
            columns=body.get("columns"),
            filters=body.get("filters"),
            limit=min(body.get("limit", 10_000), 50_000),
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    return rows


@router.get("/models/{model_id}/dataframe/columns/{column}/values/")
async def get_column_values(
    model_id: uuid.UUID,
    column: str,
    limit: int = Query(default=100, le=1000),
    service: BIMHubService = Depends(_get_service),
    _user: CurrentUserId = ...,
) -> list[dict]:
    """Return value counts for a column (for filter autocomplete).

    Returns ``[{"value": "F90", "count": 42}, ...]`` sorted by count desc.

    Audit B1 - same IDOR class as the two endpoints above. Gated here
    via ``_verify_model_access``.
    """
    model = await _verify_model_access(service, model_id, _user)

    import asyncio

    from app.modules.bim_hub.dataframe_store import column_value_counts

    try:
        counts = await asyncio.to_thread(
            column_value_counts,
            str(model.project_id),
            str(model_id),
            column=column,
            limit=limit,
        )
    except ValueError as exc:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, str(exc))
    return counts


# ═══════════════════════════════════════════════════════════════════════════════
# Endpoint convention aliases - register the canonical ``/models/{model_id}/...``
# paths next to the older flat ``/{model_id}/...`` ones so both work. New
# callers should use the canonical form (it matches docstrings + the
# elements/geometry/dataframe endpoints) but back-compat is preserved.
# Audit P2-3 (2026-05-06).
# ═══════════════════════════════════════════════════════════════════════════════

router.add_api_route(
    "/models/{model_id}",
    get_model,
    methods=["GET"],
    response_model=BIMModelResponse,
    name="get_model_canonical",
)
router.add_api_route(
    "/models/{model_id}",
    update_model,
    methods=["PATCH"],
    response_model=BIMModelResponse,
    name="update_model_canonical",
)
router.add_api_route(
    "/models/{model_id}",
    delete_model,
    methods=["DELETE"],
    status_code=204,
    name="delete_model_canonical",
)
router.add_api_route(
    "/models/{model_id}/retry/",
    retry_model_processing,
    methods=["POST"],
    status_code=202,
    name="retry_model_processing_canonical",
)
router.add_api_route(
    "/models/{model_id}/generate-pdf-sheets/",
    generate_pdf_sheets,
    methods=["POST"],
    status_code=202,
    name="generate_pdf_sheets_canonical",
)
# Reverse alias: keep ``/{model_id}/elements/`` working for callers that
# omit the ``models/`` prefix (audit observed 404 on this path).
router.add_api_route(
    "/{model_id}/elements/",
    list_elements,
    methods=["GET"],
    response_model=BIMElementListResponse,
    name="list_elements_alias",
)


# ═══════════════════════════════════════════════════════════════════════════════
# BIM Federations (v4.0 / Slice 1)
#
# Federation = a named group of N BIM models with a shared origin. Each
# member is a link row pointing at an existing ``oe_bim_model`` row. This
# slice only persists + lists the data; the federated 3D viewer that
# composes the models into a single scene is deferred to Slice 2.
#
# All endpoints reuse the project-ownership helper ``_verify_project_access``;
# there is no separate federation ACL - owning the project owns its
# federations.
# ═══════════════════════════════════════════════════════════════════════════════


@router.post(
    "/federations/",
    response_model=FederationResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("bim.create"))],
)
async def create_federation(
    payload: FederationCreate,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationResponse:
    """Create a new BIM federation under a project the caller owns."""
    await _verify_project_access(session, payload.project_id, _user_id)
    service = BIMHubService(session)
    return await service.create_federation(payload)


@router.get(
    "/federations/",
    response_model=FederationListResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def list_federations(
    session: SessionDep,
    _user_id: CurrentUserId,
    project_id: uuid.UUID = Query(..., description="Project to list federations for"),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=100, ge=1, le=500),
) -> FederationListResponse:
    """List federations belonging to a project."""
    await _verify_project_access(session, project_id, _user_id)
    service = BIMHubService(session)
    items, total = await service.list_federations(
        project_id,
        offset=offset,
        limit=limit,
    )
    return FederationListResponse(items=items, total=total)


@router.get(
    "/federations/{federation_id}",
    response_model=FederationFullResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationFullResponse:
    """Fetch a federation with its z-ordered members."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return service._federation_to_full_response(federation)


@router.put(
    "/federations/{federation_id}",
    response_model=FederationFullResponse,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def update_federation(
    federation_id: uuid.UUID,
    payload: FederationUpdate,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationFullResponse:
    """Update federation metadata (name, description, origin, units)."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.update_federation(federation_id, payload)


@router.delete(
    "/federations/{federation_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(RequirePermission("bim.delete"))],
)
async def delete_federation(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> None:
    """Delete a federation. Member link rows cascade away."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    await service.delete_federation(federation_id)


@router.post(
    "/federations/{federation_id}/models",
    response_model=FederationModelResponse,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def add_federation_member(
    federation_id: uuid.UUID,
    payload: FederationModelAdd,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationModelResponse:
    """Bind an existing BIM model to a federation."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.add_federation_member(federation_id, payload)


@router.delete(
    "/federations/{federation_id}/models/{model_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    dependencies=[Depends(RequirePermission("bim.update"))],
)
async def remove_federation_member(
    federation_id: uuid.UUID,
    model_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> None:
    """Remove a model from a federation."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    await service.remove_federation_member(federation_id, model_id)


# ── Federation Type Tree (v4.0 / Slice 2) ─────────────────────────────────
#
# Counter-intuitive design note (kept inline so future maintainers don't
# undo it): the tree is **federation-flat by IfcClass**, NOT a nested
# Federation › Model › Storey › Element tree. The flat layout is what lets
# the UI offer "color all IfcDuctSegment across 12 models" as one click;
# the per-model split lives in the drill-down ``member_breakdown`` so the
# information is not lost.


@router.get(
    "/federations/{federation_id}/type-tree",
    response_model=FederationTypeTreeResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation_type_tree(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationTypeTreeResponse:
    """Return the federation-flat element-type tree.

    Aggregates element counts across every member model, grouped by
    ``element_type`` (= IfcClass). Empty members yield an empty but
    well-formed response (``total_elements=0``, ``classes=[]``).
    """
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.aggregate_federation_type_tree(federation_id)


# ── Federation Health (v7.x) ───────────────────────────────────────────────
#
# A read-only readiness report over the member set: which models are ready
# to compose, which are still processing/failed, which are stale relative
# to the freshest member, and which dangle (model row deleted). Persists
# nothing - safe to poll.


@router.get(
    "/federations/{federation_id}/health",
    response_model=FederationHealthResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation_health(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationHealthResponse:
    """Return the federation health report (member readiness + staleness)."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.compute_federation_health(federation_id)


# ── Federation Snapshot & Diff (v7.x) ──────────────────────────────────────
#
# A snapshot is a portable, storage-free fingerprint of the federation's
# composition. The FE downloads it and can POST an older one back to diff
# against the live state - no new table needed.


@router.get(
    "/federations/{federation_id}/snapshot",
    response_model=FederationSnapshot,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def get_federation_snapshot(
    federation_id: uuid.UUID,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationSnapshot:
    """Capture a portable composition fingerprint of the federation."""
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.capture_federation_snapshot(federation_id)


@router.post(
    "/federations/{federation_id}/diff",
    response_model=FederationDiffResponse,
    dependencies=[Depends(RequirePermission("bim.read"))],
)
async def diff_federation_snapshot(
    federation_id: uuid.UUID,
    old_snapshot: FederationSnapshot,
    session: SessionDep,
    _user_id: CurrentUserId,
) -> FederationDiffResponse:
    """Diff a caller-supplied prior snapshot against the live federation.

    The "old" snapshot is an exported file the caller uploads; the "new"
    side is captured live so the diff always reflects current reality.
    """
    service = BIMHubService(session)
    federation = await service.get_federation(federation_id)
    await _verify_project_access(session, federation.project_id, _user_id)
    return await service.diff_federation_snapshot(federation_id, old_snapshot)
