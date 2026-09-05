# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Downloadable export formats for generated reports (PDF / XLSX / CSV).

Background
==========
Until this module existed the reporting feature could only hand back an
HTML body (``GET /reports/{id}/content`` returns ``text/html``). The
generated-report row carried a ``format`` column of ``pdf`` / ``excel`` /
``html`` but nothing ever produced a real binary in those formats - a user
who picked "Excel" still only had HTML to look at. This file closes that
gap by turning a report's ``data_snapshot`` (the same per-section dict the
:class:`~app.modules.reporting.renderer.ReportRenderer` consumes) into a
real downloadable file.

Design notes
============
- **Reuse, do not reinvent.** The PDF path uses the *same* reportlab
  platypus stack and the bundled Unicode fonts that the rest of the
  platform already ships (see ``app.modules.boq.pdf_export`` and
  ``app.core.pdf_fonts``). No second PDF library is introduced. The XLSX
  path uses ``openpyxl`` (already a dependency, used by the BOQ export at
  ``backend/app/modules/boq/router.py``). CSV uses the stdlib ``csv``
  module. Nothing heavy is added.
- **CSV / spreadsheet injection.** Every user-controlled string written
  into a CSV or XLSX cell is routed through
  :func:`app.core.csv_safety.neutralise_formula`, exactly as the BOQ
  exporters do.
- **Money stays Decimal-correct.** The snapshot already carries money as
  strings (e.g. ``"517103508.65 EUR"``) assembled by the service layer's
  ``_build_default_snapshot`` / retainage roll-ups. We never coerce those
  through ``float`` for display - they are emitted verbatim. Where a value
  is a bare numeric string we hand openpyxl a :class:`decimal.Decimal` so
  Excel stores it as a sortable number without precision loss.
- **One currency per report.** The resolved ISO 4217 code is stamped on the
  report row (``GeneratedReport.currency``) and into the snapshot; the
  exporters surface it in the header so every figure reads in one currency.

Public API
==========

>>> from app.modules.reporting.exporters import export_report
>>> filename, media_type, blob = export_report(
...     fmt="xlsx",
...     report_type="cost_report",
...     title="Q1 Cost Report",
...     project_name="Skyline Tower",
...     currency="EUR",
...     generated_at="2026-06-14T10:00:00",
...     template_data={"sections": [...]},
...     data_snapshot={"summary": {...}, "breakdown": [...]},
... )

The function is sync and pure - no DB, no network, no clock. The service
layer assembles the snapshot and resolves the currency before calling it.

Language
========
Pass ``locale=`` to write the file in one of
:data:`~app.modules.reporting.report_translations.SUPPORTED_REPORT_LOCALES`.
Section headings, field labels, column headers, ``Yes`` / ``No`` and the
metadata preamble all come from that catalogue, and the route serving the
file declares what was produced in ``Content-Language``. COBie is excluded
on purpose: its sheet and column names are fixed by the handover standard,
so translating them would produce a file no COBie consumer can read.
"""

from __future__ import annotations

import csv
import html
import io
from decimal import Decimal, InvalidOperation
from typing import Any

from app.core.csv_safety import neutralise_formula
from app.core.evidence import evidence_header
from app.modules.reporting.report_translations import (
    DEFAULT_REPORT_LOCALE,
    field_label,
    normalize_report_locale,
    section_title,
    tr,
)

__all__ = [
    "COBIE_ADDITIONAL_SHEETS",
    "COBIE_MEDIA_TYPE",
    "ExportFormatError",
    "SUPPORTED_FORMATS",
    "export_project_cobie",
    "export_report",
]


def _evidence_rows(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    data_snapshot: dict[str, Any] | None,
    locale: str,
) -> list[tuple[str, str]]:
    """Tamper-evident header rows: generation time + content digest.

    The digest fingerprints the report content (type, title, project,
    currency, section snapshot) but not the timestamp, so re-exporting the
    same data reproduces the same digest for verification. Rendered by each
    format's own header block. See :mod:`app.core.evidence`.

    ``evidence_header`` has always taken a ``locale``; reporting never passed
    one, so these two labels stayed English in a document whose every other
    label is now written in the reader's language.
    """
    payload = {
        "report_type": report_type,
        "title": title,
        "project_name": project_name,
        "currency": currency,
        "data": data_snapshot or {},
    }
    return evidence_header(generated_at=generated_at, payload=payload, locale=locale)


# Formats this module can produce. ``html`` is included so the download
# endpoint can serve the existing HTML body through the same code path and
# nothing regresses, but the heavy lifting here is pdf / xlsx / csv.
SUPPORTED_FORMATS: tuple[str, ...] = ("pdf", "xlsx", "csv", "html")

# MIME types keyed by format. Matches the BOQ export endpoints
# (``backend/app/modules/boq/router.py``) so the browser downloads a real
# file rather than rendering it inline.
_MEDIA_TYPES: dict[str, str] = {
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "html": "text/html; charset=utf-8",
}

# COBie is a distinct export profile: it is not a projection of a
# GeneratedReport's ``data_snapshot`` (a per-section KV/record-list bag), it
# is a projection of the platform's canonical BIM asset register (BIMModel +
# BIMElement). It is registered here, alongside the report exporters, with
# the same ``(filename, media_type, blob)`` return shape as ``export_report``
# so callers do not need to special-case it, but it is invoked through its
# own function (``export_project_cobie``) rather than through the
# ``fmt=`` dispatch of ``export_report`` because its inputs differ.
COBIE_MEDIA_TYPE = _MEDIA_TYPES["xlsx"]


class ExportFormatError(ValueError):
    """Raised when an unsupported export format is requested."""


# ── Section resolution ────────────────────────────────────────────────────
#
# Mirror the renderer's precedence so a downloaded file shows exactly the
# sections the HTML view shows: template sections first, else the built-in
# default list for the report type, else a single generic Summary block.


def _resolve_sections(
    report_type: str,
    template_data: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    """Return the ordered section list to render (template > default > generic).

    Imports the renderer's default-section map so the two stay in lockstep -
    we deliberately do not maintain a second copy of the section ordering.
    """
    if isinstance(template_data, dict):
        sections = template_data.get("sections")
        if isinstance(sections, list) and sections:
            resolved = [s for s in sections if isinstance(s, dict) and s.get("id")]
            if resolved:
                return resolved

    from app.modules.reporting.renderer import _DEFAULT_SECTIONS

    return _DEFAULT_SECTIONS.get(report_type, [{"id": "summary", "title": "Summary"}])


def _section_title(section: dict[str, Any], locale: str) -> str:
    """Human-readable section title, falling back to a Title-Cased id.

    Localized through the same table the HTML renderer uses, so a downloaded
    file and the on-screen view name a section identically.
    """
    sid = str(section.get("id", "")).strip()
    return section_title(str(section.get("title", sid.replace("_", " ").title())), locale)


def _stringify(value: Any, locale: str) -> str:
    """Stringify a scalar snapshot value for a flat text cell.

    Booleans become Yes/No (matching the renderer); ``None`` becomes an
    empty string. Dicts / lists are summarised compactly because flat CSV /
    PDF table cells cannot nest a sub-table the way the HTML renderer can.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return tr(locale, "yes" if value else "no")
    if isinstance(value, dict):
        # Compact "k: v" join so a nested object still conveys its content
        # in a single cell rather than printing a Python repr.
        return "; ".join(f"{field_label(k, locale)}: {_stringify(v, locale)}" for k, v in value.items())
    if isinstance(value, list):
        return "; ".join(_stringify(v, locale) for v in value)
    return str(value)


def _looks_numeric(text: str) -> bool:
    """True when *text* is a bare number openpyxl can store as a Decimal.

    A money string like ``"1234.56 EUR"`` is intentionally NOT numeric here
    (it carries a currency suffix and should stay text); only a clean
    ``"1234.56"`` / ``"-3"`` qualifies.
    """
    if not text:
        return False
    # An all-digit identifier ("007", "012", a zero-padded ordinal or cost
    # code) must stay text in the spreadsheet: coercing it to a number drops
    # the leading zeros and corrupts the value. A genuine number never carries
    # a significant leading zero (only "0" itself or "0.x").
    s = text[1:] if text[:1] in "+-" else text
    if len(s) > 1 and s[0] == "0" and s[1].isdigit():
        return False
    try:
        d = Decimal(text)
    except (InvalidOperation, ValueError):
        return False
    return d.is_finite()


# ── Section shape detection ─────────────────────────────────────────────────
#
# A section payload renders as one of two table shapes (mirroring the HTML
# renderer): a key/value definition table (dict, or list of scalars) or a
# columnar table (list of dicts, e.g. a cost breakdown / incident log). The
# helpers below classify the payload and materialise the matching shape so
# every export format presents the same structure.


def _is_record_list(payload: Any) -> bool:
    """True when *payload* is a non-empty list whose items are all dicts.

    This is the "columnar table" shape (e.g. ``breakdown``: a list of
    ``{"trade": ..., "amount": ...}`` records).
    """
    return isinstance(payload, list) and len(payload) > 0 and all(isinstance(it, dict) for it in payload)


def _record_columns(records: list[dict[str, Any]]) -> list[str]:
    """Ordered union of keys across record dicts (first-seen order)."""
    columns: list[str] = []
    seen: set[str] = set()
    for item in records:
        for key in item:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    return columns


def _flatten_keyvalue(payload: Any, locale: str) -> list[tuple[str, str]]:
    """Flatten a dict / scalar / scalar-list into ``(label, value)`` rows.

    Money / Decimal values pass through ``_stringify`` unchanged so no float
    rounding is ever introduced. Used for the key/value table shape; record
    lists are handled separately via :func:`_record_columns`.
    """
    rows: list[tuple[str, str]] = []
    if isinstance(payload, dict):
        for key, value in payload.items():
            rows.append((field_label(key, locale), _stringify(value, locale)))
    elif isinstance(payload, list):
        for item in payload:
            rows.append(("", _stringify(item, locale)))
    else:
        rows.append(("", _stringify(payload, locale)))
    return rows


def _section_is_empty(payload: Any) -> bool:
    """Mirror the renderer's skip rule: None / empty dict / empty list."""
    if payload is None:
        return True
    return bool(isinstance(payload, dict | list) and not payload)


# ── Export size guard ────────────────────────────────────────────────────────
#
# The CSV / XLSX / PDF writers below assemble the whole document in memory, and
# openpyxl in particular holds a Python cell object per value. A snapshot with a
# runaway row count (a project with a very large BoQ or incident log) can push
# the single worker into swap while building the download, so the public entry
# point rejects an oversized snapshot with a clear error before any writer runs.
#
# The cap is deliberately generous: a normal report is a handful of sections
# with tens to low-hundreds of rows, so only a pathological snapshot is turned
# away. This is a front-door guard, not a change to how any report is built.
_MAX_EXPORT_ROWS = 200_000


def _snapshot_row_count(data_snapshot: dict[str, Any] | None) -> int:
    """Total table rows an export would materialise from *data_snapshot*.

    Sums the rows every populated section contributes, mirroring the two table
    shapes the writers produce: a record list or scalar list counts one row per
    item, a dict counts one row per entry, a bare scalar is a single row. The
    count is a safe upper bound on what the CSV / XLSX / PDF writers build in
    memory and is computed without materialising a single cell.
    """
    if not isinstance(data_snapshot, dict):
        return 0
    total = 0
    for payload in data_snapshot.values():
        if _section_is_empty(payload):
            continue
        total += len(payload) if isinstance(payload, dict | list) else 1
    return total


# ── CSV export ──────────────────────────────────────────────────────────────


def _export_csv(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
    locale: str,
) -> bytes:
    """Render the report snapshot as a CSV file.

    Layout: a small metadata preamble, then for each populated section a
    blank separator, a section-title row, and ``Field,Value`` rows. Every
    user-controlled string is neutralised against spreadsheet formula
    injection.
    """
    snapshot = data_snapshot or {}
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Metadata preamble - first-party labels, dynamic values neutralised.
    writer.writerow([tr(locale, "meta_report"), neutralise_formula(title)])
    writer.writerow([tr(locale, "meta_project"), neutralise_formula(project_name)])
    writer.writerow([tr(locale, "meta_type"), neutralise_formula(report_type)])
    if currency:
        writer.writerow([tr(locale, "meta_currency"), neutralise_formula(currency)])
    for label, value in _evidence_rows(
        report_type=report_type,
        title=title,
        project_name=project_name,
        currency=currency,
        generated_at=generated_at,
        data_snapshot=data_snapshot,
        locale=locale,
    ):
        writer.writerow([neutralise_formula(label), neutralise_formula(value)])

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True
        writer.writerow([])
        writer.writerow([neutralise_formula(_section_title(section, locale))])
        if _is_record_list(payload):
            # Columnar table (e.g. cost breakdown): one header row of the
            # union of keys, then a row per record. Every cell neutralised.
            columns = _record_columns(payload)
            writer.writerow([neutralise_formula(field_label(c, locale)) for c in columns])
            for item in payload:
                writer.writerow([neutralise_formula(_stringify(item.get(c), locale)) for c in columns])
        else:
            writer.writerow([tr(locale, "field_column"), tr(locale, "value_column")])
            for label, value in _flatten_keyvalue(payload, locale):
                writer.writerow([neutralise_formula(label), neutralise_formula(value)])

    if not rendered_any:
        writer.writerow([])
        writer.writerow([tr(locale, "no_data_heading")])

    return buffer.getvalue().encode("utf-8-sig")


# ── XLSX export ─────────────────────────────────────────────────────────────


def _export_xlsx(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
    locale: str,
) -> bytes:
    """Render the report snapshot as a formatted .xlsx workbook.

    One worksheet: a metadata header block, then a two-column
    ``Field | Value`` table per populated section with a shaded section
    heading. Bare numeric values are written as :class:`Decimal` so Excel
    stores a sortable number; everything else is neutralised text.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    snapshot = data_snapshot or {}

    wb = Workbook()
    ws = wb.active
    ws.title = tr(locale, "worksheet_name")

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="E8E8EE", end_color="E8E8EE", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    row = 1
    # ── Title + metadata block ──
    cell = ws.cell(row=row, column=1, value=neutralise_formula(title))
    cell.font = title_font
    row += 1

    meta_rows = [
        (tr(locale, "meta_project"), project_name),
        (tr(locale, "meta_type"), report_type),
    ]
    if currency:
        meta_rows.append((tr(locale, "meta_currency"), currency))
    meta_rows.extend(
        _evidence_rows(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            data_snapshot=data_snapshot,
            locale=locale,
        )
    )
    for label, value in meta_rows:
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = bold
        ws.cell(row=row, column=2, value=neutralise_formula(value))
        row += 1
    row += 1  # blank spacer

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True

        if _is_record_list(payload):
            columns = _record_columns(payload)
            n_cols = max(len(columns), 1)

            # Section heading shaded across the table width.
            sec_cell = ws.cell(row=row, column=1, value=neutralise_formula(_section_title(section, locale)))
            sec_cell.font = section_font
            sec_cell.fill = section_fill
            for c in range(2, n_cols + 1):
                ws.cell(row=row, column=c).fill = section_fill
            row += 1

            # Header row (one cell per column).
            for c_idx, col in enumerate(columns, start=1):
                hc = ws.cell(row=row, column=c_idx, value=neutralise_formula(field_label(col, locale)))
                hc.font = header_font
                hc.fill = header_fill
            row += 1

            for item in payload:
                for c_idx, col in enumerate(columns, start=1):
                    cell_val = _stringify(item.get(col), locale)
                    if _looks_numeric(cell_val):
                        ws.cell(row=row, column=c_idx, value=Decimal(cell_val))
                    else:
                        vc = ws.cell(row=row, column=c_idx, value=neutralise_formula(cell_val))
                        vc.alignment = wrap
                row += 1
            row += 1  # blank spacer between sections
            continue

        # Key/value table shape.
        sec_cell = ws.cell(row=row, column=1, value=neutralise_formula(_section_title(section, locale)))
        sec_cell.font = section_font
        sec_cell.fill = section_fill
        ws.cell(row=row, column=2).fill = section_fill
        row += 1

        h1 = ws.cell(row=row, column=1, value=tr(locale, "field_column"))
        h1.font = header_font
        h1.fill = header_fill
        h2 = ws.cell(row=row, column=2, value=tr(locale, "value_column"))
        h2.font = header_font
        h2.fill = header_fill
        row += 1

        for label, value in _flatten_keyvalue(payload, locale):
            ws.cell(row=row, column=1, value=neutralise_formula(label))
            # Store clean numbers as Decimal so Excel treats them as numeric
            # (sortable, summable) without the lossy float roundtrip; money
            # strings with a currency suffix stay text.
            if _looks_numeric(value):
                ws.cell(row=row, column=2, value=Decimal(value))
            else:
                vcell = ws.cell(row=row, column=2, value=neutralise_formula(value))
                vcell.alignment = wrap
            row += 1
        row += 1  # blank spacer between sections

    if not rendered_any:
        ws.cell(row=row, column=1, value=tr(locale, "no_data_heading")).font = bold

    # Reasonable column widths so the file opens readable.
    ws.column_dimensions[get_column_letter(1)].width = 32
    ws.column_dimensions[get_column_letter(2)].width = 60
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── PDF export ────────────────────────────────────────────────────────────────


def _export_pdf(
    *,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
    locale: str,
) -> bytes:
    """Render an executive-summary PDF using the platform reportlab stack.

    Reuses ``app.core.pdf_fonts`` (the same bundled DejaVu Unicode faces the
    BOQ PDF export uses) and reportlab platypus. The layout is a titled
    cover header followed by one ``Field | Value`` table per populated
    section - a clean, correct executive summary rather than a faked binary.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.core.pdf_branding import branded_doc_metadata, branded_header_footer
    from app.core.pdf_fonts import BODY_FONT, BOLD_FONT, pdf_style_for_text, register_pdf_fonts

    register_pdf_fonts()

    snapshot = data_snapshot or {}
    base = getSampleStyleSheet()
    styles = {
        "title": ParagraphStyle(
            "RepTitle",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=18,
            textColor=colors.HexColor("#16213e"),
            spaceAfter=4 * mm,
        ),
        "meta": ParagraphStyle(
            "RepMeta",
            parent=base["Normal"],
            fontName=BODY_FONT,
            fontSize=9,
            textColor=colors.HexColor("#555555"),
            spaceAfter=1 * mm,
        ),
        "section": ParagraphStyle(
            "RepSection",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=13,
            textColor=colors.HexColor("#1d4ed8"),
            spaceBefore=6 * mm,
            spaceAfter=2 * mm,
        ),
        "label": ParagraphStyle(
            "RepLabel",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=9,
            textColor=colors.HexColor("#333333"),
            leading=12,
        ),
        "value": ParagraphStyle(
            "RepValue",
            parent=base["Normal"],
            fontName=BODY_FONT,
            fontSize=9,
            textColor=colors.HexColor("#111111"),
            alignment=TA_LEFT,
            leading=12,
        ),
        # The header row of a record list table, which is filled #1a1a2e. A
        # TableStyle TEXTCOLOR cannot reach a cell holding a Paragraph, so
        # asking for white there left the row drawn in the label colour,
        # #333333 on near black.
        "column_header": ParagraphStyle(
            "RepColumnHeader",
            parent=base["Normal"],
            fontName=BOLD_FONT,
            fontSize=9,
            textColor=colors.white,
            leading=12,
        ),
    }

    def _p(text: Any, style_key: str) -> Paragraph:
        """Escape arbitrary text and wrap it in a paragraph (XSS/markup safe).

        Mirrors ``boq.pdf_export._safe_para``: reportlab's Paragraph parses
        a subset of HTML, so user-controlled strings MUST be escaped or a
        payload like ``<font color="white">`` would render / a malformed tag
        would crash paraparser.

        It also picks the face, for the same reason it does the escaping: this
        is the one place every string in the report passes through. A report
        run over a Chinese project carries Chinese in its title, its project
        name and every value in its snapshot.
        """
        rendered = "" if text is None else str(text)
        return Paragraph(html.escape(rendered, quote=True), pdf_style_for_text(styles[style_key], rendered))

    flowables: list[Any] = []
    flowables.append(_p(title, "title"))
    flowables.append(_p(f"{tr(locale, 'meta_project')}: {project_name}", "meta"))
    meta_line = f"{tr(locale, 'meta_type')}: {report_type}"
    if currency:
        meta_line += f"  -  {tr(locale, 'meta_currency')}: {currency}"
    flowables.append(_p(meta_line, "meta"))
    for label, value in _evidence_rows(
        report_type=report_type,
        title=title,
        project_name=project_name,
        currency=currency,
        generated_at=generated_at,
        data_snapshot=data_snapshot,
        locale=locale,
    ):
        flowables.append(_p(f"{label}: {value}", "meta"))
    flowables.append(Spacer(1, 4 * mm))

    usable_width = A4[0] - 40 * mm
    col_widths = [usable_width * 0.35, usable_width * 0.65]

    rendered_any = False
    for section in _resolve_sections(report_type, template_data):
        sid = str(section.get("id", "")).strip()
        payload = snapshot.get(sid)
        if _section_is_empty(payload):
            continue
        rendered_any = True
        flowables.append(_p(_section_title(section, locale), "section"))

        if _is_record_list(payload):
            columns = _record_columns(payload)
            header_cells = [_p(field_label(c, locale), "column_header") for c in columns]
            table_rows = [header_cells]
            for item in payload:
                table_rows.append([_p(_stringify(item.get(c), locale), "value") for c in columns])
            n_cols = max(len(columns), 1)
            rec_col_widths = [usable_width / n_cols] * n_cols
            table = Table(table_rows, colWidths=rec_col_widths, repeatRows=1)
            table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 1.5 * mm),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                        ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                        # The fill only: the header cells are Paragraphs and
                        # carry their own colour, so a TEXTCOLOR here would
                        # describe a row nothing draws.
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f8fb")]),
                    ]
                )
            )
            flowables.append(table)
            continue

        table_rows = []
        for label, value in _flatten_keyvalue(payload, locale):
            table_rows.append([_p(label, "label"), _p(value, "value")])

        if not table_rows:
            continue
        table = Table(table_rows, colWidths=col_widths)
        table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 1.5 * mm),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5 * mm),
                    ("LEFTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 2 * mm),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f8f8fb")]),
                ]
            )
        )
        flowables.append(table)

    if not rendered_any:
        flowables.append(_p(tr(locale, "no_data_heading"), "section"))
        flowables.append(_p(tr(locale, "no_data_body_short"), "value"))

    buffer = io.BytesIO()
    # Metadata + the brand header/footer follow the workspace white-label
    # (issue #284); author/creator come from the shared helper while the
    # report's own title/subject stay as set here.
    meta = branded_doc_metadata()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        title=title,
        author=meta["author"],
        subject=tr(locale, "pdf_subject"),
        creator=meta["creator"],
    )
    doc.build(flowables, onFirstPage=branded_header_footer, onLaterPages=branded_header_footer)
    return buffer.getvalue()


# ── Public entry point ──────────────────────────────────────────────────────


def export_report(
    *,
    fmt: str,
    report_type: str,
    title: str,
    project_name: str,
    currency: str,
    generated_at: str,
    template_data: dict[str, Any] | None,
    data_snapshot: dict[str, Any] | None,
    html_body: str | None = None,
    locale: str = DEFAULT_REPORT_LOCALE,
) -> tuple[str, str, bytes]:
    """Render a generated report into a downloadable file.

    Args:
        fmt: One of :data:`SUPPORTED_FORMATS` (``pdf`` / ``xlsx`` / ``csv`` /
            ``html``).
        report_type: Report type token (drives the default section list).
        title: Report title (already schema-sanitised; re-escaped here).
        project_name: Owning project's display name.
        currency: Resolved ISO 4217 code stamped on the report row.
        generated_at: ISO timestamp shown in the file header.
        template_data: Bound template's ``template_data`` (optional).
        data_snapshot: Per-section payload dict.
        html_body: Pre-rendered HTML body, only used for ``fmt="html"`` so
            the existing HTML output is served unchanged. When ``None`` and
            ``fmt="html"`` the HTML is rendered fresh from the snapshot.
            The caller owns the language of a body it passes in: this
            function cannot re-language an already-rendered document, so a
            caller asking for a locale the stored body is not written in
            must pass ``None`` and let the renderer run again.
        locale: Language to write the file in; a member of
            ``report_translations.SUPPORTED_REPORT_LOCALES``. Defaults to
            English, which reproduces the pre-catalogue output exactly.

    Returns:
        ``(suggested_filename, media_type, file_bytes)``.

    Raises:
        ExportFormatError: when *fmt* is not supported.
    """
    fmt = (fmt or "").strip().lower()
    locale = normalize_report_locale(locale)
    if fmt not in SUPPORTED_FORMATS:
        raise ExportFormatError(f"Unsupported export format '{fmt}'. Expected one of: {', '.join(SUPPORTED_FORMATS)}.")

    # Defensive size cap for the in-memory file builders. csv / xlsx / pdf each
    # assemble the whole document in RAM, so an oversized snapshot is turned
    # away with a clear error rather than building an unbounded file and risking
    # an OOM on the single worker. html is served from its own (streamed or
    # pre-rendered) path and is not gated here.
    if fmt in ("csv", "xlsx", "pdf"):
        rows = _snapshot_row_count(data_snapshot)
        if rows > _MAX_EXPORT_ROWS:
            raise ExportFormatError(
                f"This report has {rows:,} rows, above the export limit of "
                f"{_MAX_EXPORT_ROWS:,} rows. Narrow the report scope or export in smaller batches."
            )

    if fmt == "csv":
        blob = _export_csv(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
            locale=locale,
        )
    elif fmt == "xlsx":
        blob = _export_xlsx(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
            locale=locale,
        )
    elif fmt == "pdf":
        blob = _export_pdf(
            report_type=report_type,
            title=title,
            project_name=project_name,
            currency=currency,
            generated_at=generated_at,
            template_data=template_data,
            data_snapshot=data_snapshot,
            locale=locale,
        )
    else:  # html
        if html_body is not None:
            blob = html_body.encode("utf-8")
        else:
            from app.modules.reporting.renderer import ReportRenderer

            blob = (
                ReportRenderer()
                .render_html(
                    report_type=report_type,
                    title=title,
                    project_name=project_name,
                    template_data=template_data,
                    data_snapshot=data_snapshot,
                    generated_at=generated_at,
                    locale=locale,
                )
                .encode("utf-8")
            )

    filename = f"{_safe_filename(title)}.{fmt}"
    return filename, _MEDIA_TYPES[fmt], blob


def _safe_filename(title: str) -> str:
    """Quote-free, single-line base filename derived from the report title.

    Keeps the title's real characters (umlauts, Cyrillic, CJK): the router
    builds the header with
    :func:`app.core.content_disposition.attachment_disposition`, which emits
    the RFC 6266 ASCII fallback plus UTF-8 ``filename*`` pair. What must go
    here is anything that would break the header line itself: double quotes
    are swapped for single quotes and control characters (CR/LF/tab) are
    stripped - a CR/LF in a header value is HTTP response splitting. Falls
    back to ``report`` when the title reduces to nothing.
    """
    base = (title or "").replace('"', "'")
    base = "".join(ch for ch in base if ch >= " " and ch != "\x7f").strip()
    # Collapse path separators that would confuse some download clients.
    base = base.replace("/", "-").replace("\\", "-")
    return base or "report"


# ── COBie export profile ─────────────────────────────────────────────────
#
# COBie (Construction Operations Building Information Exchange, BS 1192-4 /
# ISO 19650 handover shape) is the open, jurisdiction-neutral facility
# handover workbook: one sheet per concept, imported into whatever CAFM / FM
# system the client runs. This module does not re-derive COBie rows from
# scratch - the canonical builder already lives at
# ``app.modules.bim_hub.exporters.cobie.build_cobie_workbook`` and is fully
# covered by its own tests; we reuse it so there is exactly one place that
# knows how a BIMElement becomes a COBie Space/Type/Component/System row.
# What this module adds is the project-level "handover as a report artifact"
# framing: a project can carry several BIM models (architecture, structure,
# MEP, ...), and a COBie handover is the union of their tracked assets, not a
# single model's view.
#
# Sheet-to-source mapping (see also the docstring of
# ``app.modules.bim_hub.exporters.cobie`` for the per-column detail):
#
#     Contact    - synthesised: one default "handover" contact row (COBie
#                  requires at least one Contact; the platform does not yet
#                  model a per-person contact register for handover).
#     Facility   - the project itself (``ReportingService`` resolves the
#                  project's display name; there is one Facility row per
#                  export, standing in for the project as a whole).
#     Floor      - distinct ``BIMElement.storey`` values across every model
#                  in the project.
#     Space      - ``BIMElement`` rows whose ``element_type`` matches a room
#                  / space token (Room, Space, IfcSpace, ...).
#     Type       - ``BIMElement`` rows with ``is_tracked_asset=True``,
#                  grouped by (element_type, manufacturer, model).
#     Component  - ``BIMElement`` rows with ``is_tracked_asset=True`` (the
#                  platform's asset register), one row per element.
#     System     - ``BIMElement.asset_info["parent_system"]`` groupings of
#                  the tracked assets.
#     Zone       - emitted with the standard COBie Zone header row only. The
#                  canonical model does not yet carry a zone/space-grouping
#                  concept distinct from ``storey``, so there is nothing to
#                  populate; a future zoning feature can fill this in without
#                  changing the file shape a CAFM import already expects.
#
# The remaining COBie 2.4 spec sheets (Job, Resource, Spare, Document,
# Attribute, Coordinate, Issue, Connection, Assembly, Impact, PickLists) are
# left out entirely, same rationale as the underlying builder: they need data
# the canonical model does not hold yet (maintenance jobs, spare-parts
# stock, per-element documents/attributes) and a header-only sheet for all of
# them would only add noise without signalling anything real. Add them here
# (not in ``bim_hub``) once a module populates that data, following the same
# "emit the header, fill what's known" pattern used for Zone below.

COBIE_ADDITIONAL_SHEETS: dict[str, list[str]] = {
    "Zone": [
        "Name*",
        "CreatedBy",
        "CreatedOn",
        "Category",
        "SpaceNames",
        "ExtSystem",
        "ExtObject",
        "ExtIdentifier",
        "Description",
    ],
}


def export_project_cobie(
    model: Any,
    elements: list[Any],
    *,
    documents: list[Any] | None = None,
    options: Any | None = None,
) -> tuple[str, str, bytes]:
    """Build a COBie handover workbook and return it in the exporter shape.

    Args:
        model: A BIMModel-like object (or a stand-in with the same
            attributes) representing the facility being handed over.
            Attribute access only - see
            ``bim_hub.exporters.cobie.build_cobie_workbook`` for the exact
            attributes read.
        elements: The project's tracked-asset / space register, as
            BIMElement-like objects. Typically the union of every BIM
            model's elements for a project (multiple disciplines feed one
            handover).
        documents: Reserved, forwarded to the underlying builder unchanged.
        options: Optional ``CobieOptions`` forwarded to the underlying
            builder (project name, currency unit, frozen timestamp for
            deterministic tests, ...).

    Returns:
        ``(suggested_filename, media_type, file_bytes)`` - the same 3-tuple
        shape :func:`export_report` returns, so callers do not need to
        special-case this profile.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    from app.modules.bim_hub.exporters.cobie import build_cobie_workbook

    blob = build_cobie_workbook(model, elements, documents, options)

    # Append the header-only sheets this profile adds on top of the
    # underlying builder (see the module mapping doc above).
    if COBIE_ADDITIONAL_SHEETS:
        wb = load_workbook(BytesIO(blob))
        for sheet_name, columns in COBIE_ADDITIONAL_SHEETS.items():
            if sheet_name in wb.sheetnames:
                continue
            ws = wb.create_sheet(sheet_name)
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
        buffer = BytesIO()
        wb.save(buffer)
        blob = buffer.getvalue()

    safe_name = _safe_filename(getattr(model, "name", None) or "model")
    filename = f"COBie_{safe_name}.xlsx"
    return filename, COBIE_MEDIA_TYPE, blob
