# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Punch List service - business logic for punch list management.

Stateless service layer. Handles:
- Punch item CRUD
- Status transitions with validation (open -> in_progress -> resolved -> verified -> closed)
- Photo management (add/remove photo paths)
- Summary aggregation
- PDF export of punch list items
- Event publishing on create/update/delete/status-transition (slice E)
"""

import logging
import uuid
from collections.abc import Iterable, Mapping
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi import HTTPException, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.events import event_bus
from app.core.json_merge import merge_metadata
from app.core.party_names import resolve_party_names
from app.core.storage import find_existing_upload, module_uploads_dir
from app.modules.punchlist.models import PunchItem
from app.modules.punchlist.repository import PunchListRepository
from app.modules.punchlist.schemas import PunchItemCreate, PunchItemUpdate, PunchStatusTransition

logger = logging.getLogger(__name__)
_logger_ev = logging.getLogger(__name__ + ".events")


def _party_label(item: PunchItem, names: Mapping[str, str] | None) -> str:
    """What an export should print in an assignee column.

    Args:
        item: The punch row.
        names: Resolved names as ``resolve_party_names`` returns them.

    Returns:
        The name when one is known, otherwise the column as stored - an id is
        still better than a blank, because it can at least be looked up.
    """
    raw = item.assigned_to or ""
    return (names or {}).get(raw) or raw


def _as_utc(value: object) -> datetime | None:
    """Return ``value`` as a timezone-aware UTC datetime, or None.

    SQLite hands back naive datetimes where PostgreSQL hands back aware ones,
    so comparing a stored timestamp against ``datetime.now(UTC)`` raises a
    ``TypeError`` on one backend and not the other. A naive value is read as
    already being UTC; an aware one is converted. Anything that is not a
    datetime at all returns None rather than raising, matching the caution the
    close-duration loop below already takes with these same columns.
    """
    if not isinstance(value, datetime):
        return None
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


# Hoist heavy optional imports to module top so we pay the import cost once.
# openpyxl is a soft dependency - the Excel export falls back to CSV when
# it isn't available.
try:  # pragma: no cover - exercised in production paths
    import openpyxl as _openpyxl  # type: ignore[import-not-found]
    from openpyxl.styles import Font as _OpenpyxlFont  # type: ignore[import-not-found]

    _OPENPYXL_AVAILABLE: bool = True
except ImportError:  # pragma: no cover - fallback path
    _openpyxl = None  # type: ignore[assignment]
    _OpenpyxlFont = None  # type: ignore[assignment,misc]
    _OPENPYXL_AVAILABLE = False

# ReportLab is a soft dependency. When it's missing we fall back to the
# minimal hand-rolled PDF writer below so the export still works on slim
# installs. The actual `from reportlab...` statements stay inside the
# builder so module import remains cheap (~no cost beyond the probe).
try:  # pragma: no cover - exercised in production paths
    import reportlab as _reportlab  # noqa: F401  type: ignore[import-not-found]

    _REPORTLAB_AVAILABLE: bool = True
except ImportError:  # pragma: no cover - fallback path
    _REPORTLAB_AVAILABLE = False

# Terminal statuses - any transition FROM one of these back to an active
# status counts as a "reopen" and is appended to ``reopen_history``.
_TERMINAL_STATUSES: frozenset[str] = frozenset({"closed", "verified"})
_ACTIVE_STATUSES: frozenset[str] = frozenset({"open", "in_progress"})

# Where punchlist photos live on disk. Mirrors the path in router.py - we
# resolve photo_path entries against this base when embedding into PDFs.
# Anchored on the platform data dir; ``_resolve_photo_path`` additionally
# probes the working-directory-relative tree earlier releases wrote to, so a
# PDF export still embeds photos captured before the roots were anchored.
_PHOTOS_BASE = module_uploads_dir()


async def _safe_publish(name: str, data: dict, source_module: str = "oe_punchlist") -> None:
    """Best-effort event publish - never blocks the caller on failure."""
    try:
        event_bus.publish_detached(name, data, source_module=source_module)
    except Exception:
        _logger_ev.debug("Event publish skipped: %s", name)


# Valid status transitions: current_status -> list of allowed next statuses
# Valid status transitions: current_status -> list of allowed next statuses
# FSM: open → assigned → in_progress → resolved → verified → closed
# ``reopened`` is a display alias that always resolves to ``open`` in service code.
VALID_TRANSITIONS: dict[str, list[str]] = {
    "open": ["assigned", "in_progress"],  # direct assign or fast-track
    "assigned": ["in_progress", "open"],  # accepted or unassigned
    "in_progress": ["resolved", "verified", "assigned", "open"],  # done via either path
    "resolved": ["verified", "open"],  # legacy pre-verify step
    "verified": ["closed", "open"],  # approved or re-opened
    "closed": ["open"],  # can always reopen
}

# Terminal statuses trigger a reopen_history entry when transitioning back
# to an active status.
# Statuses that require special role checks:
# resolved -> verified: must be a different user than the resolver
# verified -> closed: admin/manager only (handled via permissions in router)


class PunchListService:
    """Business logic for punch list operations."""

    def __init__(self, session: AsyncSession) -> None:
        self.session = session
        self.repo = PunchListRepository(session)

    # ── Create ────────────────────────────────────────────────────────────

    async def create_item(
        self,
        data: PunchItemCreate,
        user_id: str | None = None,
    ) -> PunchItem:
        """Create a new punch list item."""
        item = PunchItem(
            project_id=data.project_id,
            title=data.title,
            description=data.description,
            document_id=data.document_id,
            page=data.page,
            location_x=data.location_x,
            location_y=data.location_y,
            priority=data.priority,
            status="open",
            assigned_to=data.assigned_to,
            due_date=data.due_date,
            category=data.category,
            trade=data.trade,
            geo_lat=data.geo_lat,
            geo_lon=data.geo_lon,
            rework_cost=getattr(data, "rework_cost", None),
            rework_cost_currency=getattr(data, "rework_cost_currency", "USD") or "USD",
            created_by=user_id,
            metadata_=data.metadata,
        )
        item = await self.repo.create(item)

        await _safe_publish(
            "punchlist.item.created",
            {
                "item_id": str(item.id),
                "project_id": str(item.project_id),
                "priority": item.priority,
                "status": item.status,
                "assigned_to": item.assigned_to,
                "created_by": user_id,
            },
        )

        logger.info("Punch item created: %s for project %s", item.title[:40], data.project_id)
        return item

    async def resolve_party_names(self, values: Iterable[str | None]) -> dict[str, str]:
        """Map the ids among ``values`` onto readable names.

        ``assigned_to`` and ``verified_by`` are free-text columns, and a name,
        a contact id and a user id are all legitimate contents. The seeder and
        the field integrations write a contact id; the assignment control on
        the punch screen is a list of platform users and writes a user id. The
        list printed whichever it got, so a row read "Assigned to
        3f2b8c1e-..." where a name belonged.

        The punch list is where that was first worked out, and every other
        register with a person column had the same problem and no map at all,
        so the rule itself now lives in ``app.core.party_names``. This method
        stays because it is what the module's callers ask, and its tests are
        the guard that moving the rule changed none of its answers.

        Args:
            values: Raw column values, ids and names mixed, nulls allowed.

        Returns:
            ``{raw value: display name}`` for the ids that resolved.
        """
        return await resolve_party_names(self.session, values)

    # ── Read ──────────────────────────────────────────────────────────────

    async def get_item(self, item_id: uuid.UUID) -> PunchItem:
        """Get punch item by ID. Raises 404 if not found."""
        item = await self.repo.get_by_id(item_id)
        if item is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Punch item not found",
            )
        return item

    async def list_items(
        self,
        project_id: uuid.UUID,
        *,
        offset: int = 0,
        limit: int = 50,
        status_filter: str | None = None,
        priority_filter: str | None = None,
        assigned_to: str | None = None,
        category_filter: str | None = None,
        trade_filter: str | None = None,
    ) -> tuple[list[PunchItem], int]:
        """List punch items for a project."""
        return await self.repo.list_for_project(
            project_id,
            offset=offset,
            limit=limit,
            status=status_filter,
            priority=priority_filter,
            assigned_to=assigned_to,
            category=category_filter,
            trade=trade_filter,
        )

    # ── Update ────────────────────────────────────────────────────────────

    async def update_item(
        self,
        item_id: uuid.UUID,
        data: PunchItemUpdate,
    ) -> PunchItem:
        """Update punch item fields."""
        item = await self.get_item(item_id)

        fields = data.model_dump(exclude_unset=True)
        if "metadata" in fields:
            _incoming = fields.pop("metadata")
            fields["metadata_"] = (
                merge_metadata(getattr(item, "metadata_", None), _incoming)
                if isinstance(_incoming, dict)
                else _incoming
            )

        if not fields:
            return item

        await self.repo.update_fields(item_id, **fields)
        await self.session.refresh(item)

        await _safe_publish(
            "punchlist.item.updated",
            {
                "item_id": str(item_id),
                "project_id": str(item.project_id),
                "updated_fields": list(fields.keys()),
            },
        )

        logger.info("Punch item updated: %s (fields=%s)", item_id, list(fields.keys()))
        return item

    # ── Delete ────────────────────────────────────────────────────────────

    async def delete_item(self, item_id: uuid.UUID) -> None:
        """Delete a punch item."""
        item = await self.get_item(item_id)  # Raises 404 if not found
        project_id = str(item.project_id)
        await self.repo.delete(item_id)

        await _safe_publish(
            "punchlist.item.deleted",
            {
                "item_id": str(item_id),
                "project_id": project_id,
            },
        )

        logger.info("Punch item deleted: %s", item_id)

    # ── Status transition ─────────────────────────────────────────────────

    async def transition_status(
        self,
        item_id: uuid.UUID,
        transition: PunchStatusTransition,
        user_id: str,
    ) -> PunchItem:
        """Transition a punch item to a new status with validation.

        Rules:
        - open -> in_progress (anyone)
        - in_progress -> resolved (assigned user or admin)
        - resolved -> verified (different user than resolver - enforced here)
        - verified -> closed (admin/manager - enforced via permission in router)
        - Any -> open (reopen)
        """
        item = await self.get_item(item_id)
        current = item.status
        # ``reopened`` is a client-facing alias for ``open`` that always
        # triggers the reopen audit path regardless of current state.
        target = "open" if transition.new_status == "reopened" else transition.new_status

        # Reopen is always allowed from any status
        if target == "open":
            update_fields: dict[str, Any] = {"status": "open"}
            if transition.notes:
                update_fields["resolution_notes"] = transition.notes
            self._record_reopen_if_needed(
                item,
                new_status="open",
                user=user_id,
                reason=transition.notes,
                update_fields=update_fields,
            )
            await self.repo.update_fields(item_id, **update_fields)
            await self.session.refresh(item)

            # Epic H - universal audit trail (reopen branch).
            from app.core.audit_log import log_activity as _log_activity

            await _log_activity(
                self.session,
                actor_id=user_id,
                entity_type="punch_item",
                entity_id=str(item_id),
                action="status_changed",
                from_status=current,
                to_status="open",
                reason=transition.notes,
                metadata={"reopened": True},
                module="punchlist",
                parent_entity_type="project",
                parent_entity_id=str(item.project_id),
                before_state={"status": current},
                after_state={"status": "open"},
            )

            await _safe_publish(
                "punchlist.item.status_changed",
                {
                    "item_id": str(item_id),
                    "project_id": str(item.project_id),
                    "from_status": current,
                    "to_status": "open",
                    "user_id": user_id,
                },
            )

            logger.info("Punch item reopened: %s by %s", item_id, user_id)
            return item

        # Validate allowed transitions
        allowed = VALID_TRANSITIONS.get(current, [])
        if target not in allowed:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot transition from '{current}' to '{target}'",
            )

        now = datetime.now(UTC)
        update_fields: dict[str, Any] = {"status": target}

        if transition.notes:
            update_fields["resolution_notes"] = transition.notes

        # open/assigned → assigned: record assigned_at timestamp in metadata
        if target == "assigned":
            existing_meta = dict(getattr(item, "metadata_", None) or {})
            existing_meta["assigned_at"] = now.isoformat()
            update_fields["metadata_"] = existing_meta

        # in_progress → resolved: set resolved_at and record who resolved it so the
        # four-eyes verify gate can reject the same user later, even when the item
        # has no assignee.
        if target == "resolved":
            update_fields["resolved_at"] = now
            resolved_meta = dict(getattr(item, "metadata_", None) or {})
            resolved_meta["resolved_by"] = user_id
            update_fields["metadata_"] = resolved_meta

        # resolved/in_progress → verified: must be a different user than the one who
        # resolved the item. A null assignee must not disable the guard, so we compare
        # against the recorded resolver (metadata_.resolved_by) instead of assigned_to.
        if target == "verified":
            resolved_by = (getattr(item, "metadata_", None) or {}).get("resolved_by")
            if resolved_by and resolved_by == user_id:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Verification must be done by a different user than the resolver",
                )
            update_fields["verified_at"] = now
            update_fields["verified_by"] = user_id

        # verified -> closed: block if critical items remain open in the same project
        if target == "closed" and item.priority == "critical":
            open_critical = await self.repo.count_open_critical(item.project_id, exclude_id=item_id)
            if open_critical > 0:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=(
                        f"Cannot close this critical punch item while {open_critical} other "
                        f"critical item(s) remain unresolved in the project"
                    ),
                )

        # Record reopen audit for the rare path where an allowed transition
        # moves a terminal item back to an active status without going via
        # the explicit "open" reopen branch above (defence in depth).
        self._record_reopen_if_needed(
            item,
            new_status=target,
            user=user_id,
            reason=transition.notes,
            update_fields=update_fields,
        )

        await self.repo.update_fields(item_id, **update_fields)
        await self.session.refresh(item)

        # Epic H - universal audit trail.
        from app.core.audit_log import log_activity as _log_activity

        await _log_activity(
            self.session,
            actor_id=user_id,
            entity_type="punch_item",
            entity_id=str(item_id),
            action="status_changed",
            from_status=current,
            to_status=target,
            reason=transition.notes,
            metadata={"priority": item.priority},
            module="punchlist",
            parent_entity_type="project",
            parent_entity_id=str(item.project_id),
            before_state={"status": current},
            after_state={"status": target},
        )

        await _safe_publish(
            "punchlist.item.status_changed",
            {
                "item_id": str(item_id),
                "project_id": str(item.project_id),
                "from_status": current,
                "to_status": target,
                "user_id": user_id,
            },
        )

        logger.info(
            "Punch item transitioned: %s %s -> %s by %s",
            item_id,
            current,
            target,
            user_id,
        )
        return item

    # ── Reopen audit ──────────────────────────────────────────────────────

    def _record_reopen_if_needed(
        self,
        item: PunchItem,
        *,
        new_status: str,
        user: str | None,
        reason: str | None,
        update_fields: dict[str, Any],
    ) -> None:
        """Append a reopen-history entry when transitioning from terminal -> active.

        Mutates ``update_fields`` in place so the new ``reopen_history`` list
        is persisted alongside the status change in a single update.
        """
        previous = item.status
        if previous not in _TERMINAL_STATUSES or new_status not in _ACTIVE_STATUSES:
            return

        existing = list(getattr(item, "reopen_history", None) or [])
        entry: dict[str, Any] = {
            "reopened_at": datetime.now(UTC).isoformat(),
            "reopened_by": user,
            "previous_status": previous,
        }
        if reason:
            entry["reason"] = reason
        existing.append(entry)
        update_fields["reopen_history"] = existing

    # ── Bulk close ────────────────────────────────────────────────────────

    async def bulk_close(
        self,
        project_id: uuid.UUID,
        item_ids: list[uuid.UUID],
        *,
        user_id: str,
        comment: str | None = None,
    ) -> dict[str, Any]:
        """Close many punch items at once.

        - Items already ``closed`` are counted as ``skipped``.
        - Items not found, owned by another project, or violating close rules
          (e.g. critical items with open peers) are returned in ``errors``.
        - Successful closes emit ``punchlist.item.status_changed`` events.
        """
        closed = 0
        skipped = 0
        errors: list[dict[str, Any]] = []

        for item_id in item_ids:
            try:
                item = await self.repo.get_by_id(item_id)
                if item is None:
                    errors.append({"id": str(item_id), "error": "not_found"})
                    continue
                if item.project_id != project_id:
                    errors.append({"id": str(item_id), "error": "project_mismatch"})
                    continue
                if item.status == "closed":
                    skipped += 1
                    continue

                # Critical-with-open-peers guard mirrors transition_status().
                if item.priority == "critical":
                    open_critical = await self.repo.count_open_critical(project_id, exclude_id=item_id)
                    if open_critical > 0:
                        errors.append(
                            {
                                "id": str(item_id),
                                "error": (f"critical_blocked:{open_critical}_other_open"),
                            }
                        )
                        continue

                # Snapshot the prior status before the close write, so the
                # transition below is logged from where the item actually came
                # without a reload, which would raise MissingGreenlet.
                from_status = item.status

                update_fields: dict[str, Any] = {"status": "closed"}
                if comment:
                    update_fields["resolution_notes"] = comment

                await self.repo.update_fields(item_id, **update_fields)
                closed += 1

                await _safe_publish(
                    "punchlist.item.status_changed",
                    {
                        "item_id": str(item_id),
                        "project_id": str(project_id),
                        "from_status": from_status,
                        "to_status": "closed",
                        "user_id": user_id,
                        "bulk": True,
                    },
                )
            except HTTPException as exc:
                errors.append({"id": str(item_id), "error": str(exc.detail)})
            except Exception as exc:  # noqa: BLE001
                logger.exception("Bulk-close failed for punch item %s", item_id)
                errors.append({"id": str(item_id), "error": exc.__class__.__name__})

        logger.info(
            "Bulk-closed punch items for project %s: %d closed, %d skipped, %d errors",
            project_id,
            closed,
            skipped,
            len(errors),
        )
        return {"closed": closed, "skipped": skipped, "errors": errors}

    # ── Pin to sheet ──────────────────────────────────────────────────────

    async def pin_to_sheet(
        self,
        item_id: uuid.UUID,
        *,
        sheet_id: str | None = None,
        document_id: str | None = None,
        page: int,
        location_x: float,
        location_y: float,
    ) -> PunchItem:
        """Pin a punch item to a location on a document sheet.

        Updates document_id, page, location_x, and location_y on the item.
        If sheet_id is given but document_id is not, sheet_id is stored as
        the document_id (sheets are a logical subset of documents).
        """
        item = await self.get_item(item_id)

        effective_doc_id = document_id or sheet_id

        update_fields: dict[str, Any] = {
            "document_id": effective_doc_id,
            "page": page,
            "location_x": location_x,
            "location_y": location_y,
        }

        await self.repo.update_fields(item_id, **update_fields)
        await self.session.refresh(item)

        logger.info(
            "Punch item pinned to sheet: %s -> doc=%s page=%d (%.2f, %.2f)",
            item_id,
            effective_doc_id,
            page,
            location_x,
            location_y,
        )
        return item

    # ── Photos ────────────────────────────────────────────────────────────

    async def add_photo(self, item_id: uuid.UUID, photo_path: str) -> PunchItem:
        """Add a photo path to the punch item's photos list."""
        item = await self.get_item(item_id)
        photos = list(item.photos or [])
        photos.append(photo_path)
        await self.repo.update_fields(item_id, photos=photos)
        await self.session.refresh(item)
        logger.info("Photo added to punch item %s: %s", item_id, photo_path)
        return item

    async def remove_photo(self, item_id: uuid.UUID, index: int) -> PunchItem:
        """Remove a photo by index from the punch item's photos list."""
        item = await self.get_item(item_id)
        photos = list(item.photos or [])

        if not photos:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No photos to remove",
            )

        if index < 0 or index >= len(photos):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Photo index {index} out of range (0..{len(photos) - 1})",
            )

        removed = photos.pop(index)
        await self.repo.update_fields(item_id, photos=photos)
        await self.session.refresh(item)
        logger.info("Photo removed from punch item %s: %s", item_id, removed)
        return item

    # ── Summary ───────────────────────────────────────────────────────────

    async def get_summary(self, project_id: uuid.UUID) -> dict[str, Any]:
        """Get aggregated stats for a project's punch list."""
        agg = await self.repo.summary_aggregates(project_id)
        overdue = await self.repo.count_overdue(project_id)

        # closed_timestamps is a list of (created_at, verified_at, resolved_at,
        # updated_at) tuples for closed/verified items only - SQL diff isn't
        # portable across SQLite/PostgreSQL so we still walk in Python.
        now = datetime.now(UTC)
        week_ago = now - timedelta(days=7)

        closed_durations: list[float] = []
        closed_last_7_days = 0
        for created_at, verified_at, resolved_at, updated_at in agg["closed_timestamps"]:
            end_time = verified_at or resolved_at or updated_at
            end_utc = _as_utc(end_time)
            if end_utc is not None and end_utc >= week_ago:
                closed_last_7_days += 1
            if not created_at:
                continue
            if end_time is None:
                continue
            try:
                days = (end_time - created_at).total_seconds() / 86400.0
            except (TypeError, AttributeError):
                continue
            if days >= 0:
                closed_durations.append(days)

        avg_days_to_close: float | None = None
        if closed_durations:
            avg_days_to_close = round(sum(closed_durations) / len(closed_durations), 1)

        # Age of the still-open work. Averaged here rather than in SQL for the
        # same reason the close durations are: SQLite has no portable date diff.
        # Clamped at zero so a row stamped in the future cannot pull the mean
        # negative and report the backlog as younger than it is.
        open_ages = [
            max(0.0, (now - created_utc).total_seconds() / 86400.0)
            for created_utc in (_as_utc(v) for v in agg["open_created_at"])
            if created_utc is not None
        ]
        avg_open_age_days: float | None = None
        if open_ages:
            avg_open_age_days = round(sum(open_ages) / len(open_ages), 1)

        return {
            "total": agg["total"],
            "by_status": agg["by_status"],
            "by_priority": agg["by_priority"],
            "overdue": overdue,
            "avg_days_to_close": avg_days_to_close,
            "urgent_open": agg["urgent_open"],
            "closed_last_7_days": closed_last_7_days,
            "avg_open_age_days": avg_open_age_days,
        }

    # ── PDF Export ────────────────────────────────────────────────────────

    async def export_pdf(self, project_id: uuid.UUID) -> bytes:
        """Generate a rich PDF report with all punch list items.

        Uses ReportLab when available (cover page, per-item cards, embedded
        photo thumbnails, sheet-pin captions). Falls back to the minimal
        hand-rolled PDF writer when ReportLab is not installed so the
        endpoint always returns a valid ``application/pdf``.
        """
        items = await self.repo.all_for_project(project_id)
        # The exported list is the artefact that leaves the building, so it
        # has to name the same party the screen names. An id in a printed
        # column cannot even be clicked.
        names = await self.resolve_party_names(item.assigned_to for item in items)

        if _REPORTLAB_AVAILABLE:
            pdf = _build_reportlab_pdf(project_id, items, names)
        else:
            pdf = _build_minimal_pdf(_render_punchlist_text(project_id, items, names))

        logger.info(
            "Punch list PDF exported for project %s (%d items, reportlab=%s)",
            project_id,
            len(items),
            _REPORTLAB_AVAILABLE,
        )
        return pdf

    async def export_excel(self, project_id: uuid.UUID) -> bytes:
        """Generate an Excel report with all punch list items.

        Returns raw xlsx bytes when ``openpyxl`` is available, otherwise
        falls back to UTF-8 CSV bytes. The ``openpyxl`` import is resolved
        once at module load - :data:`_OPENPYXL_AVAILABLE` tells us which
        branch to take without repeatedly catching ``ImportError``.
        """
        items = await self.repo.all_for_project(project_id)
        names = await self.resolve_party_names(item.assigned_to for item in items)

        if _OPENPYXL_AVAILABLE:
            import io

            assert _openpyxl is not None  # for type-checkers
            assert _OpenpyxlFont is not None
            wb = _openpyxl.Workbook()
            ws = wb.active
            ws.title = "Punch List"

            headers = [
                "No.",
                "Title",
                "Status",
                "Priority",
                "Category",
                "Trade",
                "Assigned To",
                "Due Date",
                "Description",
                "Resolution Notes",
                "Created",
            ]

            bold = _OpenpyxlFont(bold=True)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = bold

            for row_idx, item in enumerate(items, 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=item.title)
                ws.cell(row=row_idx, column=3, value=item.status)
                ws.cell(row=row_idx, column=4, value=item.priority)
                ws.cell(row=row_idx, column=5, value=item.category or "")
                ws.cell(row=row_idx, column=6, value=item.trade or "")
                ws.cell(row=row_idx, column=7, value=_party_label(item, names))
                ws.cell(row=row_idx, column=8, value=str(item.due_date) if item.due_date else "")
                ws.cell(row=row_idx, column=9, value=(item.description or "")[:500])
                ws.cell(row=row_idx, column=10, value=(item.resolution_notes or "")[:500])
                ws.cell(row=row_idx, column=11, value=str(item.created_at) if item.created_at else "")

            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()

            logger.info("Punch list Excel exported for project %s (%d items)", project_id, len(items))
            return excel_bytes

        # Fallback: return CSV bytes if openpyxl is not installed
        import csv
        import io as _io

        output = _io.StringIO()
        writer = csv.writer(output)
        writer.writerow(
            [
                "No.",
                "Title",
                "Status",
                "Priority",
                "Category",
                "Trade",
                "Assigned To",
                "Due Date",
                "Description",
                "Resolution Notes",
                "Created",
            ]
        )
        for idx, item in enumerate(items, 1):
            writer.writerow(
                [
                    idx,
                    item.title,
                    item.status,
                    item.priority,
                    item.category or "",
                    item.trade or "",
                    _party_label(item, names),
                    str(item.due_date) if item.due_date else "",
                    (item.description or "")[:500],
                    (item.resolution_notes or "")[:500],
                    str(item.created_at) if item.created_at else "",
                ]
            )
        logger.info(
            "Punch list CSV exported (openpyxl not available) for project %s (%d items)",
            project_id,
            len(items),
        )
        return output.getvalue().encode("utf-8")


def _build_minimal_pdf(text: str) -> bytes:
    """Build a minimal valid PDF document from plain text.

    This produces a basic but valid PDF without requiring any external library.
    """
    # Escape special PDF characters in text
    safe_text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")

    # Split into lines for the PDF text block
    text_lines = safe_text.split("\n")
    # Build BT/ET text block with Td positioning
    text_commands: list[str] = []
    text_commands.append("BT")
    text_commands.append("/F1 10 Tf")
    text_commands.append("50 750 Td")
    text_commands.append("12 TL")  # leading
    for line in text_lines:
        text_commands.append(f"({line}) '")
    text_commands.append("ET")
    stream_content = "\n".join(text_commands)

    objects: list[str] = []

    # Object 1: Catalog
    objects.append("1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj")
    # Object 2: Pages
    objects.append("2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj")
    # Object 3: Page
    objects.append(
        "3 0 obj\n<< /Type /Page /Parent 2 0 R "
        "/MediaBox [0 0 612 792] "
        "/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\nendobj"
    )
    # Object 4: Content stream
    objects.append(f"4 0 obj\n<< /Length {len(stream_content)} >>\nstream\n{stream_content}\nendstream\nendobj")
    # Object 5: Font
    objects.append("5 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>\nendobj")

    # Build the PDF
    parts: list[str] = ["%PDF-1.4"]
    offsets: list[int] = []
    current = len(parts[0]) + 1  # +1 for newline

    for obj in objects:
        offsets.append(current)
        parts.append(obj)
        current += len(obj) + 1

    # Cross-reference table
    xref_offset = current
    xref_lines = [f"xref\n0 {len(objects) + 1}", "0000000000 65535 f "]
    for off in offsets:
        xref_lines.append(f"{off:010d} 00000 n ")
    parts.append("\n".join(xref_lines))

    # Trailer
    parts.append(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF")

    # Courier is a Type1 base font covering Latin-1 only, and this document is a
    # single-byte stream, so anything outside Latin-1 in a punch item (a Cyrillic
    # title, a CJK assignee name, a Polish stroke) has to be substituted rather
    # than crash the export with a UnicodeEncodeError -> HTTP 500. errors="replace"
    # emits one '?' byte per unencodable character, which keeps every character one
    # byte so the /Length written above stays correct.
    #
    # This writer is a near-duplicate of the field report one in
    # app/modules/fieldreports/service.py, which has carried this argument and this
    # reason from the start. The duplication is known and is deliberately not
    # resolved here: the shared writer belongs in core rather than in either module,
    # because a module importing another module's service would make punch list
    # export stop working depending on which other modules happen to be installed.
    return "\n".join(parts).encode("latin-1", "replace")


def _render_punchlist_text(
    project_id: uuid.UUID,
    items: list[PunchItem],
    names: Mapping[str, str] | None = None,
) -> str:
    """Render a flat text view of the punch list - used by the minimal-PDF fallback."""
    lines: list[str] = []
    lines.append("PUNCH LIST REPORT")
    lines.append(f"Project: {project_id}")
    lines.append(f"Generated: {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append(f"Total Items: {len(items)}")
    lines.append("")
    lines.append("-" * 80)

    for idx, item in enumerate(items, 1):
        lines.append(f"\n{idx}. {item.title}")
        lines.append(f"   Status: {item.status} | Priority: {item.priority}")
        if item.category:
            lines.append(f"   Category: {item.category}")
        if item.trade:
            lines.append(f"   Trade: {item.trade}")
        if item.assigned_to:
            lines.append(f"   Assigned to: {_party_label(item, names)}")
        if item.due_date:
            lines.append(f"   Due: {item.due_date}")
        if item.description:
            lines.append(f"   Description: {item.description[:200]}")
        if item.resolution_notes:
            lines.append(f"   Resolution: {item.resolution_notes[:200]}")
        lines.append(f"   Created: {item.created_at}")

    return "\n".join(lines)


def _resolve_photo_path(rel_or_abs: str) -> Path | None:
    """Resolve a stored photo path to a readable file or return None.

    Photos are persisted as relative paths like ``punchlist/photos/<uuid>.jpg``
    underneath the ``uploads/`` directory (see ``router.upload_photo``). A
    relative entry is resolved against the active data-dir root first and then
    the working-directory-relative tree earlier releases wrote to, so a project
    whose photos predate the anchoring still renders them.
    Defensive: any error => return None so the PDF builder simply skips the
    thumbnail without breaking export.
    """
    if not rel_or_abs:
        return None
    try:
        p = Path(rel_or_abs)
        if p.is_absolute():
            return p if p.is_file() else None
        return find_existing_upload(p, _PHOTOS_BASE)
    except Exception:  # noqa: BLE001
        return None


def _build_reportlab_pdf(
    project_id: uuid.UUID,
    items: list[PunchItem],
    names: Mapping[str, str] | None = None,
) -> bytes:
    """Build a styled PDF using ReportLab.

    Layout:
        * Cover page - title, project id, generated date, open / closed totals.
        * One block per item - code, title, location, assignee, status,
          severity, due date.
        * If the item has a photo on disk, the first photo is embedded as
          an 80×80 px thumbnail.
        * If the item has sheet-pin coordinates (``document_id`` / ``page``
          / ``location_x``+ ``location_y``) a small caption is rendered.
    """
    # Lazy-import - only paid when ReportLab is actually used.
    import io

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Image as RLImage,
    )
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from app.core.pdf_fonts import (
        BODY_FONT,
        BOLD_FONT,
        pdf_style_for_text,
        pdf_table_paragraph_rows,
        register_pdf_fonts,
    )

    register_pdf_fonts()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Punch List {project_id}",
    )

    styles = getSampleStyleSheet()
    h1 = styles["Heading1"]
    h1.fontName = BOLD_FONT
    h2 = styles["Heading2"]
    h2.fontName = BOLD_FONT
    body = styles["BodyText"]
    body.fontName = BODY_FONT
    small = ParagraphStyle(
        "punch_small",
        parent=body,
        fontSize=8,
        leading=10,
        textColor=colors.grey,
    )
    caption = ParagraphStyle(
        "punch_caption",
        parent=body,
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#1f3a8a"),
    )
    # The item meta table. Its face, size and label colour were table commands
    # until these cells became Paragraphs, and a table command cannot reach a
    # flowable, so they live here instead.
    meta_value = ParagraphStyle(
        "punch_meta_value",
        parent=body,
        fontName=BODY_FONT,
        fontSize=9,
        leading=11,
        spaceBefore=0,
        spaceAfter=0,
    )
    meta_label = ParagraphStyle(
        "punch_meta_label",
        parent=meta_value,
        fontName=BOLD_FONT,
        textColor=colors.HexColor("#444444"),
    )

    def _meta_face(_row_index: int, col_index: int):
        """Columns 0 and 2 are labels, columns 1 and 3 are the item's data."""
        return meta_label if col_index in (0, 2) else None

    open_count = sum(1 for it in items if it.status not in ("closed", "verified"))
    closed_count = len(items) - open_count

    story: list = []
    story.append(Paragraph("Punch List Report", h1))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"<b>Project:</b> {project_id}", body))
    story.append(
        Paragraph(
            f"<b>Generated:</b> {datetime.now(UTC).strftime('%Y-%m-%d %H:%M UTC')}",
            body,
        )
    )
    story.append(Paragraph(f"<b>Total Items:</b> {len(items)}", body))
    story.append(Paragraph(f"<b>Open:</b> {open_count}", body))
    story.append(Paragraph(f"<b>Closed:</b> {closed_count}", body))
    story.append(PageBreak())

    for idx, item in enumerate(items, 1):
        code = (item.metadata_ or {}).get("code") if hasattr(item, "metadata_") else None
        heading = f"#{idx} - {item.title}"
        if code:
            heading = f"#{idx} · {code} - {item.title}"
        story.append(Paragraph(heading, pdf_style_for_text(h2, heading)))

        meta_rows = [
            ["Status", item.status or "-", "Priority", item.priority or "-"],
            [
                "Assignee",
                _party_label(item, names) or "-",
                "Due Date",
                item.due_date.strftime("%Y-%m-%d") if item.due_date else "-",
            ],
            [
                "Category",
                item.category or "-",
                "Trade",
                item.trade or "-",
            ],
        ]
        # Paragraph cells rather than bare strings. A bare cell is drawn
        # through canvas.drawString, which neither wraps nor shapes: an
        # assignee or a trade longer than its 55mm column was printed over the
        # label beside it, and the last column ran off the sheet, while a Thai
        # or Devanagari value was mis-arranged whatever face it was given. A
        # Paragraph does both and carries its own face, size and colour, so the
        # FONT and TEXTCOLOR commands that used to sit here are gone rather
        # than left to describe a layout nothing obeys. ROWBACKGROUNDS stays,
        # because a fill is drawn by the table and not by the cell.
        meta_rows = pdf_table_paragraph_rows(meta_rows, meta_value, style_for=_meta_face)
        meta_table = Table(meta_rows, colWidths=[26 * mm, 55 * mm, 26 * mm, 55 * mm])
        meta_table.setStyle(
            TableStyle(
                [
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 1),
                    (
                        "ROWBACKGROUNDS",
                        (0, 0),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f6f7f9")],
                    ),
                ]
            )
        )
        story.append(meta_table)
        story.append(Spacer(1, 2 * mm))

        if item.description:
            description = item.description[:1000]
            story.append(Paragraph(description, pdf_style_for_text(body, description)))

        # Sheet-pin caption (sheet-pin style).
        sheet_ref = getattr(item, "document_id", None) or (item.metadata_ or {}).get("sheet_id")
        pin_x = item.location_x
        pin_y = item.location_y
        if sheet_ref and pin_x is not None and pin_y is not None:
            story.append(
                Paragraph(
                    f"&#128205; ({pin_x:.3f}, {pin_y:.3f}) on sheet {sheet_ref}"
                    + (f" · page {item.page}" if item.page else ""),
                    caption,
                )
            )

        # First photo as 80×80 thumbnail when available on disk.
        photos = list(item.photos or [])
        # Also accept legacy single ``photo_path`` attribute for forward-compat.
        legacy_photo = getattr(item, "photo_path", None)
        if legacy_photo:
            photos.insert(0, legacy_photo)
        for raw in photos[:1]:
            disk_path = _resolve_photo_path(raw)
            if disk_path is None:
                continue
            try:
                img = RLImage(str(disk_path), width=80, height=80)
                story.append(Spacer(1, 1 * mm))
                story.append(img)
            except Exception:  # noqa: BLE001 - defensive
                # If reportlab can't decode the image we silently skip it.
                story.append(Paragraph(f"[photo {disk_path.name} could not be embedded]", small))

        if item.resolution_notes:
            story.append(Spacer(1, 1 * mm))
            notes = item.resolution_notes[:500]
            story.append(Paragraph(f"<b>Resolution:</b> {notes}", pdf_style_for_text(small, notes)))

        # Reopen-history chronology (defensive: schema may not yet be migrated).
        history = list(getattr(item, "reopen_history", None) or [])
        if history:
            story.append(Spacer(1, 1 * mm))
            for entry in history[-3:]:  # last 3 reopens at most
                ts = entry.get("reopened_at", "?")
                prev = entry.get("previous_status", "?")
                by = entry.get("reopened_by", "?")
                story.append(
                    Paragraph(
                        f"&#8634; reopened from <b>{prev}</b> by {by} at {ts}",
                        pdf_style_for_text(small, f"{prev}{by}"),
                    )
                )

        story.append(Spacer(1, 6 * mm))

    if not items:
        story.append(Paragraph("No punch list items recorded for this project.", body))

    doc.build(story)
    return buffer.getvalue()
