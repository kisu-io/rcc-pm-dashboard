# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Hungarian bill-of-quantities workbook profiles.

Two workbook shapes are in production use in Hungary and neither survives the
generic spreadsheet reader, for the same reason: neither is a table with one
header row.

**Building works.** One worksheet per chapter, seventeen of them, plus a cover
summary and a chapter summary. Each chapter sheet carries a five-row header
block, and the item code is not in a column: it is spread across nine columns
with separator columns between them, and the line's name sits in whichever
column follows the last filled segment. Every priced line carries two unit
prices, material (anyag) and fee (díj), and two totals to match.

**Infrastructure works.** One flat sheet with a real header row, followed by
several code-dictionary sheets. The line's place in the work breakdown is
spelled out twice, once as a set of level columns and once as a structure code
joining them with ``+``; its identity for the client's monitoring system is a
row number joined to a catalogue item number; and a set of tag columns carry
codes resolved against the dictionary sheets, with the programme activity and
its dates among them.

Why this is a profile inside the Excel importer rather than an importer of its
own: the importer protocol picks a parser from the first four kilobytes of the
upload, and an ``.xlsx`` is a zip whose first four kilobytes name
``xl/worksheets/sheet1.xml`` and nothing about what is in it. Registering a
Hungarian importer ahead of the generic one would mean claiming every
spreadsheet and hoping. Recognition here happens after the workbook is open,
where the question can actually be answered, and a workbook that is neither
shape is handed straight back.

Nothing in this module names a cost-catalogue vendor. The building sheets carry
a column of item numbers from a purchased catalogue; the column is recognised by
what it holds and the value is carried through as
``catalogue_item_number``.
"""

from __future__ import annotations

import io
import logging
import re
from datetime import date, datetime
from typing import Any

from app.modules.boq.importers._base import ImportedBOQ, ImportedPosition

logger = logging.getLogger(__name__)

# ── Recognition ──────────────────────────────────────────────────────────

# The building workbook names itself on the header row of every chapter sheet.
# Matched case- and whitespace-insensitively because the cell is styled across
# a merge and picks up stray spacing.
_BUILDING_TITLE = "egyseges magasepitesi agazati tetelrend"

# Chapter sheets are named for the sector and the chapter: MA-01_ALT and so on.
_BUILDING_SECTOR = "MA"
_BUILDING_SHEET_RE = re.compile(r"^MA-(\d{2})[_ ]", re.IGNORECASE)

# The infrastructure sheet is recognised by two headers that do not occur
# together anywhere else: the structure code and the per-project item number.
_INFRA_REQUIRED_HEADERS = frozenset({"struktura kod", "pde tetelszam"})

# Where the nine code segments sit on a chapter sheet, zero-based: B, D, F, H,
# J, L, N, P, R, with the separator columns between them. Read out of the
# workbook's own numbering row rather than assumed, but kept here as the
# fallback for a sheet whose numbering row is missing.
_BUILDING_SEGMENT_COLUMNS = (1, 3, 5, 7, 9, 11, 13, 15, 17)

_ACCENTS = str.maketrans(
    "áéíóöőúüűÁÉÍÓÖŐÚÜŰ",
    "aeiooouuuAEIOOOUUU",
)


def _fold(value: object) -> str:
    """Lower-case, strip accents, collapse whitespace. For matching only."""
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text.translate(_ACCENTS).strip().lower())


def _text(value: object) -> str:
    """A cell as trimmed text, with the empty cell and ``None`` both blank."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: object) -> float | None:
    """A cell as a float, or ``None`` when it does not hold one.

    Hungarian workbooks are written with a comma decimal separator and a space
    or non-breaking space between thousands, and openpyxl hands those back as
    text whenever the cell was typed rather than computed.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(" ", "").replace(" ", "").replace(" ", "")
    # A comma is the decimal separator; a dot in the same string is a thousands
    # separator and goes. With no comma, a dot is the decimal separator.
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


# ── Building profile ─────────────────────────────────────────────────────

# Chapter sheet columns, matched on what the header cell says. The catalogue
# reference is matched on the bare word for "item number" so the column is
# found by its content and not by the vendor whose catalogue fills it.
_BUILDING_COLUMNS: dict[str, tuple[str, ...]] = {
    "catalogue_item_number": ("tetelszam",),
    "description": ("tetel szovege",),
    "quantity": ("mennyiseg",),
    "unit": ("egyseg",),
    "material_unit_rate": ("anyag egysegar",),
    "fee_unit_rate": ("dij egysegar",),
    "material_total": ("netto anyag osszesen (ft)", "netto anyag osszesen"),
    "fee_total": ("netto dij osszesen (ft)", "netto dij osszesen"),
    "combined_total": ("netto a+d osszesen",),
    "note": ("megjegyzes",),
}


def _at(row: tuple[Any, ...], columns: dict[str, int], field: str) -> Any:
    """The cell a mapped field points at on this row, or ``None``."""
    column = columns.get(field)
    if column is None or column >= len(row):
        return None
    return row[column]


def _building_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """Find the header row of a chapter sheet and map its columns.

    Returns ``(row_index, {field: column_index})``, or ``None`` when the sheet
    carries no header the profile recognises. The row index is zero-based into
    ``rows``.
    """
    for index, row in enumerate(rows[:12]):
        folded = [_fold(cell) for cell in row]
        if not any(cell == _BUILDING_TITLE for cell in folded):
            continue
        mapping: dict[str, int] = {}
        claimed: set[int] = {column for column, cell in enumerate(folded) if cell == _BUILDING_TITLE}

        # Exact matches first, and a column can only be claimed once. Doing it
        # the other way round loses two columns to the same word: the sheet
        # names itself "EGYSEGES ..." on the header row, which contains the
        # word for "unit", and "MENNYISEG" contains it too. Reading the sector
        # letters out of the title cell as a unit made every heading look
        # priced, and a heading that looks priced is imported as an item
        # carrying the total of everything beneath it.
        for exact in (True, False):
            for field, aliases in _BUILDING_COLUMNS.items():
                if field in mapping:
                    continue
                for column, cell in enumerate(folded):
                    if not cell or column in claimed:
                        continue
                    hit = cell in aliases if exact else any(alias in cell for alias in aliases)
                    if hit:
                        mapping[field] = column
                        claimed.add(column)
                        break
        return index, mapping
    return None


def _building_segments(row: tuple[Any, ...]) -> tuple[dict[int, str], int]:
    """The code segments this row fills, and the deepest one it reaches.

    A chapter sheet does not repeat the whole code on every line. A heading
    writes the segments down to its own level and the lines below it write
    only the segment that changed, so the full code of a row is the running
    composition of everything above it. This returns what the row itself says;
    :func:`_parse_building` keeps the running prefix.

    Returns ``({segment_index: value}, deepest_index)`` with ``deepest_index``
    at ``-1`` when the row fills no segment at all.
    """
    filled: dict[int, str] = {}
    deepest = -1
    for index, column in enumerate(_BUILDING_SEGMENT_COLUMNS):
        value = _text(row[column]) if column < len(row) else ""
        if not value or value == "-":
            continue
        # A segment is the sector letters or a run of digits. Anything else in
        # a segment column is the row's label, which the workbook writes into
        # whichever column follows the last segment, and that column is
        # sometimes a segment column of the next level down.
        if not re.fullmatch(r"[A-Za-z]{1,3}|\d{1,3}", value):
            break
        filled[index] = value
        deepest = index
    return filled, deepest


def _building_label(row: tuple[Any, ...], deepest_column: int) -> str:
    """The row's own text: the first real cell after its deepest segment."""
    for column in range(deepest_column + 1, min(len(row), _BUILDING_SEGMENT_COLUMNS[-1] + 2)):
        candidate = _text(row[column])
        if candidate and candidate != "-":
            return candidate
    return ""


def _parse_building(workbook: Any) -> ImportedBOQ:
    """Read every chapter sheet of a building workbook into positions."""
    result = ImportedBOQ(source_format="xlsx", currency="HUF")
    chapters_seen: list[str] = []

    for sheet_name in workbook.sheetnames:
        match = _BUILDING_SHEET_RE.match(sheet_name)
        if not match:
            continue
        chapter = match.group(1)
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header = _building_header(rows)
        if header is None:
            result.warnings.append(
                {
                    "sheet": sheet_name,
                    "severity": "warning",
                    "message": f"Sheet {sheet_name} looks like a chapter sheet but carries no recognised header row.",
                }
            )
            continue
        header_index, columns = header
        chapters_seen.append(chapter)

        # The running code prefix, one slot per segment, and the text of the
        # heading a bare item row sits under.
        prefix: list[str | None] = [None] * len(_BUILDING_SEGMENT_COLUMNS)
        heading_label = ""

        for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            filled, deepest = _building_segments(row)
            description = _text(_at(row, columns, "description"))
            quantity = _number(_at(row, columns, "quantity"))
            unit = _text(_at(row, columns, "unit"))
            material = _number(_at(row, columns, "material_unit_rate"))
            fee = _number(_at(row, columns, "fee_unit_rate"))

            if deepest < 0:
                # The numbering row and the blank spacers between blocks.
                if description or quantity or unit:
                    result.skipped += 1
                continue

            for index, value in filled.items():
                prefix[index] = value
            for index in range(deepest + 1, len(prefix)):
                prefix[index] = None
            code = "-".join(value for value in prefix[: deepest + 1] if value)
            if not code.upper().startswith(_BUILDING_SECTOR):
                # A row whose composed code never reached the sector segment is
                # not a coded line: the sheet's own scaffolding, or a stray.
                result.skipped += 1
                continue

            label = _building_label(row, _BUILDING_SEGMENT_COLUMNS[deepest])
            priced = any(v not in (None, 0, 0.0) for v in (quantity, material, fee)) or bool(unit)

            if not priced:
                if not (description or label):
                    # An unfilled placeholder row. The blank template ships
                    # thousands of them, five under every item, and importing
                    # them as headings named after their own code would bury
                    # the structure the sheet is there to carry.
                    result.skipped += 1
                    continue
                # A heading. It carries the totals of everything beneath it,
                # which is a computed number rather than a price, so it is
                # imported as a section and priced at zero.
                if label:
                    heading_label = label
                result.positions.append(
                    ImportedPosition(
                        description=description or label,
                        ordinal=code,
                        unit="section",
                        quantity=0.0,
                        unit_rate=0.0,
                        classification={"tetelrend": code},
                        source="hungary_workbook",
                        metadata={
                            "import_row_index": row_index,
                            "section_header": True,
                            "hu": _building_hu(sheet_name, code, row, columns),
                        },
                        is_section=True,
                    )
                )
                continue

            text = description or label or heading_label
            if not text:
                # A priced row with nothing to call it is an unfilled template
                # line, and the template ships thousands of them.
                result.skipped += 1
                continue

            hu = _building_hu(sheet_name, code, row, columns)
            hu["material_unit_rate"] = material or 0.0
            hu["fee_unit_rate"] = fee or 0.0
            hu["material_total"] = _number(_at(row, columns, "material_total")) or 0.0
            hu["fee_total"] = _number(_at(row, columns, "fee_total")) or 0.0
            result.positions.append(
                ImportedPosition(
                    description=text,
                    ordinal=code,
                    unit=unit or "pcs",
                    quantity=quantity or 0.0,
                    # The rate is the two halves together. Reading either one
                    # as the rate would halve the bill.
                    unit_rate=(material or 0.0) + (fee or 0.0),
                    classification={"tetelrend": code},
                    source="hungary_workbook",
                    metadata={"import_row_index": row_index, "hu": hu},
                )
            )

    result.metadata = {
        "hu_profile": "building",
        "chapters": sorted(set(chapters_seen)),
        "sheet_names": list(workbook.sheetnames),
    }
    return result


def _building_hu(sheet_name: str, code: str, row: tuple[Any, ...], columns: dict[str, int]) -> dict[str, Any]:
    """The Hungarian payload common to both kinds of building row."""
    segments = code.split("-")
    hu: dict[str, Any] = {
        "profile": "building",
        "sheet": sheet_name,
        "chapter": segments[1] if len(segments) > 1 else "",
    }
    if len(segments) > 2:
        hu["subchapter"] = segments[2]
    catalogue = _text(_at(row, columns, "catalogue_item_number"))
    if catalogue:
        hu["catalogue_item_number"] = catalogue
    note = _text(_at(row, columns, "note"))
    if note:
        hu["note"] = note
    return hu


# ── Infrastructure profile ───────────────────────────────────────────────

_INFRA_COLUMNS: dict[str, tuple[str, ...]] = {
    "structure_code": ("struktura kod",),
    "row_number": ("sorszam",),
    "item_number": ("pde tetelszam",),
    "catalogue_item_number": ("tetel-szam", "tetelszam"),
    "description": ("megnevezes",),
    "unit": ("mertek egyseg", "mertekegyseg"),
    "quantity": ("mennyiseg",),
    "unit_rate": ("egysegar (ft)", "egysegar"),
    "total": ("osszesen (ft)", "osszesen"),
    "work_process": ("mfolyamat",),
    "start": ("kezdes",),
    "finish": ("befejezes",),
}


def _infra_header(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]] | None:
    """Find the header row of the infrastructure sheet and map its columns."""
    for index, row in enumerate(rows[:10]):
        folded = [_fold(cell) for cell in row]
        present = set(folded)
        if not present >= _INFRA_REQUIRED_HEADERS:
            continue
        mapping: dict[str, int] = {}
        for field, aliases in _INFRA_COLUMNS.items():
            for column, cell in enumerate(folded):
                if cell and cell in aliases:
                    mapping.setdefault(field, column)
                    break
        return index, mapping
    return None


def _parse_infrastructure(workbook: Any) -> ImportedBOQ:
    """Read the flat infrastructure sheet into positions."""
    result = ImportedBOQ(source_format="xlsx", currency="HUF")

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
        header = _infra_header(rows)
        if header is None:
            continue
        header_index, columns = header

        for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            description = _text(_at(row, columns, "description"))
            if not description:
                result.skipped += 1
                continue

            code = re.sub(r"\s+", "", _text(_at(row, columns, "catalogue_item_number")))
            quantity = _number(_at(row, columns, "quantity"))
            unit_rate = _number(_at(row, columns, "unit_rate"))

            hu: dict[str, Any] = {"profile": "infrastructure", "sheet": sheet_name}
            for field in ("structure_code", "row_number", "item_number", "work_process", "start", "finish"):
                value = _text(_at(row, columns, field))
                if value:
                    hu[field] = value

            unit = _text(_at(row, columns, "unit"))
            # A coding file carries no prices at all and still has a unit
            # on every real line, so the unit is what separates an item
            # awaiting a price from a heading that will never carry one.
            is_section = quantity is None and unit_rate is None and not unit
            result.positions.append(
                ImportedPosition(
                    description=description,
                    ordinal=_text(_at(row, columns, "row_number")) or code,
                    unit="section" if is_section else (unit or "pcs"),
                    quantity=0.0 if is_section else (quantity or 0.0),
                    unit_rate=0.0 if is_section else (unit_rate or 0.0),
                    classification={"tetelrend": code} if code else {},
                    source="hungary_workbook",
                    metadata={"import_row_index": row_index, "hu": hu}
                    | ({"section_header": True} if is_section else {}),
                    is_section=is_section,
                )
            )

        result.metadata = {
            "hu_profile": "infrastructure",
            "item_sheet": sheet_name,
            "dictionary_sheets": [name for name in workbook.sheetnames if name != sheet_name],
            "sheet_names": list(workbook.sheetnames),
        }
        break

    return result


# ── Entry point ──────────────────────────────────────────────────────────


def detect_profile(workbook: Any) -> str | None:
    """Which Hungarian profile an open workbook matches, or ``None``.

    Cheap enough to run on every spreadsheet upload: the building test reads
    sheet names only, and the infrastructure test reads the first ten rows of
    each sheet and stops at the first that answers.
    """
    if any(_BUILDING_SHEET_RE.match(name) for name in workbook.sheetnames):
        return "building"
    for sheet_name in workbook.sheetnames:
        rows = list(workbook[sheet_name].iter_rows(min_row=1, max_row=10, values_only=True))
        if _infra_header(rows) is not None:
            return "infrastructure"
    return None


def parse_hungarian_workbook(content: bytes) -> ImportedBOQ | None:
    """Parse an ``.xlsx`` upload if it is a Hungarian bill, else ``None``.

    Never raises. A workbook that opens but is not one of the two shapes is
    not this module's, and one that will not open at all is the generic
    reader's problem to report, with its own error text.
    """
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 - the generic reader reports this properly
        return None

    try:
        profile = detect_profile(workbook)
        if profile == "building":
            return _parse_building(workbook)
        if profile == "infrastructure":
            return _parse_infrastructure(workbook)
        return None
    except Exception as exc:  # noqa: BLE001 - fall through to the generic reader
        logger.warning("Hungarian workbook profile failed, falling back to the generic reader: %s", exc)
        return None
    finally:
        workbook.close()
