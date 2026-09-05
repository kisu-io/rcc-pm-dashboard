"""``GET /sets/{set_id}/comparison/`` and its spreadsheet twin.

The comparison is the point of the module: every option in a set is read from
its OWN bill, rebased to one comparison currency, laid out as columns and trade
rows, and topped with a transparent recommendation and an honest fairness
banner. These tests drive the real aggregator over seeded bills - only the
matcher is out of the picture here, because nothing in this path calls it.
"""

from __future__ import annotations

import io
import uuid

from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.boq.models import BOQMarkup
from app.modules.design_options.models import DesignOption
from tests.modules.design_options.conftest import (
    API_PREFIX,
    build_app,
    http_client,
    make_boq,
    make_option,
    make_position,
    make_project,
    make_set,
    make_user,
)


async def _priced_option(
    session: AsyncSession,
    option_set,  # noqa: ANN001 - DesignOptionSet, kept loose to avoid a circular import
    *,
    name: str,
    sort_order: int,
    lines: list[tuple[str, str, str, dict[str, str]]],
    gfa: str = "0",
    markup_pct: str | None = None,
) -> DesignOption:
    """Create an option with its own bill of quantities.

    Each entry in ``lines`` is ``(unit, quantity, unit_rate, classification)``.
    """
    boq = await make_boq(session, option_set.project_id, name=f"{name} bill")
    for index, (unit, quantity, unit_rate, classification) in enumerate(lines, start=1):
        await make_position(
            session,
            boq.id,
            ordinal=f"01.{index:03d}",
            unit=unit,
            quantity=quantity,
            unit_rate=unit_rate,
            total=str(float(quantity) * float(unit_rate)),
            classification=classification,
        )
    if markup_pct is not None:
        session.add(
            BOQMarkup(
                boq_id=boq.id,
                name="Overhead",
                markup_type="percentage",
                percentage=markup_pct,
                apply_to="direct_cost",
            )
        )
    return await make_option(
        session,
        option_set,
        name=name,
        sort_order=sort_order,
        boq_id=boq.id,
        gfa=gfa,
        status="priced",
    )


def _column(body: dict, name: str) -> dict:
    return next(col for col in body["options"] if col["name"] == name)


# ── Columns and deltas ───────────────────────────────────────────────────────


async def test_each_option_becomes_a_column_totalled_from_its_own_bill(session: AsyncSession) -> None:
    """Two options with different bills produce two independent columns."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id, name="Frame options")
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
        markup_pct="10",
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "150", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["set_name"] == "Frame options"
    assert [col["name"] for col in body["options"]] == ["Steel", "Timber"]

    steel = _column(body, "Steel")
    assert steel["direct_cost"] == "1000.00"
    assert steel["markups_total"] == "100.00"
    assert steel["grand_total"] == "1100.00"
    # Cost per m2 uses direct cost against the project floor area.
    assert steel["cost_per_m2"] == "10.00"
    assert steel["gfa"] == "100"
    assert steel["position_count"] == 1
    assert steel["currency"] == "EUR"

    timber = _column(body, "Timber")
    assert timber["grand_total"] == "1500.00"


async def test_deltas_are_measured_against_the_chosen_baseline(session: AsyncSession) -> None:
    """The baseline reads zero; the others read their signed difference."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    baseline = await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "125", {"din276": "330"})],
    )
    option_set.baseline_option_id = baseline.id
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["baseline_option_id"] == str(baseline.id)
    assert _column(body, "Steel")["delta_vs_baseline"] == "0.00"
    assert _column(body, "Steel")["delta_pct"] == "0.00"
    assert _column(body, "Timber")["delta_vs_baseline"] == "250.00"
    assert _column(body, "Timber")["delta_pct"] == "25.00"


async def test_without_a_baseline_no_percentage_is_invented(session: AsyncSession) -> None:
    """Deltas stay at zero and the percentage is null, not a fabricated number."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "125", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["baseline_option_id"] is None
    assert all(col["delta_pct"] is None for col in body["options"])
    assert all(col["delta_vs_baseline"] == "0" for col in body["options"])


async def test_an_unpriced_draft_still_gets_a_column(session: AsyncSession) -> None:
    """A draft option is shown with zeros rather than dropped from the set."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await make_option(session, option_set, name="Sketch", sort_order=1)
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    sketch = _column(body, "Sketch")
    assert sketch["grand_total"] == "0.00"
    assert sketch["position_count"] == 0


async def test_an_empty_set_compares_cleanly(session: AsyncSession) -> None:
    """A set with no options answers with empty tables, not an error."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id, name="Empty")
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["options"] == []
    assert body["by_trade"] == []
    assert body["recommendation"]["option_id"] is None
    assert body["recommendation"]["reason_key"] == "designOptions.recommendation.none"


# ── Comparison currency ──────────────────────────────────────────────────────


async def test_a_blank_comparison_currency_uses_the_project_base(session: AsyncSession) -> None:
    """No requested currency means the project base, with no warning."""
    user = await make_user(session)
    project = await make_project(session, user.id, currency="GBP")
    option_set = await make_set(session, project.id, comparison_currency="")
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["comparison_currency"] == "GBP"
    keys = {w["key"] for w in body["fairness"]["warnings"]}
    assert "designOptions.fairness.comparisonCurrencyUnavailable" not in keys


async def test_a_requested_currency_with_a_rate_rebases_every_column(session: AsyncSession) -> None:
    """One uniform factor relabels the whole set, leaving the ranking intact."""
    user = await make_user(session)
    project = await make_project(
        session,
        user.id,
        currency="EUR",
        gross_floor_area="100",
        fx_rates=[{"code": "USD", "rate": "0.5"}],
    )
    option_set = await make_set(session, project.id, comparison_currency="USD")
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    # The project rate is base-per-foreign, so a EUR 1000 bill at 0.5 EUR per
    # USD presents as USD 2000.
    assert body["comparison_currency"] == "USD"
    assert _column(body, "Steel")["grand_total"] == "2000.00"
    assert _column(body, "Steel")["currency"] == "USD"


async def test_a_requested_currency_without_a_rate_stays_in_base_and_says_so(session: AsyncSession) -> None:
    """The label never lies: with no usable rate the numbers stay in base."""
    user = await make_user(session)
    project = await make_project(session, user.id, currency="EUR")
    option_set = await make_set(session, project.id, comparison_currency="JPY")
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    assert body["comparison_currency"] == "EUR"
    assert _column(body, "Steel")["grand_total"] == "1000.00"
    notice = next(
        w for w in body["fairness"]["warnings"] if w["key"] == "designOptions.fairness.comparisonCurrencyUnavailable"
    )
    assert notice["context"] == {"requested": "JPY", "used": "EUR"}


# ── By-trade table ───────────────────────────────────────────────────────────


async def test_trade_rows_cover_every_option_and_rank_by_baseline_cost(session: AsyncSession) -> None:
    """A trade only one option prices still gets a row, with a zero cell."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id)
    baseline = await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[
            ("m3", "10", "300", {"din276": "330"}),
            ("m2", "20", "50", {"din276": "420"}),
        ],
    )
    timber = await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "12", "200", {"din276": "330"})],
    )
    option_set.baseline_option_id = baseline.id
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    rows = res.json()["by_trade"]
    assert [row["key"] for row in rows] == ["300", "400"]

    structure = rows[0]
    assert structure["label"] == "Building construction"
    assert structure["classification_system"] == "din276"
    assert structure["baseline_cost"] == "3000.00"
    assert structure["baseline_quantity"] == "10"
    cells = {cell["option_id"]: cell for cell in structure["per_option"]}
    assert cells[str(timber.id)]["cost"] == "2400.00"
    assert cells[str(timber.id)]["unit"] == "m3"

    services = rows[1]
    services_cells = {cell["option_id"]: cell for cell in services["per_option"]}
    assert services_cells[str(baseline.id)]["cost"] == "1000.00"
    # Timber prices nothing in this trade, so its cell is an explicit zero.
    assert services_cells[str(timber.id)]["cost"] == "0"
    assert services_cells[str(timber.id)]["unit"] == ""


async def test_an_unclassified_line_lands_in_its_own_bucket(session: AsyncSession) -> None:
    """An unclassified position is bucketed, never dropped from the totals."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    rows = res.json()["by_trade"]
    assert [row["key"] for row in rows] == ["unclassified"]
    assert rows[0]["classification_system"] == "none"


# ── Recommendation ───────────────────────────────────────────────────────────


async def test_the_cheapest_cost_per_m2_is_recommended(session: AsyncSession) -> None:
    """The rule is explainable: lowest normalised cost per m2 wins."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    cheap = await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "75", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    recommendation = res.json()["recommendation"]
    assert recommendation["option_id"] == str(cheap.id)
    assert recommendation["reason_key"] == "designOptions.recommendation.lowestCostPerM2"
    # 750 against a runner-up of 1000 is a quarter clear.
    assert recommendation["confidence"] == "0.25"


async def test_a_lone_priced_option_is_recommended_as_such(session: AsyncSession) -> None:
    """With one candidate the reason says so instead of claiming a contest."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    only = await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await make_option(session, option_set, name="Sketch", sort_order=1)
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    recommendation = res.json()["recommendation"]
    assert recommendation["option_id"] == str(only.id)
    assert recommendation["reason_key"] == "designOptions.recommendation.onlyOption"


async def test_without_a_floor_area_the_lowest_total_wins(session: AsyncSession) -> None:
    """The fallback is the lowest grand total, and it names itself."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area=None)
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    cheap = await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "60", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    recommendation = res.json()["recommendation"]
    assert recommendation["option_id"] == str(cheap.id)
    assert recommendation["reason_key"] == "designOptions.recommendation.lowestTotal"


async def test_an_option_whose_bill_mixes_currencies_is_not_recommended(session: AsyncSession) -> None:
    """A blended bill is disqualified from winning on price."""
    user = await make_user(session)
    project = await make_project(session, user.id, currency="EUR", gross_floor_area="100")
    option_set = await make_set(session, project.id)
    clean = await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    mixed_boq = await make_boq(session, project.id, name="Timber bill")
    await make_position(session, mixed_boq.id, ordinal="01.001", quantity="1", unit_rate="10", total="10")
    await make_position(
        session,
        mixed_boq.id,
        ordinal="01.002",
        quantity="1",
        unit_rate="10",
        total="10",
        metadata_={"currency": "USD"},
    )
    await make_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        boq_id=mixed_boq.id,
        status="priced",
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    # Timber is far cheaper but its bill blends currencies, so it cannot win.
    assert body["recommendation"]["option_id"] == str(clean.id)
    keys = {w["key"] for w in body["fairness"]["warnings"]}
    assert "designOptions.fairness.mixedCurrencyOption" in keys


# ── Fairness banner ──────────────────────────────────────────────────────────


async def test_a_single_option_set_says_there_is_nothing_to_compare(session: AsyncSession) -> None:
    """One option is an informational notice, not a clean comparison."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    keys = {w["key"] for w in res.json()["fairness"]["warnings"]}
    assert "designOptions.fairness.singleOption" in keys


async def test_unpriced_options_are_counted_in_the_banner(session: AsyncSession) -> None:
    """An unpriced option makes the banner amber and says how many."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await make_option(session, option_set, name="Sketch", sort_order=1)
    await make_option(session, option_set, name="Napkin", sort_order=2)
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    notice = next(w for w in body["fairness"]["warnings"] if w["key"] == "designOptions.fairness.unpricedOptions")
    assert notice["context"]["count"] == 2
    assert notice["severity"] == "warning"
    assert body["fairness"]["status"] in ("warnings", "error")


async def test_two_priced_options_without_a_baseline_are_flagged(session: AsyncSession) -> None:
    """Nothing is being measured against anything, and the banner says so."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "120", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    keys = {w["key"] for w in res.json()["fairness"]["warnings"]}
    assert "designOptions.fairness.noBaseline" in keys


async def test_a_priced_option_with_no_floor_area_is_flagged(session: AsyncSession) -> None:
    """Without an area there is no cost per m2 to compare on."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area=None)
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    keys = {w["key"] for w in res.json()["fairness"]["warnings"]}
    assert "designOptions.fairness.missingGfa" in keys


async def test_options_with_different_floor_areas_are_flagged(session: AsyncSession) -> None:
    """Comparing cost per m2 across different programmes gets a notice."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
        gfa="100",
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "100", {"din276": "330"})],
        gfa="180",
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    keys = {w["key"] for w in res.json()["fairness"]["warnings"]}
    assert "designOptions.fairness.mixedGfa" in keys


async def test_the_same_floor_area_written_two_ways_is_not_a_difference(session: AsyncSession) -> None:
    """Equal areas must compare as equal, whatever their stored spelling.

    Floor areas are stored as decimal strings, so an area edited from "100" to
    "100.00" is the same number written differently. Comparing the rendered
    strings rather than the values would report a programme difference that
    does not exist.
    """
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
        gfa="100",
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m3", "10", "100", {"din276": "330"})],
        gfa="100.00",
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    keys = {w["key"] for w in res.json()["fairness"]["warnings"]}
    assert "designOptions.fairness.mixedGfa" not in keys


# ── Validation wiring ────────────────────────────────────────────────────────


async def test_the_comparison_runs_the_modules_own_validation_rules(session: AsyncSession) -> None:
    """The banner carries rule findings, and never claims validation is pending.

    This is the one assertion that proves the comparison actually reaches the
    ``design_options`` rule set. Two options measure the same trade in different
    units, which the unit-consistency rule rejects; if the rules were not
    registered the engine would report the set as unsupported, the banner would
    carry the honest "not validated yet" notice instead, and every rule in this
    module would be dead code in production.
    """
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await _priced_option(
        session,
        option_set,
        name="Timber",
        sort_order=1,
        lines=[("m2", "40", "25", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    body = res.json()
    keys = {w["key"] for w in body["fairness"]["warnings"]}
    assert "designOptions.validation.design_options.unit_consistency" in keys
    assert "designOptions.fairness.validationPending" not in keys
    assert body["fairness"]["status"] == "error"


async def test_a_validated_option_column_carries_a_traffic_light(session: AsyncSession) -> None:
    """The per-option status comes from the engine, not the persisted default."""
    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id)
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison/")

    assert res.status_code == 200, res.text
    # "pending" is the persisted floor, so anything else proves a rule ran.
    assert _column(res.json(), "Steel")["validation_status"] != "pending"


# ── Spreadsheet export ───────────────────────────────────────────────────────


async def test_the_export_returns_a_readable_workbook(session: AsyncSession) -> None:
    """The download is a real .xlsx carrying both appraisal sheets."""
    from openpyxl import load_workbook

    user = await make_user(session)
    project = await make_project(session, user.id, gross_floor_area="100")
    option_set = await make_set(session, project.id, name="Frame options")
    await _priced_option(
        session,
        option_set,
        name="Steel",
        sort_order=0,
        lines=[("m3", "10", "100", {"din276": "330"})],
    )
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison.xlsx")

    assert res.status_code == 200, res.text
    assert res.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "Design Option Appraisal - Frame options.xlsx" in res.headers["content-disposition"]
    assert res.headers["content-length"] == str(len(res.content))

    workbook = load_workbook(io.BytesIO(res.content))
    assert workbook.sheetnames == ["Option Appraisal", "By Trade"]
    rendered = "\n".join(
        str(cell.value)
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "Steel" in rendered
    assert "Frame options" in rendered


async def test_the_export_survives_a_set_name_that_would_break_a_header(session: AsyncSession) -> None:
    """A crafted set name cannot inject a line break into the response header."""
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id, name="Frames\r\nX-Injected: 1")
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison.xlsx")

    assert res.status_code == 200, res.text
    assert "x-injected" not in {k.lower() for k in res.headers}
    assert "\n" not in res.headers["content-disposition"]


async def test_the_export_of_an_empty_set_is_still_a_workbook(session: AsyncSession) -> None:
    """Nothing to appraise still downloads rather than erroring."""
    from openpyxl import load_workbook

    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id, name="Empty")
    await session.commit()

    app = build_app(session, caller_id=user.id)
    async with http_client(app) as client:
        res = await client.get(f"{API_PREFIX}/sets/{option_set.id}/comparison.xlsx")

    assert res.status_code == 200, res.text
    workbook = load_workbook(io.BytesIO(res.content))
    assert workbook.sheetnames == ["Option Appraisal", "By Trade"]


async def test_a_missing_project_is_not_a_crash(session: AsyncSession) -> None:
    """A set whose project row is gone still reports, with an empty currency.

    The project link cascades, so this is defensive rather than reachable
    through the API; the aggregator resolves its context with ``getattr``
    defaults for exactly this reason.
    """
    user = await make_user(session)
    project = await make_project(session, user.id)
    option_set = await make_set(session, project.id)
    await make_option(session, option_set, name="Steel")
    await session.commit()

    from app.modules.design_options.comparison import DesignOptionComparator
    from app.modules.design_options.service import DesignOptionsService

    fetched = await DesignOptionsService(session).get_set(option_set.id)
    # Detach before repointing, so the rewritten project id is only a read-time
    # fact and never reaches the database as an update.
    session.expunge(fetched)
    fetched.project_id = uuid.uuid4()
    comparison = await DesignOptionComparator(session).build(fetched)
    assert comparison.comparison_currency == ""
    assert len(comparison.options) == 1
