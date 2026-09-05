"""DB-free tests for the approvals-register Excel export.

Mirrors the style of ``test_notifications.py``: build transient
workflow/step objects in memory and exercise the pure workbook builder
``build_approvals_workbook`` - no database, no HTTP. We assert the workbook
is well-formed AND that the OWASP CSV-injection neutralisation is applied to
every user-controlled cell (a ``=cmd|...`` payload comes back
apostrophe-prefixed, so a colleague opening the file never triggers a
formula).

Run: pytest backend/tests/modules/file_approvals/test_register_export.py -q
"""

import io
import uuid
from datetime import UTC, datetime

from openpyxl import load_workbook

from app.modules.file_approvals.models import FileApprovalStep, FileApprovalWorkflow
from app.modules.file_approvals.router import _EXPORT_HEADERS, build_approvals_workbook


def _step(
    order: int,
    *,
    decision: str = "pending",
    role_label: str | None = None,
    decision_note: str | None = None,
    approver: uuid.UUID | None = None,
) -> FileApprovalStep:
    return FileApprovalStep(
        id=uuid.uuid4(),
        sort_order=order,
        approver_id=approver or uuid.uuid4(),
        role_label=role_label,
        decision=decision,
        decision_note=decision_note,
    )


def _wf(
    *,
    status: str = "in_review",
    steps: list[FileApprovalStep] | None = None,
    notes: str | None = None,
    file_id: str = "drawing-A-101",
    version: str | None = "C",
) -> FileApprovalWorkflow:
    wf = FileApprovalWorkflow(
        id=uuid.uuid4(),
        project_id=uuid.uuid4(),
        file_kind="document",
        file_id=file_id,
        file_version_snapshot=version,
        submitted_by_id=uuid.uuid4(),
        submitted_at=datetime(2026, 7, 24, 9, 30, tzinfo=UTC),
        status=status,
        notes=notes,
    )
    wf.steps = steps or []
    return wf


def _roundtrip(wb):
    """Save the workbook to bytes and reopen it, proving it serialises."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf).active


# ── shape ───────────────────────────────────────────────────────────────────


def test_workbook_builds_with_header_and_one_row_per_workflow():
    wfs = [
        _wf(steps=[_step(0, decision="approved", role_label="Engineer"), _step(1, role_label="PM")]),
        _wf(status="approved", steps=[_step(0, decision="approved", role_label="Engineer")]),
    ]
    ws = _roundtrip(build_approvals_workbook(wfs))
    assert ws.title == "Approval Register"
    assert tuple(c.value for c in ws[1]) == _EXPORT_HEADERS
    # One data row per workflow, plus the header row.
    assert ws.max_row == 1 + len(wfs)


def test_empty_register_still_produces_a_header_only_sheet():
    ws = _roundtrip(build_approvals_workbook([]))
    assert tuple(c.value for c in ws[1]) == _EXPORT_HEADERS
    assert ws.max_row == 1


# ── derived "current step / approver" ────────────────────────────────────────


def test_current_step_is_the_first_pending_approver():
    wf = _wf(
        steps=[
            _step(0, decision="approved", role_label="Engineer"),
            _step(1, decision="pending", role_label="Project Manager"),
        ]
    )
    ws = _roundtrip(build_approvals_workbook([wf]))
    # Column 7 = "Current Step / Approver": first pending step, 1-based ordinal.
    assert ws.cell(row=2, column=7).value == "#2: Project Manager"


def test_terminal_workflow_has_no_current_step():
    wf = _wf(status="approved", steps=[_step(0, decision="approved", role_label="Engineer")])
    ws = _roundtrip(build_approvals_workbook([wf]))
    # No pending step -> blank cell (openpyxl reads an empty cell back as None).
    assert ws.cell(row=2, column=7).value in (None, "")


# ── CSV-injection neutralisation (the security assertion) ────────────────────


def test_notes_formula_is_neutralised():
    payload = "=cmd|'/c calc'!A0"
    wf = _wf(notes=payload, steps=[_step(0, role_label="Engineer")])
    ws = _roundtrip(build_approvals_workbook([wf]))
    # Column 11 = Notes. The OWASP defence prepends a single apostrophe so
    # Excel renders the text literally instead of evaluating the formula.
    assert ws.cell(row=2, column=11).value == "'" + payload


def test_step_role_label_formula_is_neutralised():
    # A malicious role_label would otherwise open the step-digest cell with
    # a leading '=' and be evaluated on open.
    wf = _wf(steps=[_step(0, decision="pending", role_label='=HYPERLINK("http://evil")')])
    ws = _roundtrip(build_approvals_workbook([wf]))
    # Column 8 = Steps / Decisions; the digest starts with the role label.
    assert ws.cell(row=2, column=8).value.startswith("'=")


# ── tz-aware timestamps ──────────────────────────────────────────────────────


def test_tz_aware_dates_are_stringified_not_rejected():
    # openpyxl raises on tz-aware datetimes; the builder ISO-formats them.
    wf = _wf(steps=[_step(0, role_label="Engineer")])
    ws = _roundtrip(build_approvals_workbook([wf]))
    assert ws.cell(row=2, column=5).value == "2026-07-24T09:30:00+00:00"
