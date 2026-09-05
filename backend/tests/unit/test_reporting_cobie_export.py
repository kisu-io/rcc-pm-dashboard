# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""Unit tests for the reporting module's COBie export profile.

Mirrors the structural-assertion style of
``backend/tests/unit/test_cobie_exporter.py`` (sheet names + header rows,
not a byte-for-byte snapshot, since openpyxl embeds build metadata). This
file only tests what the reporting module adds on top of the canonical
``bim_hub`` COBie builder: the ``(filename, media_type, blob)`` exporter
shape and the extra header-only ``Zone`` sheet.
"""

from __future__ import annotations

import uuid
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.modules.reporting.exporters import (
    COBIE_ADDITIONAL_SHEETS,
    COBIE_MEDIA_TYPE,
    export_project_cobie,
)


def _fixture_model() -> SimpleNamespace:
    return SimpleNamespace(
        id=uuid.UUID("11111111-1111-1111-1111-111111111111"),
        project_id=uuid.UUID("22222222-2222-2222-2222-222222222222"),
        name="Skyline Tower",
        discipline="Architectural",
    )


def _fixture_elements() -> list[SimpleNamespace]:
    return [
        SimpleNamespace(
            stable_id="room-101",
            element_type="IfcSpace",
            name="Office 101",
            storey="Floor 1",
            discipline="Architectural",
            asset_info={},
            is_tracked_asset=False,
            quantities={"area": 42.5, "height": 3.0},
            properties={},
        ),
        SimpleNamespace(
            stable_id="ahu-01",
            element_type="AirHandlingUnit",
            name="AHU-01",
            storey="Floor 1",
            discipline="MEP",
            asset_info={
                "manufacturer": "Siemens",
                "model": "SV-100",
                "serial_number": "SN-AHU-001",
                "parent_system": "HVAC",
                "asset_tag": "AHU-01",
            },
            is_tracked_asset=True,
            quantities={},
            properties={},
        ),
    ]


def _workbook_from_bytes(xlsx: bytes):
    return load_workbook(BytesIO(xlsx), read_only=True, data_only=True)


class TestExportProjectCobie:
    def test_returns_exporter_shaped_tuple(self):
        filename, media_type, blob = export_project_cobie(_fixture_model(), _fixture_elements())
        assert filename == "COBie_Skyline Tower.xlsx"
        assert media_type == COBIE_MEDIA_TYPE
        assert isinstance(blob, bytes)
        assert len(blob) > 0

    def test_workbook_has_canonical_sheets_plus_zone(self):
        _, _, blob = export_project_cobie(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(blob)
        expected = [
            "Contact",
            "Facility",
            "Floor",
            "Space",
            "Type",
            "Component",
            "System",
            "Zone",
        ]
        assert wb.sheetnames == expected

    def test_zone_sheet_is_header_only(self):
        _, _, blob = export_project_cobie(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(blob)
        ws = wb["Zone"]
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1
        assert list(rows[0]) == COBIE_ADDITIONAL_SHEETS["Zone"]

    def test_underlying_sheets_still_carry_element_data(self):
        _, _, blob = export_project_cobie(_fixture_model(), _fixture_elements())
        wb = _workbook_from_bytes(blob)
        space_rows = list(wb["Space"].iter_rows(min_row=2, values_only=True))
        component_rows = list(wb["Component"].iter_rows(min_row=2, values_only=True))
        assert len(space_rows) == 1
        assert len(component_rows) == 1

    def test_filename_falls_back_when_model_has_no_name(self):
        model = SimpleNamespace(id=uuid.uuid4(), project_id=uuid.uuid4(), name=None)
        filename, _, _ = export_project_cobie(model, [])
        assert filename == "COBie_model.xlsx"

    def test_empty_elements_still_produces_all_sheets_with_headers_only(self):
        _, _, blob = export_project_cobie(_fixture_model(), [])
        wb = _workbook_from_bytes(blob)
        # Facility and Contact always carry exactly one row (synthesised);
        # Floor falls back to a single "Floor 1" placeholder; Space / Type /
        # Component / System / Zone are header-only with no data.
        for sheet_name in ("Space", "Type", "Component", "System", "Zone"):
            rows = list(wb[sheet_name].iter_rows(min_row=2, values_only=True))
            assert rows == [], f"{sheet_name} should have no data rows"
