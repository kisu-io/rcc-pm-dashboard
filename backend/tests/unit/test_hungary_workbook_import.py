# DDC-CWICR-OE: DataDrivenConstruction · OpenConstructionERP
# Copyright (c) 2026 Artem Boiko / DataDrivenConstruction
"""The Hungarian workbook profiles, driven by workbooks built to the real shape.

Every fixture here reproduces something the delivered files actually do, and
each one is here because writing the profile without it produced a wrong
answer that still looked plausible:

* the item code is not in a column. A heading writes the segments down to its
  own level and the lines under it write only the segment that changed, so the
  code of a row is the running composition of everything above it. Reading
  each row on its own gave item codes like ``001``.
* a chapter sheet names the standard on its header row, and the Hungarian word
  for "unit" is a substring of both that title and the word for "quantity". A
  header matcher that took the first substring hit read the sector letters out
  of the title as a unit, which made every heading look priced, and a heading
  that looks priced imports as an item carrying the total of everything under
  it.
* the blank template ships five unfilled placeholder rows under every item.
  Importing those as headings named after their own code buried the structure.
* a priced line carries two unit prices. Taking either one as the rate halves
  the bill.
* the infrastructure file in hand is a coding file with no prices at all and a
  unit on two thirds of its lines, so the unit is what separates an item
  awaiting a price from a heading that will never carry one.

No database, no application settings: openpyxl in, dataclasses out.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.modules.boq.importers.hungary_workbook import (
    detect_profile,
    parse_hungarian_workbook,
)

# The header block of a chapter sheet, as the real workbook lays it out. The
# code segments live in columns B, D, F, H, J, L, N, P and R with separator
# columns between them, which is why the fixtures below address cells by
# letter rather than appending rows.
_HEADER = {
    "B": "EGYSÉGES MAGASÉPÍTÉSI ÁGAZATI TÉTELREND",
    "S": "Külső tételszám",
    "T": "Tétel szövege",
    "U": "Mennyiség",
    "V": "Egység",
    "X": "Anyag egységár",
    "Y": "Díj egységár",
    "Z": "Nettó anyag összesen (Ft)",
    "AA": "Nettó díj összesen (Ft)",
    "AB": "Nettó A+D összesen",
    "AC": "Megjegyzés",
}


def _building_workbook(rows: list[dict[str, object]], sheet: str = "MA-01_ÁLT") -> bytes:
    """A one-chapter building workbook with the header block and given rows.

    ``rows`` are dicts of column letter to value, written from row 6 down, so
    the header sits where the real sheets put it.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Címlap"
    chapter = workbook.create_sheet(sheet)
    for column, value in _HEADER.items():
        chapter[f"{column}5"] = value
    for index, row in enumerate(rows, start=6):
        for column, value in row.items():
            chapter[f"{column}{index}"] = value
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


_INFRA_HEADER = {
    "P": "Struktúra kód",
    "Q": "Sorszám",
    "R": "PDE tételszám",
    "X": "TÉTEL-SZÁM",
    "Y": "MEGNEVEZÉS",
    "Z": "MÉRTÉK EGYSÉG",
    "AA": "MENNYISÉG",
    "AB": "EGYSÉGÁR (Ft)",
    "AC": "ÖSSZESEN (Ft)",
    "AG": "MFOLYAMAT",
    "AQ": "kezdés",
    "AR": "befejezés",
}


def _infra_workbook(rows: list[dict[str, object]]) -> bytes:
    """A flat infrastructure sheet with the header on row 1, plus a dictionary sheet."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CSV kódolás"
    for column, value in _INFRA_HEADER.items():
        worksheet[f"{column}1"] = value
    for index, row in enumerate(rows, start=2):
        for column, value in row.items():
            worksheet[f"{column}{index}"] = value
    workbook.create_sheet("V03016TKOD kódszótár")
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── Recognition ──────────────────────────────────────────────────────────


def test_a_spreadsheet_that_is_neither_shape_is_handed_back() -> None:
    """The profile must decline, not guess. The generic reader owns everything else."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for column, value in enumerate(("Description", "Unit", "Quantity", "Rate"), start=1):
        sheet.cell(row=1, column=column, value=value)
    sheet.append(["Concrete C30/37", "m3", 12, 140])
    buffer = io.BytesIO()
    workbook.save(buffer)
    assert parse_hungarian_workbook(buffer.getvalue()) is None


def test_bytes_that_are_not_a_workbook_at_all_are_handed_back() -> None:
    assert parse_hungarian_workbook(b"not a zip, not a workbook") is None


def test_the_two_shapes_are_told_apart_by_what_they_carry() -> None:
    from openpyxl import load_workbook

    building = load_workbook(io.BytesIO(_building_workbook([])), read_only=True)
    infra = load_workbook(io.BytesIO(_infra_workbook([])), read_only=True)
    assert detect_profile(building) == "building"
    assert detect_profile(infra) == "infrastructure"


# ── Building profile ─────────────────────────────────────────────────────


def _rows_like_the_template() -> list[dict[str, object]]:
    """The opening of a real chapter sheet: headings, then one priced line."""
    return [
        {"B": "1.", "C": "-", "D": "2.", "E": "-", "F": "3.", "G": "-", "H": "4.", "I": "-", "J": "5."},
        {},
        {"B": "MA", "C": "-", "D": "MAGASÉPÍTÉSI MUNKÁK"},
        {},
        {"B": "MA", "C": "-", "D": "01", "E": "ÁLTALÁNOS, JÁRULÉKOS KÖLTSÉGEK", "Z": 1000000, "AA": 100000},
        {"B": "MA", "C": "-", "D": "01", "E": "-", "F": "11", "G": "FELVONULÁSI, IDEIGLENES LÉTESÍTMÉNYEK"},
        {"B": "MA", "C": "-", "D": "01", "E": "-", "F": "11", "G": "-", "H": "01", "I": "Ideiglenes utak, hidak"},
        {"T": "Ideiglenes útalap zúzottkőből", "U": 10, "V": "m2", "X": 100000, "Y": 10000, "R": "001"},
        {"R": "002"},
        {"R": "003"},
    ]


def test_a_rows_code_is_composed_from_the_headings_above_it() -> None:
    result = parse_hungarian_workbook(_building_workbook(_rows_like_the_template()))
    assert result is not None
    codes = [p.classification["tetelrend"] for p in result.positions]
    assert codes[:4] == ["MA", "MA-01", "MA-01-11", "MA-01-11-01"]
    # The priced line writes only its own last segment and inherits the rest.
    priced = [p for p in result.positions if not p.is_section]
    assert len(priced) == 1
    assert priced[0].classification["tetelrend"] == "MA-01-11-01-001"


def test_the_rate_is_the_two_halves_together() -> None:
    result = parse_hungarian_workbook(_building_workbook(_rows_like_the_template()))
    assert result is not None
    priced = next(p for p in result.positions if not p.is_section)
    assert priced.metadata["hu"]["material_unit_rate"] == 100000
    assert priced.metadata["hu"]["fee_unit_rate"] == 10000
    assert priced.unit_rate == 110000, "the rate must be material plus fee, not either half"
    assert priced.quantity == 10
    assert priced.unit == "m2"


def test_headings_import_as_sections_priced_at_zero() -> None:
    """A heading carries the total of everything beneath it, which is not a price."""
    result = parse_hungarian_workbook(_building_workbook(_rows_like_the_template()))
    assert result is not None
    heading = next(p for p in result.positions if p.classification["tetelrend"] == "MA-01")
    assert heading.is_section
    assert heading.unit == "section"
    assert heading.unit_rate == 0.0
    assert heading.description == "ÁLTALÁNOS, JÁRULÉKOS KÖLTSÉGEK"


def test_the_unit_column_is_not_stolen_by_the_sheet_title() -> None:
    """The regression that made every heading look priced.

    The Hungarian for "unit" is a substring of the sheet's own title on the
    header row and of the word for "quantity". If the matcher takes the first
    substring hit, the unit column lands on the title cell, every heading reads
    the sector letters as its unit, and the whole sheet imports as priced items
    carrying rolled-up totals.
    """
    result = parse_hungarian_workbook(_building_workbook(_rows_like_the_template()))
    assert result is not None
    sections = [p for p in result.positions if p.is_section]
    assert len(sections) == 4, f"expected four headings, got {[p.ordinal for p in result.positions]}"
    assert all(p.unit_rate == 0.0 for p in sections)


def test_unfilled_placeholder_rows_are_skipped() -> None:
    result = parse_hungarian_workbook(_building_workbook(_rows_like_the_template()))
    assert result is not None
    assert result.skipped >= 2, "the two empty placeholder rows should not become positions"
    assert not any(p.description in ("MA-01-11-01-002", "MA-01-11-01-003") for p in result.positions)


def test_every_chapter_sheet_is_read_not_just_the_active_one() -> None:
    """The generic reader takes the active worksheet, which on these files is the cover."""
    workbook = Workbook()
    workbook.active.title = "Címlap"
    for chapter in ("01", "07", "17"):
        sheet = workbook.create_sheet(f"MA-{chapter}_X")
        for column, value in _HEADER.items():
            sheet[f"{column}5"] = value
        sheet["B6"], sheet["C6"], sheet["D6"], sheet["E6"] = "MA", "-", chapter, f"Chapter {chapter}"
    buffer = io.BytesIO()
    workbook.save(buffer)
    result = parse_hungarian_workbook(buffer.getvalue())
    assert result is not None
    assert result.metadata["chapters"] == ["01", "07", "17"]
    assert {p.classification["tetelrend"] for p in result.positions} == {"MA-01", "MA-07", "MA-17"}


def test_a_deeper_row_clears_the_levels_below_the_one_it_sets() -> None:
    """Two items under the same heading must not inherit each other's tail."""
    rows = [
        {"B": "MA", "C": "-", "D": "03", "E": "-", "F": "31", "G": "-", "H": "01", "I": "Földmunka"},
        {"T": "Kitermelés", "U": 5, "V": "m3", "X": 100, "Y": 50, "R": "001"},
        {"B": "MA", "C": "-", "D": "03", "E": "-", "F": "31", "G": "-", "H": "02", "I": "Feltöltés"},
        {"T": "Visszatöltés", "U": 5, "V": "m3", "X": 100, "Y": 50, "R": "001"},
    ]
    result = parse_hungarian_workbook(_building_workbook(rows))
    assert result is not None
    priced = [p.classification["tetelrend"] for p in result.positions if not p.is_section]
    assert priced == ["MA-03-31-01-001", "MA-03-31-02-001"]


# ── Infrastructure profile ───────────────────────────────────────────────


def _infra_rows() -> list[dict[str, object]]:
    return [
        {"P": "1", "Q": 100, "R": "100_7", "X": "7", "Y": "Balatonfenyves vasútállomás"},
        {"P": "1+1", "Q": 101, "R": "101_000 000", "X": "000 000", "Y": "Általános tételek"},
        {
            "P": "1+1+1",
            "Q": 103,
            "R": "103_010 010",
            "X": "010 010",
            "Y": "Tervezés",
            "Z": "Ft",
            "AG": "10010",
            "AQ": "2023-01-15",
            "AR": "2023-03-15",
        },
        {
            "P": "1+1+2",
            "Q": 104,
            "R": "104_1000000_k",
            "X": "1000000_k",
            "Y": "Mérföldkő",
            "Z": "db",
            "AG": "1000000_k",
        },
    ]


def test_the_item_number_and_the_structure_code_survive_the_import() -> None:
    result = parse_hungarian_workbook(_infra_workbook(_infra_rows()))
    assert result is not None
    by_row = {p.metadata["hu"]["row_number"]: p for p in result.positions}
    line = by_row["103"]
    assert line.metadata["hu"]["item_number"] == "103_010 010"
    assert line.metadata["hu"]["structure_code"] == "1+1+1"
    assert line.metadata["hu"]["work_process"] == "10010"
    assert line.classification["tetelrend"] == "010010", "the space inside the code is presentation"


def test_a_line_with_a_unit_and_no_price_is_an_item_not_a_heading() -> None:
    """The delivered coding file carries no prices at all and is still a bill."""
    result = parse_hungarian_workbook(_infra_workbook(_infra_rows()))
    assert result is not None
    items = [p for p in result.positions if not p.is_section]
    assert {p.metadata["hu"]["row_number"] for p in items} == {"103", "104"}
    assert all(p.unit_rate == 0.0 for p in items)


def test_the_programme_dates_come_across() -> None:
    result = parse_hungarian_workbook(_infra_workbook(_infra_rows()))
    assert result is not None
    line = next(p for p in result.positions if p.metadata["hu"]["row_number"] == "103")
    assert line.metadata["hu"]["start"].startswith("2023-01-15")
    assert line.metadata["hu"]["finish"].startswith("2023-03-15")


def test_the_dictionary_sheets_are_named_but_not_imported_as_positions() -> None:
    result = parse_hungarian_workbook(_infra_workbook(_infra_rows()))
    assert result is not None
    assert result.metadata["item_sheet"] == "CSV kódolás"
    assert "V03016TKOD kódszótár" in result.metadata["dictionary_sheets"]
    assert len(result.positions) == len(_infra_rows())


@pytest.mark.parametrize("currency_source", ["building", "infrastructure"])
def test_both_profiles_declare_the_forint(currency_source: str) -> None:
    content = (
        _building_workbook(_rows_like_the_template())
        if currency_source == "building"
        else _infra_workbook(_infra_rows())
    )
    result = parse_hungarian_workbook(content)
    assert result is not None
    assert result.currency == "HUF"
